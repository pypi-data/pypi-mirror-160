import logging
logger = logging.getLogger(__name__)
import numpy as np
import openpyxl
import xlrd
import os.path
import re
import string
from dataUncert.variable import variable


def readData(xlFile, dataRange, uncertRange=None):
    logger.info(f'Creating a data object from the file {xlFile} with the dataRange {dataRange} and the uncertRange {uncertRange}')
    dat = _readData(xlFile, dataRange, uncertRange)
    return dat.dat


class _readData():

    def __init__(self, xlFile, dataRange, uncertRange=None) -> None:

        if not '-' in dataRange:
            logger.error('The data range has to include a hyphen (-)')
            raise ValueError('The data range has to include a hyphen (-)')
        index = dataRange.find('-')
        dataStartCol = dataRange[0:index]
        dataEndCol = dataRange[index + 1:]

        if '-' in dataStartCol or '-' in dataEndCol:
            logger.error('The data range can only include a singly hyphen (-)')
            raise ValueError('The data range can only include a singly hyphen (-)')

        if not uncertRange is None:
            if not '-' in uncertRange:
                logger.error('The uncertanty range has to include a hyphen (-)')
                raise ValueError('The uncertanty range has to include a hyphen (-)')
            index = uncertRange.find('-')
            uncertStartCol = uncertRange[0:index]
            uncertEndCol = uncertRange[index + 1:]

            if '-' in uncertStartCol or '-' in uncertEndCol:
                logger.error('The data range can only include a singly hyphen (-)')
                raise ValueError('The data range can only include a singly hyphen (-)')
        else:
            uncertStartCol = None
            uncertEndCol = None

        # convert the coloumns
        self.dataStartCol = self.colToIndex(dataStartCol)
        self.dataEndCol = self.colToIndex(dataEndCol)
        self.uncertStartCol = self.colToIndex(uncertStartCol)
        self.uncertEndCol = self.colToIndex(uncertEndCol)

        # check the uncertanty range
        uncertCols = [self.uncertStartCol is None, self.uncertEndCol is None]
        if sum(uncertCols) not in [0, 2]:
            logger.error('You have provided one of the coloumn for the uncertanty but not the other')
            raise ValueError('You have provided one of the coloumn for the uncertanty but not the other')

        # check the number of coloumns
        nColsData = self.dataEndCol - self.dataStartCol + 1
        if not self.uncertStartCol is None:
            nColsUncert = self.uncertEndCol - self.uncertStartCol + 1
            if nColsData != nColsUncert:
                logger.error('The number of coloumns of the data is not equal to the number of coloumns for the uncertanty')
                raise ValueError('The number of coloumns of the data is not equal to the number of coloumns for the uncertanty')
        self.nCols = nColsData

        # check the extension
        extension = os.path.splitext(xlFile)[1]
        supportedExtensions = ['.xls', '.xlsx']
        if extension not in supportedExtensions:
            logger.error(f'The file extension is not supported. The supported extension are {supportedExtensions}')
            raise ValueError(f'The file extension is not supported. The supported extension are {supportedExtensions}')

        # parse functions for the specific extension and get all sheets
        if extension == '.xls':
            self.wb = xlrd.open_workbook(xlFile)
            self.sheets = self.wb.sheets()

            def readCell(sheet, row, col):
                return sheet.cell(row, col).value

            def readRow(sheet, row):
                return [elem.value for elem in sheet.row(row)]

            def readCol(sheet, col):
                return [elem.value for elem in sheet.col(col)]

        elif extension == '.xlsx':
            self.wb = openpyxl.load_workbook(xlFile)
            self.sheets = [self.wb[elem] for elem in self.wb.sheetnames]

            def readCell(sheet, row, col):
                return sheet.cell(row + 1, col + 1).value

            def readRow(sheet, row):
                return [elem.value for elem in list(sheet.iter_rows())[row]]

            def readCol(sheet, col):
                return [elem.value for elem in list(sheet.iter_cols())[col]]

        self.readCell = readCell
        self.readRow = readRow
        self.readCol = readCol

        # read the data
        self.readData()

    def colToIndex(self, col):
        if col is None:
            return None
        if not isinstance(col, str):
            logger.error('The coloumn has to be a string')
            raise ValueError('The coloumn has to be a string')
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def formatHeaders(self, header):
        out = []
        for head in header:

            # remove symbols and replace with _
            head = re.sub(r'[^\w]', '_', head.lower())

            # determine places with repeated "_"
            indexes_to_remove = []
            for i in range(len(head) - 1):
                if head[i] == "_":
                    if head[i + 1] == "_":
                        indexes_to_remove.append(i)

            # remove the indexes found in the previous step
            head = [char for char in head]
            for i in reversed(indexes_to_remove):
                head.pop(i)
            head = "".join(head)

            # add "_" to the begining of the name if the first letter is a digit
            if head[0].isnumeric():
                head = '_' + head

            # remove "_" if the last letter is "_"
            if head[-1] == "_" and len(head) != 1:
                head = head[0:-1]

            i, imax, done = 0, 100, False
            while not done and i <= imax:
                if i > 0:
                    h = head + f'_{i+2}'
                else:
                    h = head
                if h not in out:
                    out.append(h)
                    done = True

        return out

    def formatUnits(self, units):
        out = []
        for unit in units:
            if unit is None:
                unit = ''
            if len(unit) > 0:
                # remove symbols and replace with _
                allowedCharacters = []
                allowedCharacters += list(string.ascii_letters)
                allowedCharacters += [str(num) for num in range(10)]
                allowedCharacters += ['/', '-']
                unit = list(unit)
                for i, char in enumerate(unit):
                    if char not in allowedCharacters:
                        unit.pop(i)
                unit = ''.join(unit)

                # remove "_" if the last letter is "_"
                if unit[-1] == "_" and len(unit) != 1:
                    unit = unit[0:-1]

                # remove "_" if the first letter is "_"
                if unit[0] == "_":
                    unit = unit[1:]

            out.append(unit)
        return out

    def readData(self):
        self.dat = _Data()

        # Looping over the sheets in the data file
        for i, sheet in enumerate(self.sheets):
            sheetData = _Sheet(f's{i+1}')

            # determine the number of variables
            headers = self.readRow(sheet, 0)[0:self.nCols]
            headers = self.formatHeaders(headers)
            units = self.readRow(sheet, 1)[0:self.nCols]
            units = self.formatUnits(units)

            # determine the number of datapoints
            nDataPoints = []
            for i in range(self.nCols):
                nDataPoint = self.readCol(sheet, i)[2:]
                nDataPoint = sum([1 if elem not in ['', None] else 0 for elem in nDataPoint])
                nDataPoints.append(nDataPoint)
            if not all(elem == nDataPoints[0] for elem in nDataPoints):
                logger.error('There are not an equal amount of rows in the data')
                raise ValueError('There are not an equal amount of rows in the data')
            nDataPoint = nDataPoints[0]

            # read the data
            data = np.zeros([nDataPoint, self.nCols])
            for i in range(nDataPoint):
                for j in range(self.nCols):
                    data[i, j] = float(self.readCell(sheet, 2 + i, j))

            if not self.uncertStartCol is None:
                # determine the number of rows in the uncertanty
                nUncertanties = []
                for i in range(self.nCols):
                    nUncertanty = self.readCol(sheet, self.nCols + i)[2:]
                    nUncertanty = sum([1 if elem not in ['', None] else 0 for elem in nUncertanty])
                    nUncertanties.append(nUncertanty)
                if not all(elem == nUncertanties[0] for elem in nUncertanties):
                    logger.error('There are not an equal amount of rows in the uncertanty')
                    raise ValueError('There are not an equal amount of rows in the uncertanty')
                nUncertanty = nUncertanties[0]

                # evaluate the number of rows of the uncertanty
                if nUncertanty not in [nDataPoint, nDataPoint * self.nCols]:
                    logger.error('The number of rows in the uncertanty has to be equal to the number of rows of data or equal to the number of rows of data multiplied with the number of coloumns in the data')
                    raise ValueError('The number of rows in the uncertanty has to be equal to the number of rows of data or equal to the number of rows of data multiplied with the number of coloumns in the data')

                if nUncertanty == nDataPoint:
                    # There is one row of uncertanty for each row of data. Therefore there are no covariance data in the sheet

                    # read the uncertanty
                    uncert = np.zeros([nDataPoint, self.nCols])
                    for i in range(nDataPoint):
                        for j in range(self.nCols):
                            uncert[i, j] = float(self.readCell(sheet, 2 + i, self.nCols + j))

                    # create the measurements uncertanties
                    for i in range(self.nCols):
                        name = headers[i]
                        unit = units[i]
                        val = np.array(data[:, i])
                        u = np.array(uncert[:, i])
                        var = variable(val, unit, uncert=u)

                        sheetData._addMeasurement(name, var)
                else:
                    # There are covariance data in the sheet

                    # read the uncertanty
                    uncert = []
                    for i in range(nDataPoint):
                        u = np.zeros([self.nCols, self.nCols])
                        for j in range(self.nCols):
                            for k in range(self.nCols):
                                u[j, k] = float(self.readCell(sheet, 2 + i * self.nCols + j, self.nCols + k))
                        uncert.append(u)

                    # check if each element in the uncertanty is symmetric
                    for elem in uncert:
                        if (elem.shape == elem.transpose().shape) and (elem == elem.transpose()).all():
                            pass
                        else:
                            logger.error('The covariances has to be symmetric')
                            raise ValueError('The covariances has to be symmetric')

                    # create the measurements with covariance uncertanties
                    vars = []
                    for i in range(self.nCols):
                        name = headers[i]
                        unit = units[i]
                        val = np.array(data[:, i])
                        u = np.array([elem[i, i] for elem in uncert])
                        var = variable(val, unit, uncert=u)
                        vars.append(var)

                    for i in range(self.nCols):
                        covariance = [elem[:, i] for elem in uncert]
                        for j in range(self.nCols):
                            if i != j:
                                cov = [elem[j] for elem in covariance]
                                vars[i]._addCovariance(vars[j], cov)

                    for head, var in zip(headers, vars):
                        sheetData._addMeasurement(head, var)
            else:
                # There are no uncertaty data in the sheet

                # create the measurements without uncertanties
                for i in range(self.nCols):
                    name = headers[i]
                    unit = units[i]
                    val = np.array(data[:, i])
                    var = variable(val, unit)
                    sheetData._addMeasurement(name, var)

            self.dat._addSheet(sheetData.name, sheetData)


class _Data():
    def __init__(self, name=''):
        self.name = name
        self.sheets = []

    def _addSheet(self, name, sheet):
        sheet.name = name
        sheetNames = [elem.name for elem in self.sheets]
        if name in sheetNames:
            index = sheetNames.index(name)
            self.sheets[index] = sheet
            logger.warning(f'A sheet with the name {name} already existed in the object {self}. The first sheet with the same name is overwritten.')
            raise Warning(f'A sheet with the name {name} already existed in the object {self}. The first sheet with the same name is overwritten.')
        else:
            self.sheets.append(sheet)
        setattr(self, name, sheet)

    def printContents(self):
        for sheet in self.sheets:
            sheet.printContents(self.name)
            print('')

    def __iter__(self):
        return iter(self.sheets)


class _Sheet():
    def __init__(self, name=''):
        self.name = name
        self.measurements = []
        self.measurementNames = []

    def _addMeasurement(self, name, var):
        self.measurements.append(var)
        self.measurementNames.append(name)
        setattr(self, name, var)

    def printContents(self, suffix=None):

        for name in self.measurementNames:
            if suffix is None:
                print(f'{self.name}.{name}')
            else:
                print(f'{suffix}.{self.name}.{name}')

    def __getitem__(self, index):

        logger.info(f'Indexing the sheet {self} with the indexes {index}')
        measurements = []
        for meas in self.measurements:
            val = meas.value[index]
            unit = meas.unit
            uncert = meas.uncert[index]
            measurements.append(variable(val, unit, uncert))

        sheet = _Sheet(self.name)

        for measurement, measurementName in zip(measurements, self.measurementNames):
            sheet._addMeasurement(measurementName, measurement)

        return sheet

    def append(self, other):

        if not isinstance(other, _Sheet):
            logger.error('You can only append two sheets together')
            raise ValueError('You can only append two sheets together')

        # Test if all names are the same
        for elem in self.measurementNames:
            if elem not in other.measurementNames:
                logger.error('You can only append sheets with the excact same measurements. The names did not match')
                raise ValueError('You can only append sheets with the excact same measurements. The names did not match')

        # get the measurements in the same order
        measA = self.measurements
        measB = []
        for elem in self.measurementNames:
            index = other.measurementNames.index(elem)
            measB.append(other.measurements[index])

        # test if all units are the same
        for elemA, elemB in zip(measA, measB):
            if str(elemA.unit) != str(elemB.unit):
                logger.error('You can only append sheets with the excact same measurements. The units did not match')
                raise ValueError('You can only append sheets with the excact same measurements. The units did not match')

        # append the data
        for i in range(len(self.measurements)):
            meas_i = measB[i]
            self.measurements[i].value = np.append(self.measurements[i].value, meas_i.value)
            self.measurements[i].uncert = np.append(self.measurements[i].uncert, meas_i.uncert)
            if len(self.measurements[i].covariance) == 0:
                if len(meas_i.covariance) == 0:
                    # do nothing
                    pass
                else:
                    # create covarinace for self.measurements[i] and fill it with zeros
                    nData = int(len(self.measurements[i].value) / 2)
                    for j in range(len(self.measurements)):
                        if i != j:
                            self.measurements[i].covariance[self.measurements[j]] = [0] * nData

                    # append the covariance from meas_i
                    keys = list(self.measurements[i].covariance.keys())
                    keys_i = list(meas_i.covariance.keys())
                    for key_i, name in zip(keys_i, other.measurementNames):
                        index = self.measurementNames.index(name)
                        self.measurements[i].covariance[keys[index]] = np.append(self.measurements[i].covariance[keys[index]], meas_i.covariance[key_i])
            else:
                if len(meas_i.covariance) == 0:
                    # append zeros to the covariance of self.measurements[i]
                    nData = int(len(self.measurements[i].value) / 2)
                    for key in self.measurements[i].covariance.keys():
                        self.measurements[i].covariance[key] = np.append(self.measurements[i].covariance[key], [0] * nData)
                else:
                    # append the covariance of meas_i to the covariance of self.measurements[i]
                    keys = list(self.measurements[i].covariance.keys())
                    keys_i = list(meas_i.covariance.keys())
                    for key_i, name in zip(keys_i, other.measurementNames):
                        index = self.measurementNames.index(name)
                        self.measurements[i].covariance[keys[index]] = np.append(self.measurements[i].covariance[keys[index]], meas_i.covariance[key_i])

    def __iter__(self):
        return iter(self.measurements)
