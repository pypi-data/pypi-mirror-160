from datetime import datetime

from openpyxl import Workbook

from easytradesdk import Serializer


class BackTestResult:

    def __init__(self, startDate, endDate, backTestOrders, positions):
        self.startDate = startDate
        self.endDate = endDate
        self.backTestOrders = backTestOrders
        self.positions = positions
        self.profits = []

    def calculateProfits(self):
        for _position in self.positions.values():
            _tc = _position.tc
            _symbol = _position.symbol
            _currentAmount = _position.remainAmount + _position.remainHolding * _position.lastTicker
            _profit = Profit(_tc, _symbol, _position.initialTotalAmount, _currentAmount)
            self.profits.append(_profit)

    def printProfits(self):
        for _profit in self.profits:
            _s = _profit.symbol.split("_")[1]
            print(
                "{}, {} profitAmount:{} {}, profitRate:{} %, cost:{} {}, remain:{} {} \n".format(
                    _profit.tc, _profit.symbol, _profit.profitAmount, _s, _profit.profitRate,
                    _profit.initialTotalAmount, _s, _profit.remainTotalAmount, _s)
            )

    def exportOrders(self, fileName=None):

        if not fileName:
            fileName = 'backTestOrder-' + str(datetime.now().timestamp()) + '.xlsx'

        if self.backTestOrders:
            excelTitle = ['序号', '订单id', '交易所', '交易标的', '交易类型', '交易方向', '成交价', '成交量', '成交额', '回测下单时间', '回测下单时间戳', '订单信号参数', '订单数据', '交易前仓位快照', '交易后仓位快照', '交易后仓位总价值']
            wb = Workbook()
            ws = wb.worksheets[0]
            ws.title = "回测订单"
            for idx in range(len(excelTitle)):
                ws.cell(1, idx + 1).value = excelTitle[idx]

            for idx in range(len(self.backTestOrders)):
                backTestOrder = self.backTestOrders[idx]
                ws.cell(idx + 2, 1).value = backTestOrder.id
                ws.cell(idx + 2, 2).value = backTestOrder.clientOrderId
                ws.cell(idx + 2, 3).value = backTestOrder.tc
                ws.cell(idx + 2, 4).value = backTestOrder.symbol
                ws.cell(idx + 2, 5).value = backTestOrder.type
                ws.cell(idx + 2, 6).value = backTestOrder.side
                ws.cell(idx + 2, 7).value = backTestOrder.price
                ws.cell(idx + 2, 8).value = backTestOrder.dealQty
                ws.cell(idx + 2, 9).value = backTestOrder.actualDealAmount
                ws.cell(idx + 2, 10).value = backTestOrder.time.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(idx + 2, 11).value = str(backTestOrder.timeMills)
                ws.cell(idx + 2, 12).value = Serializer.objectToJson(backTestOrder.orderSignal)
                ws.cell(idx + 2, 13).value = Serializer.objectToJson(backTestOrder.orderData)
                ws.cell(idx + 2, 14).value = Serializer.objectToJson(backTestOrder.posBeforeDeal)
                ws.cell(idx + 2, 15).value = Serializer.objectToJson(backTestOrder.posAfterDeal)
                ws.cell(idx + 2, 16).value = backTestOrder.posTotalAmountAfterDeal

            wb.save(fileName)
            wb.close()


class Profit:

    def __init__(self, tc, symbol, initialTotalAmount, remainTotalAmount):
        self.tc = tc
        self.symbol = symbol
        self.initialTotalAmount = initialTotalAmount
        self.remainTotalAmount = remainTotalAmount
        self.profitAmount = self.remainTotalAmount - self.initialTotalAmount
        self.profitRate = 0 if self.initialTotalAmount == 0 else self.profitAmount / self.initialTotalAmount * 100
