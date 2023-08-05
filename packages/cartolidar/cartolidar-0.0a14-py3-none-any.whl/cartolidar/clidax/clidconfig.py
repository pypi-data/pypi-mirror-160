#!/usr/bin/python
# -*- coding: UTF-8 -*-
'''
Created on 7/12/2018

@author: JB
# -*- coding: latin-1 -*-
'''
from __future__ import division, print_function

import os
import sys
import pathlib
import time
# import datetime
import types # Ver https://docs.python.org/2/library/types.html
# import csv
import re
import math
# import random
import platform
import inspect
# import traceback
# import subprocess
# import argparse
from configparser import RawConfigParser
import logging
# import importlib
# import struct
import shutil
# import gc
import unicodedata
import xml.etree.ElementTree as ET

# Paquetes de terceros
from dbfread import DBF
try:
    import psutil
    psutilOk = True
except:
    psutilOk = False

# ==============================================================================
if __name__ == '__main__':
    print('\nclidconfig-> ATENCION: este modulo no se puede ejecutar de forma autonoma')
    sys.exit(0)
# ==============================================================================
if '--cargadoClidconfig' in sys.argv:
    moduloPreviamenteCargado = True
    # print(f'\nclidconfig->1> moduloPreviamenteCargado: {moduloPreviamenteCargado}; sys.argv: {sys.argv}')
else:
    moduloPreviamenteCargado = False
    # print(f'\nclidconfig->1> moduloPreviamenteCargado: {moduloPreviamenteCargado}; sys.argv: {sys.argv}')
    sys.argv.append('--cargadoClidconfig')
# ==============================================================================
if '--idProceso' in sys.argv and len(sys.argv) > sys.argv.index('--idProceso') + 1:
    ARGS_idProceso = sys.argv[sys.argv.index('--idProceso') + 1]
else:
    # ARGS_idProceso = str(random.randint(1, 999998))
    ARGS_idProceso = '999999'
    sys.argv.append('--idProceso')
    sys.argv.append(ARGS_idProceso)
# ==============================================================================
if type(ARGS_idProceso) == int:
    MAIN_idProceso = ARGS_idProceso
elif type(ARGS_idProceso) == str:
    try:
        MAIN_idProceso = int(ARGS_idProceso)
    except:
        MAIN_idProceso = 0
        print(f'clidconfig-> ATENCION: revisar asignacion de idProceso.')
        print(f'ARGS_idProceso: {type(ARGS_idProceso)} {ARGS_idProceso}')
        print(f'sys.argv: {sys.argv}')
else:
    MAIN_idProceso = 0
    print(f'clidconfig-> ATENCION: revisar codigo de idProceso.')
    print(f'ARGS_idProceso: {type(ARGS_idProceso)} {ARGS_idProceso}')
    print(f'sys.argv: {sys.argv}')
# ==============================================================================
# Verbose provisional para la version alpha
if '-vvv' in sys.argv:
    __verbose__ = 3
elif '-vv' in sys.argv:
    __verbose__ = 2
elif '-v' in sys.argv or '--verbose' in sys.argv:
    __verbose__ = 1
else:
    # En eclipse se adopta el valor indicado en Run Configurations -> Arguments
    __verbose__ = 0
# ==============================================================================
if '-q' in sys.argv:
    __quiet__ = 1
    __verbose__ = 0
else:
    __quiet__ = 0
# ==============================================================================

# ==============================================================================
# ============================ Variables GLOBALES ==============================
# ==============================================================================
# TB = '\t'
TB = ' ' * 13
TV = ' ' * 3
TW = ' ' * 2
# ==============================================================================

# ==============================================================================
# ============================== Variables MAIN ================================
# ==============================================================================
# Directorio que depende del entorno:
MAIN_HOME_DIR = str(pathlib.Path.home())
# Directorios de la aplicacion:
MAIN_FILE_DIR = os.path.dirname(os.path.abspath(__file__))
# Cuando estoy en un modulo principal (clidbase.py o clidflow.py):
# MAIN_PROJ_DIR = MAIN_FILE_DIR
# Cuando estoy en un modulo dentro de un paquete (subdirectorio):
MAIN_PROJ_DIR = os.path.abspath(os.path.join(MAIN_FILE_DIR, '../..'))
MAIN_RAIZ_DIR = os.path.abspath(os.path.join(MAIN_PROJ_DIR, '..'))
# Directorio desde el que se lanza la app (estos dos coinciden):
# MAIN_BASE_DIR = os.path.abspath('.')
MAIN_BASE_DIR = os.path.abspath(os.path.dirname(sys.argv[0]))
MAIN_THIS_DIR = os.getcwd()
# ==============================================================================
# Unidad de disco si MAIN_ENTORNO = 'windows'
MAIN_DRIVE = os.path.splitdrive(MAIN_FILE_DIR)[0]  # 'D:' o 'C:'
# ==============================================================================
if MAIN_FILE_DIR[:12] == '/LUSTRE/HOME':
    MAIN_ENTORNO = 'calendula'
    MAIN_PC = 'calendula'
elif MAIN_FILE_DIR[:8] == '/content':
    MAIN_ENTORNO = 'colab'
    MAIN_PC = 'colab'
else:
    MAIN_ENTORNO = 'windows'
    try:
        if MAIN_DRIVE[0] == 'D':
            MAIN_PC = 'Casa'
        else:
            MAIN_PC = 'JCyL'
    except:
        MAIN_ENTORNO = 'calendula'
        MAIN_PC = 'calendula'
# ==============================================================================

# ==============================================================================
# Ver https://peps.python.org/pep-0008/#module-level-dunder-names
# Ver https://stackoverflow.com/questions/458550/standard-way-to-embed-version-into-python-package
# ==============================================================================
VERSIONFILE = os.path.abspath(os.path.join(MAIN_FILE_DIR, '_version.py'))
if not os.path.exists(VERSIONFILE):
    VERSIONFILE = os.path.abspath(os.path.join(MAIN_FILE_DIR, '..', '_version.py'))
    if not os.path.exists(VERSIONFILE):
        VERSIONFILE = os.path.abspath(os.path.join(MAIN_FILE_DIR, '../..', '_version.py'))
        if not os.path.exists(VERSIONFILE):
            VERSIONFILE = os.path.abspath('_version.py')
if os.path.exists(VERSIONFILE):
    try:
        verstrline = open(VERSIONFILE, "rt").read()
        VSRE = r"^__version__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            # __version__ = mo.groups()[0]
            __version__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __version__ = "a.b.c"')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __version__ = "a.b.c"\n')
            __version__ = '0.0a0'
        VSRE = r"^__date__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            __date__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __date__ = "year1-year2"')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __date__ = "year1-year2"\n')
            __date__ = '2016-2022'
        VSRE = r"^__updated__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            __updated__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __updated__ = "date"')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __updated__ = "date"\n')
            __updated__ = '2022-07-01'
        VSRE = r"^__copyright__ = ['\"]([^'\"]*)['\"]"
        mo = re.search(VSRE, verstrline, re.M)
        mo = re.search(VSRE, verstrline, re.M)
        if mo:
            __copyright__ = mo.group(1)
        else:
            # raise RuntimeError(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __copyright__ = "..."')
            sys.stderr.write(f'Revisar fichero {VERSIONFILE} -> Debe incluir la linea __copyright__ = "..."')
            __copyright__ = '@clid 2016-22'
    except:
        sys.stderr.write(f'clidtools.__init__-> no se ha podido leer {VERSIONFILE}\n')
        __version__ = '0.0a0'
        __date__ = '2016-2022'
        __updated__ = '2022-07-01'
        __copyright__ = '@clid 2016-22'
else:
    __version__ = '0.0a0'
    __date__ = '2016-2022'
    __updated__ = '2022-07-01'
    __copyright__ = '@clid 2016-22'
# ==============================================================================


# ==============================================================================
# Duplico esta funcion de clidaux para no importar clidaux
def infoUsuario(verbose=False):
    if psutilOk:
        try:
            esteUsuario = psutil.users()[0].name
            if verbose:
                print('clidconfig-> Usuario:', esteUsuario)
        except:
            esteUsuario = psutil.users()
            if verbose:
                print('clidconfig-> Users:', esteUsuario)
        if not isinstance(esteUsuario, str) or esteUsuario == '':
            esteUsuario = 'local'
    else:
        esteUsuario = 'SinUsuario'
    return esteUsuario


# ==============================================================================
# Duplico esta funcion de clidaux para no importar clidaux
def showCallingModules(inspect_stack=inspect.stack(), verbose=False):
    # print('->->->inspect_stack  ', inspect_stack
    # print('->->->inspect.stack()', inspect.stack())
    if len(inspect_stack) > 1:
        try:
            esteModuloFile0 = inspect_stack[0][1]
            esteModuloNum0 = inspect_stack[0][2]
            esteModuloFile1 = inspect_stack[1][1]
            esteModuloNum1 = inspect_stack[1][2]
            esteModuloName0 = inspect.getmodulename(esteModuloFile0)
            esteModuloName1 = inspect.getmodulename(esteModuloFile1)
        except:
            print(f'{TB}clidconfig-> Error identificando el modulo 1')
            return 'desconocido1', 'desconocido1'
    else:
        if verbose:
            print(f'{TB}clidconfig-> No hay modulos que identificar')
        return 'noHayModuloPrevio', 'esteModulo'

    if not esteModuloName0 is None:
        esteModuloName = esteModuloName0
        esteModuloNum = esteModuloNum0
        stackSiguiente = 1
    else:
        esteModuloName = esteModuloName1
        esteModuloNum = esteModuloNum1
        stackSiguiente = 2

    callingModulePrevio = ''
    callingModuleInicial = ''
    if verbose:
        print(f'{TB}clidconfig-> El modulo {esteModuloName} ({esteModuloNum}) ha sido', end=' ')
    for llamada in inspect_stack[stackSiguiente:]:
        if 'cartolid' in llamada[1] or 'clid' in llamada[1] or 'qlid' in llamada[1]:
            callingModule = inspect.getmodulename(llamada[1])
            if callingModule != esteModuloName and callingModulePrevio == '':
                callingModulePrevio = callingModule
            callingModuleInicial = callingModule
            if verbose:
                print(f'importado desde: {callingModule} ({llamada[2]})', end='; ')
    if verbose:
        print()
    return callingModulePrevio, callingModuleInicial


# ==============================================================================
def iniciaConsLog(myModule='clidconfig', myVerbose=False, myQuiet=False):
    if myVerbose == 3:
        logLevel = logging.DEBUG  # 10
    elif myVerbose == 2:
        logLevel = logging.DEBUG  # 10
    elif myVerbose == 1:
        logLevel = logging.INFO  # 20
    elif myVerbose == 0:
        if myQuiet:
            # logLevel = logging.CRITICAL
            logLevel = logging.ERROR  # 40
        else:
            # logLevel = logging.WARNING  # 30
            logLevel = logging.INFO  # 20
    # ==============================================================================
    # class ContextFilter(logging.Filter):
    #     """
    #     This is a filter which injects contextual information into the log.
    #     """
    #
    #     def filter(self, record):
    #         record.thisUser = myUser
    #         record.thisFile = myModule[:10]
    #         return True
    # myFilter = ContextFilter()
    # ==============================================================================
    # formatter1 = '{asctime}|{name:10s}|{levelname:8s}|{thisUser:8s}|> {message}'
    # formatterFile = logging.Formatter(formatter1, style='{', datefmt='%d-%m-%y %H:%M:%S')
    formatterCons = logging.Formatter('{message}', style='{')
    
    myLog = logging.getLogger(myModule)
    if sys.argv[0].endswith('__main__.py') and 'cartolidar' in sys.argv[0]:
        # qlidtwins.py se ejecuta lanzando el paquete cartolidar desde linea de comandos:
        #  python -m cartolidar
        # En __main__.py ya se ha confiigurado el logging.basicConfig()
        # if myModule == __name__.split('.')[-1]:
        #     print(f'{myModule}-> En __main.py se va a crear el loggin de consola para todos los modulos en __main__.py')
        # else:
        #     print(f'{myModule}-> Ya se ha creado el loggin de consola para todos los modulos en __main__.py')
        pass
    consLog = logging.StreamHandler()
    consLog.setFormatter(formatterCons)
    consLog.setLevel(logLevel)
    myLog.setLevel(logLevel)
    myLog.addHandler(consLog)
    return myLog


# ==============================================================================
def foo0():
    pass

# ==============================================================================
CONFIGverbose = __verbose__ > 2
if CONFIGverbose:
    print(f'\nclidconfig-> AVISO: CONFIGverbose True; __verbose__: {__verbose__}')
# ==============================================================================

# ==============================================================================
myUser = infoUsuario()
myModule = __name__.split('.')[-1]
# ==============================================================================
if not moduloPreviamenteCargado or True:
    # print('\nclidconfig-> AVISO: creando myLog (ConsLog)')
    myLog = iniciaConsLog(myModule=myModule, myVerbose=__verbose__)
# ==============================================================================
if CONFIGverbose:
    myLog.debug(f'{"":_^80}')
    myLog.debug(f'clidconfig-> Debug & alpha version info:')
    myLog.debug(f'{TB}-> ENTORNO:          {MAIN_ENTORNO}')
    myLog.debug(f'{TB}-> Modulo principal: <{sys.argv[0]}>') # = __file__
    myLog.debug(f'{TB}-> __package__ :     <{__package__ }>')
    myLog.debug(f'{TB}-> __name__:         <{__name__}>')
    myLog.debug(f'{TB}-> __verbose__:      <{__verbose__}>')
    myLog.debug(f'{TB}-> IdProceso         <{MAIN_idProceso:006}>')
    # myLog.debug(f'{TB}-> configFile:       <{GLO.configFileNameCfg}>')
    myLog.debug(f'{TB}-> sys.argv:         <{sys.argv}>')
    myLog.debug(f'{"":=^80}')
# ==============================================================================

# ==============================================================================
if CONFIGverbose:
    myLog.debug(f'\nclidconfig-> Cargando clidconfig...')
    myLog.debug(f'{TB}-> Directorio desde el que se lanza la aplicacion-> os.getcwd(): {os.getcwd()}')
    myLog.debug(f'{TB}-> Revisando la pila de llamadas...')
callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
if CONFIGverbose:
    myLog.debug(f'{TB}{TV}-> callingModulePrevio:  {callingModulePrevio}')
    myLog.debug(f'{TB}{TV}-> callingModuleInicial: {callingModuleInicial}')
# ==============================================================================


# ==============================================================================
def creaLog(consLogYaCreado=False, myModule='module', myUser=myUser, myPath='.', myVerbose=0, myVerboseFile=0, myQuiet=False):
    myVerboseCons = myVerbose

    if myVerboseFile == 3:
        logLevelFile = logging.DEBUG
    elif myVerboseFile == 2:
        logLevelFile = logging.DEBUG
    elif myVerboseFile == 1:
        logLevelFile = logging.INFO
    elif myVerboseFile == 0:
        if myQuiet:
            # logLevelFile = logging.CRITICAL
            logLevelFile = logging.ERROR
        else:
            # logLevelFile = logging.WARNING
            logLevelFile = logging.INFO

    if myVerbose == 3:
        logLevelCons = logging.DEBUG  # 10
    elif myVerbose == 2:
        logLevelCons = logging.DEBUG  # 10
    elif myVerbose == 1:
        logLevelCons = logging.INFO  # 20
    elif myVerbose == 0:
        if myQuiet:
            # logLevelCons = logging.CRITICAL
            logLevelCons = logging.ERROR  # 40
        else:
            # logLevelCons = logging.WARNING  # 30
            logLevelCons = logging.INFO  # 20

    if myPath == '':
        myLogPath = myPath
    else:
        myLogPath = os.path.abspath(myPath)
        if not os.path.isdir(myLogPath):
            try:
                os.makedirs(myLogPath)
            except:
                print(f'No se ha podido crear el directorio {myLogPath}. Revisar derechos de escritura en esa ruta.')
                sys.exit(0)
    mainLogFile = os.path.join(myLogPath, f'clidbase.log')
    thisLogFile = os.path.join(myLogPath, f'{myModule}.log')
    # print(f'clidconfig-> mainLogFile: {mainLogFile}')
    # print(f'clidconfig-> thisLogFile: {thisLogFile}')
    try:
        if not os.path.exists(mainLogFile):
            controlConfigFile = open(mainLogFile, mode='w')
            controlConfigFile.close()
            os.remove(mainLogFile)
        else:
            controlConfigFile = open(mainLogFile, mode='r+')
            controlConfigFile.close()
    except:
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
        print(f'\n{"":_^80}')
        print(f'clidconfig-> AVISO:')
        print(f'{TB}-> No se puede guardar el fichero log:  {mainLogFile}')
        print(f'{TB}-> Es posible que no tenga permisos de escritura en: {myLogPath}')
        print(f'{TB}-> O que exista el fichero {mainLogFile} y este bloqueado.')
        rutaAlternativa = True
        try:
            MAIN_HOME_DIR = str(pathlib.Path.home())
            MAIN_LOG_DIR = os.path.join(MAIN_HOME_DIR, 'Documents')
            print(f'\tSe intenta la ruta alternativa: {MAIN_LOG_DIR}')
            mainLogFile = os.path.join(MAIN_LOG_DIR, f'clidbase.log')
            thisLogFile = os.path.join(MAIN_LOG_DIR, f'{myModule}.log')
            if not os.path.exists(mainLogFile):
                controlConfigFile = open(mainLogFile, mode='w')
                controlConfigFile.close()
                os.remove(mainLogFile)
            else:
                controlConfigFile = open(mainLogFile, mode='r+')
                controlConfigFile.close()
            print(f'\tOk log file1: {mainLogFile}')
            print(f'\tOk log file2: {thisLogFile}')
            myLogPath = MAIN_LOG_DIR
        except:
            print(f'\tTampoco se puede escribir en la ruta {MAIN_LOG_DIR}')
            MAIN_HOME_DIR = str(pathlib.Path.home())
            mainLogFile = os.path.join(MAIN_HOME_DIR, f'clidbase.log')
            thisLogFile = os.path.join(MAIN_HOME_DIR, f'{myModule}.log')
            print(f'\tSe opta por log file1: {mainLogFile}')
            print(f'\tlog file2: {thisLogFile}')
            myLogPath = MAIN_HOME_DIR
        print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')
        print(f'{"":=^80}')

    # return (myLogPath, mainLogFile, thisLogFile)


    # ==============================================================================
    class ContextFilter(logging.Filter):
        """
        This is a filter which injects contextual information into the log.
        """

        def filter(self, record):
            record.thisUser = myUser
            record.thisFile = os.path.join(myLogPath, myModule[:10])
            return True
    # ==============================================================================
    myFilter = ContextFilter()
    # ==============================================================================
    # formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    # formatter = logging.Formatter('%(name)-12s: %(message)s')
    # formatter = logging.Formatter('{name:16s}: {levelname:8s} {message}', style='{')
    # formatter3 = '%(asctime)s|%(process)d|%(thisFile)-10s|%(levelname)-8s|%(thisUser)-8s|>%(message)s' # tyle='%'
    
    formatter0 = '{message}'
    formatter1 = '{asctime}|{name:10s}|{levelname:8s}|> {message}'
    formatter2 = '{asctime}|{name:10s}|{levelname:8s}|{thisUser:8s}|> {message}'
    formatterFile = logging.Formatter(formatter2, style='{', datefmt='%d-%m-%y %H:%M:%S')
    formatterCons = logging.Formatter(formatter0, style='{')

    if not consLogYaCreado:
        # print(f'clidconfig-> {myModule}-> Iniciando logging.basicConfig<> para fileLog & consLog')
        # Este logger se mantiene activo para todos los modulos,
        # pero no puedo utilizar myFilter con todos los modulos porque algunas librerias (matplotlib)
        # lanzan mensajes de debug y no tienen definido el filtro, con lo que da error
        # https://docs.python.org/3/library/logging.html#logging.basicConfig
        logging.basicConfig(
            filename=mainLogFile,
            filemode='w',
            format=formatter1,
            style='{',
            datefmt='%d-%m-%y %H:%M:%S',
            # datefmt='%d-%b-%y %H:%M:%S',
            level=logLevelFile,
            # level=logging.DEBUG,
            # level=logging.INFO,
            # level=logging.WARNING,
            # level=logging.ERROR,
            # level=logging.CRITICAL,
        )
    else:
        # qlidtwins.py se ejecuta lanzando el paquete cartolidar desde linea de comandos:
        #  python -m cartolidar
        # En __main__.py ya se ha confiigurado el logging.basicConfig()
        # print(f'clidconfig-> {myModule}-> Ya se ha creado el loggin de consola para todos los modulos en __main__.py')
        pass

    myLog = logging.getLogger(myModule)
    # myLog.setLevel(logging.DEBUG)
    myLog.addFilter(myFilter)
    
    # Este logger solo actua para este modulo:
    # https://docs.python.org/3/library/logging.handlers.html#filehandler
    fileLog = logging.FileHandler(thisLogFile, mode='w')
    # fileLog.terminator = ''
    fileLog.set_name(myModule)
    fileLog.setLevel(logLevelFile)
    fileLog.setFormatter(formatterFile)
    fileLog.addFilter(myFilter)
    # logging.getLogger().addHandler(fileLog)
    myLog.addHandler(fileLog)
    if not consLogYaCreado:
        # https://docs.python.org/3/library/logging.handlers.html#logging.StreamHandler
        consLog = logging.StreamHandler()
        # https://docs.python.org/3/library/logging.handlers.html#logging.StreamHandler.terminator
        # consLog.terminator = ''  # Sustituye al valor por defecto que es '\n'
        consLog.setFormatter(formatterCons)
        consLog.setLevel(logLevelCons)
        # logging.getLogger().addHandler(consLog)
        myLog.addHandler(consLog)
    # ==============================================================================
    # myLog.debug('clidconfig-> debug')
    # myLog.info('clidconfig-> info')
    # myLog.warning('clidconfig-> warning')
    # myLog.error('clidconfig-> error')
    # myLog.critical('clidconfig-> critical')
    # ==============================================================================

    return myLog


# ==============================================================================
def leerTablaDBF(elArchivo, GLOBAL_seleccionadosParaRepetir='', GLOBAL_listaRepetir=[]):
    # miTabla = dbf_old.Table(elArchivo)
    # miTabla = dbf.Table(elArchivo)
    miTabla = DBF(elArchivo)
    print('\nListado de campos del fichero dbf:')
    for miCampo in miTabla.fields:
        print(miCampo.name, miCampo.type, miCampo.length, miCampo.decimal_count)
    usarPropiedad_meta = False  # No funciona
    if usarPropiedad_meta:
        print('\nListado de campos (con _meta):')
        for campo in miTabla._meta.keys():
            if miTabla._meta[campo][0] == 'N':
                decimales = '; dec: ' + str(miTabla._meta[campo][4])
            else:
                decimales = ''
            print(campo.ljust(10), 'Tipo: ' + miTabla._meta[campo][0] + '; ancho:', str(miTabla._meta[campo][2]) + decimales)
    usarPropiedad_field_layout = False  # No funciona
    if usarPropiedad_field_layout:
        print('\nListado de campos (con _field_layout):')
        for i in range(len(miTabla.field_names)):
            print('miTabla._field_layout({:1}):'.format(i), miTabla._field_layout(i))

    print('Nombre del fichero:', elArchivo)
    print('Datos generales de', elArchivo + ':', miTabla)
    # Este es el string method de la "class Table(_Navigation):"= print( miTabla )
    print('Datos individuales de', elArchivo + ':')
    print('Nombre:              ', miTabla.filename)
    # print( 'Version:             ', miTabla.version ) # No en dfbread
    print('Version:             ', miTabla.dbversion, '=', miTabla.header.dbversion)  # Especifica de dbfread
    # print( 'Ult actualizacion:   ', miTabla.last_update ) # No en dfbread
    print('Ult actualizacion:   ', miTabla.date, '=', miTabla.header.year, miTabla.header.month, miTabla.header.day)  # Especifica de dbfread
    # print( 'codepage:            ', miTabla.codepage ) # No en dfbread
    print('codepage:            ', miTabla.encoding)  # Especifica de dbfread
    # Para que el texto con acentos en el dbf no de error establecer el encoding a 'cp1252' o 'latin-1':
    if miTabla.encoding == 'ascii':
        miTabla.encoding = 'cp1252'
    # print( 'Num de campos:       ', miTabla.field_count, '=', len(miTabla.field_names) ) # No en dfbread
    print('Num de campos:       ', len(miTabla.field_names))
    # print( 'Num de registros:    ', miTabla.__len__(), '=', len(miTabla) ) # No en dfbread
    print('Num de registros:    ', miTabla.header.numrecords, '(', len(miTabla.records), ')')
    # print( 'Ancho de cabecera:   ', miTabla.record_length ) # No en dfbread
    print('Ancho de registro:   ', miTabla.header.recordlen)  # Especifica de dbfread
    print('Ancho de cabecera:   ', miTabla.header.headerlen)  # Especifica de dbfread
    print('Campos (field_names):', miTabla.field_names)  # -> List

    # print( 'Registros:     ', type(miTabla.records))
    nContador = 0
    for registro in miTabla.records:
        nContador += 1
        # print( 'Registros - tipo:', type(registro), dir(registro))
        # print( 'Registros - claves:', registro.keys())
        # print( 'Registro', nContador, '->', registro )
        print('Registro', nContador, '-> Fichero lidar:', registro['Lidar'])
        if nContador > 5:
            print('clidconfig-> Etc.')
            break

    # miTabla.open() # No en dfbread
    # print('clidconfig-> miTabla', type(miTabla), dir(miTabla))
    # Metodos de miTabla: 'add_fields', 'allow_nulls', 'append', 'backup', 'bof', 'bottom', 'close', 'codepage', 'create_backup', 'create_index', 'create_template', 'current', 'current_record', 'delete_fields', 'disallow_nulls', 'eof', 'field_count', 'field_info', 'field_names', 'filename', 'first_record', 'goto', 'index', 'last_record', 'last_update', 'memoname', 'new', 'next_record', 'nullable_field', 'open', 'pack', 'prev_record', 'query', 'record_length', 'reindex', 'rename_field', 'resize_field', 'skip', 'status', 'structure', 'supported_tables', 'top', 'version', 'zap']
    for miRegistr in miTabla:
        # print('clidconfig-> miRegistr', type(miRegistr), dir(miRegistr))
        # Metodos de miRegistr <class 'dbf.Record'> ['__class__', '__contains__', '__delattr__', '__dir__', '__doc__', '__enter__', '__eq__', '__exit__', '__format__', '__ge__', '__getattr__', '__getattribute__', '__getitem__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__len__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__setitem__', '__sizeof__', '__slots__', '__str__', '__subclasshook__', '__weakref__', '_commit_flux', '_create_blank_data', '_data', '_dirty', '_memos', '_meta', '_old_data', '_recnum', '_reindex_record', '_retrieve_field_value', '_rollback_flux', '_start_flux', '_update_disk', '_update_field_value', '_write', '_write_to_disk'
        # repetir = miRegistr._retrieve_field_value(0,'repetir')
        # repetir = miRegistr._retrieve_field_value('repetir') # No en dfbread
        repetir = miRegistr['Repetir']
        # nombreLas = miRegistr._retrieve_field_value(0,'Lidar')
        # nombreLas = miRegistr._retrieve_field_value('Lidar') # No en dfbread
        nombreLas = miRegistr['Lidar']
        if repetir >= GLOBAL_seleccionadosParaRepetir:
            GLOBAL_listaRepetir.append(nombreLas)
            # print( nombreLas, str(repetir) )
    return len(miTabla), GLOBAL_listaRepetir


# # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# def quitarContrabarrasAgregarBarraFinal(ruta=''):
#     if not ruta:
#         return None
#     nuevaRuta = ''
#     for letra in ruta:
#         letraSinBackslash = letra if letra != '\\' else '/'
#         nuevaRuta += letraSinBackslash
#     if nuevaRuta[-1:] != '/':
#         nuevaRuta += '/'
#     return nuevaRuta


# ooooooooooo Variable gobal precalculada: GLBLminimoDeMemoriaRAM ooooooooooooooo
def asignarMinimoDeMemoriaRAM(GLBLminimoDeMemoriaRAM):
    if platform.processor()[:3] == 'x86':
        minimoDeMemoriaRAM_recomendada = 0
    else:
        minimoDeMemoriaRAM_recomendada = GLBLminimoDeMemoriaRAM
    if TRNSpreguntarMemoria:
        print('Valor minimo de RAM para interrumpir lectura MB)?')
        print('Sugerencia: Para PC de 64 bits: 100 MB; para PC de 32 bits, 0 MB')
        selec = input(f'{TB}clidconfig-> Pulsa un numero o [enter] directamente para valor por defecto ({minimoDeMemoriaRAM_recomendada} MB) -> ')
        try:
            minimoDeMemoriaRAM = int(selec)
        except:
            minimoDeMemoriaRAM = minimoDeMemoriaRAM_recomendada
        print('    -> La lectura se interrumpe cuando la RAM disponible baja de %i MB' % minimoDeMemoriaRAM)
    else:
        minimoDeMemoriaRAM = minimoDeMemoriaRAM_recomendada
    return minimoDeMemoriaRAM


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def mostrarVariablesGlobales(momento='inicial'):
    print('clidconfig-> Fichero actual (file):', __file__)
    print('\nVariables globales MAIN, TRNS y USER del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if variableGlobal[0:4] == 'MAIN' or variableGlobal[0:4] == 'TRNS' or variableGlobal[0:4] == 'USER':
            print('clidconfig-> {:>50}\t{}'.format(variableGlobal, globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nVariables globales GLBL del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if variableGlobal[0:4] == 'GLBL':
            print('clidconfig-> {:>50}\t{}'.format(variableGlobal, globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nVariables globales no GLBL, MAIN, TRNS o USER del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if (
            variableGlobal[0:3] != 'GLB'
            and variableGlobal[0:4] != 'MAIN'
            and variableGlobal[0:4] != 'TRNS'
            and variableGlobal[0:4] != 'USER'
            and variableGlobal[0:2] != '__'
            and not isinstance(globals()[variableGlobal], type)
        ):
            print('clidconfig-> {:>30}\t{}\t{}'.format(variableGlobal, type(globals()[variableGlobal]), globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nVariables globales __nativas__ del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if variableGlobal[0:2] == '__':
            print('clidconfig-> {:>15}\t{}\t{}'.format(variableGlobal, type(globals()[variableGlobal]), globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print('\nModulos y clases del modulo %s:' % os.path.basename(__file__))
    for variableGlobal in globals().keys():
        if type(globals()[variableGlobal]) == types.ModuleType or isinstance(globals()[variableGlobal], type):
            print('clidconfig-> {:>12}\t{}\t{}'.format(variableGlobal, type(globals()[variableGlobal]), globals()[variableGlobal]))
            # print(f'{TB} {variableGlobal} {globals()[variableGlobal]}')
    print(
        '\nVariables globales almacenadas en el momento {} en el array GLOBALconfigDict del modulo {}:'.format(
            momento, os.path.basename(__file__)
        )
    )
    for variableGlobalAlmacenada in globals()['GLOBALconfigDict']:
        print('clidconfig-> {:>50}\t{}'.format(variableGlobalAlmacenada, globals()['GLOBALconfigDict'][variableGlobalAlmacenada]))
        # print(f'{TB} {variableGlobalAlmacenada} {globals()["GLOBALconfigDict"][variableGlobalAlmacenada]}')


# ==============================================================================
# Una vez importado este modulo as CG puedo usar:
#    from cartolidar.clidax.clidconfig import GLOBALconfigDict
#    variable global -> GLOBALconfigDict['miVariable']
# o bien:
#    from cartolidar.clidax import clidconfig as CG
#    variable global -> CG.configVarsDict['miVariable']
#    #o bien:
#    variable global -> CG.GC.miVariable
#    #o bien:
#    VG = CG.VariablesGlobales()
#    variable global -> VG.miVariable
# o bien:
#    from cartolidar.clidax.clidconfig import GLO
#    variable global -> GLO.miVariable

# Opto por esta ultima
# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# Para controlFileLas y controlFileGlobal accedo directamente con
#    GL.controlFileLas
#    GL.controlFileGlobal
# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def initConfigDicts(idProceso=MAIN_idProceso, LOCL_verbose=0):
    global GLOBALconfigDict

    # ==========================================================================
    # Lectura del config (cfg) generado en un carga anterior de clidconfig o en una ejecucion anterior
    configFileNameCfg = getConfigFileNameCfg(idProceso, LOCL_verbose=LOCL_verbose)

    if CONFIGverbose:
        print('clidconfig-> Ejecutando initConfigDicts<> para leer el fichero de configuracion cfg')
        print(f'{TB}-> Verifico si hay un cfg con idProceso={idProceso} generado')
        print(f'{TB}-> en un carga anterior de clidconfig o en una ejecucion anterior:')
        print(f'{TB}{TV}{configFileNameCfg}')

    if os.path.exists(configFileNameCfg):
        if CONFIGverbose:
            print(f'{TB}clidconfig-> Ok, se leen los parametros del cfg: {os.path.basename(configFileNameCfg)}')
        config = RawConfigParser()
        config.optionxform = str  # Avoid change to lowercase
        numObjetivosExtraMax = 0
        LOCALconfigDict = {}
        try:
            config.read(configFileNameCfg)
            for grupoParametroConfiguracion in config.sections():
                for nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion):
                    strParametroConfiguracion = config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion)
                    listaParametroConfiguracion = strParametroConfiguracion.split('|+|')
                    valorParametroConfiguracion = valorConfig(
                        listaParametroConfiguracion[0],
                        nombreParametro=nombreParametroDeConfiguracion,
                        tipoVariable=listaParametroConfiguracion[1]
                    )
                    if len(listaParametroConfiguracion) > 1:
                        tipoParametroConfiguracion = listaParametroConfiguracion[1]
                    else:
                        tipoParametroConfiguracion = 'str'
                    if len(listaParametroConfiguracion) > 2:
                        descripcionParametroConfiguracion = listaParametroConfiguracion[2]
                    else:
                        descripcionParametroConfiguracion = ''
                    if nombreParametroDeConfiguracion[:1] == '_':
                        grupoParametroConfiguracion_new = '_%s' % grupoParametroConfiguracion
                    else:
                        grupoParametroConfiguracion_new = grupoParametroConfiguracion
                    LOCALconfigDict[nombreParametroDeConfiguracion] = [
                        valorParametroConfiguracion,
                        grupoParametroConfiguracion_new,
                        descripcionParametroConfiguracion,
                        tipoParametroConfiguracion,
                    ]
                    numObjetivosExtra = max(len(listaParametroConfiguracion) - 3, 0)
                    numObjetivosExtraMax = max(numObjetivosExtra, numObjetivosExtraMax)
                    if numObjetivosExtraMax:
                        valObjetivosExtra = []
                        for nObjetivoExtra in range(numObjetivosExtraMax):
                            if 3 + nObjetivoExtra < len(listaParametroConfiguracion):
                                valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[3 + nObjetivoExtra], tipoVariable=listaParametroConfiguracion[1]))
                            else:
                                valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[0], tipoVariable=listaParametroConfiguracion[1]))
                        LOCALconfigDict[nombreParametroDeConfiguracion].extend(valObjetivosExtra)
    
            LOCALconfigDict['configFileNameCfg'] = [
                configFileNameCfg,
                'GrupoMAIN',
                'Nombre del fichero de configuracion (este fichero).',
                'str',
            ]
            if CONFIGverbose:
                print(f'\tclidconfig-> Parametros leidos ok del fichero cfg: {LOCALconfigDict["configFileNameCfg"]}')
    
            configLeidoDelCfgOk = True
            GLOBALconfigDict = LOCALconfigDict
        except:
            configLeidoDelCfgOk = False
    else:
        configLeidoDelCfgOk = False

    if configLeidoDelCfgOk:
        return GLOBALconfigDict
    # ==========================================================================

    # ==========================================================================
    # Si no hay cfg generado en un carga anterior de clidconfig o en una ejecucion anterior
    # se busca el fichero de configuracion original (xlsx).
    (
        filenameXlsCfg1Clidbase,
        filenameXlsCwdClidbase,
        filenameXlsCwdModulo,
        filenameXlsCfg2Clidbase,
        filenameXlsProjClidbase,
    ) = getConfigFileNameXls(configFileNameCfg, LOCL_verbose=LOCL_verbose)

    directorioActual = (os.getcwd()).replace(os.sep, '/')
    directorioProyecto = os.path.dirname(sys.argv[0])
    moduloInicialXls = (os.path.basename(sys.argv[0])).replace('.py', '.xlsx')

    if CONFIGverbose:
        print('clidconfig-> Ejecutando initConfigDicts<> para leer el fichero de configuracion xlsx')
        print(f'{TB}-> Fichero cfg no disponible, se leen los parametros del xls:')
        print(f'{TB}{TV}-> {filenameXlsCfg1Clidbase}')
        print(f'{TB}-> Aqui leo las columnas correspondientes a todos los objetivoEjecucion.')
        print(f'{TB}-> Mas adelante, segun el valor de MAINobjetivoEjecucion usare uno u otro valor.')

    usarXLS = True
    if os.path.exists(filenameXlsCfg1Clidbase):
        if (directorioActual.replace('/','')).endswith('cartolidar') and moduloInicialXls == 'clidbase.xlsx':
            filenameXLS = filenameXlsCfg1Clidbase
            if CONFIGverbose:
                print('clidconfig-> Se usa un fichero de configuracion principal: {}'.format(filenameXlsCfg1Clidbase))
        else:
            filenameXLS = filenameXlsCwdModulo
            if CONFIGverbose:
                print(
                    'clidconfig-> Se usa un fichero de configuracion especifico de modulo ({}) o directorio ({}):'.format(
                        moduloInicialXls,
                        directorioActual,
                    )
                )
                print(f'{TB}-> {filenameXlsCwdModulo}')
    elif os.path.exists(filenameXlsCwdClidbase):
        if (directorioActual.replace('/','')).endswith('cartolidar'):
            filenameXLS = filenameXlsCwdClidbase
            if CONFIGverbose:
                print(f'clidconfig-> Se usa un fichero de configuracion principal del directorio desde el que se lanza la aplicacion: {filenameXlsCwdClidbase}')
        else:
            filenameXLS = filenameXlsCwdModulo
            if CONFIGverbose:
                print(f'clidconfig-> Se usa un fichero de configuracion para el modulo inicial en el directorio desde el que lanza la aplicacion: ({directorioActual})')
    elif os.path.exists(filenameXlsCfg2Clidbase):
        filenameXLS = filenameXlsCfg2Clidbase
        if CONFIGverbose:
            print(f'clidconfig-> Se usa un fichero de configuracion general ubicado en el directorio elegido finalmente para el cfg: {filenameXlsCfg2Clidbase}')
    elif os.path.exists(filenameXlsProjClidbase):
        filenameXLS = filenameXlsProjClidbase
        if CONFIGverbose:
            print(f'clidconfig-> Se usa un fichero de configuracion general ubicado en el directorio de la aplicacion cartolidar: {filenameXlsProjClidbase}')
    else:
        if 'cartolidar' in directorioActual:
            # Por si ejecuto directamente un modulo que no esta en cartolid/ 
            # sino en cartolid/package/ y no tengo un xml propio para ese package.
            if not (directorioActual.replace('/','')).endswith('cartolidar'):
                directorioActual = os.path.abspath(os.path.join(directorioActual, '..'))
        else:
            # Por si llamo a cartolid desde su directorio raiz (en calendula)
            directorioActual = os.path.join(directorioActual, 'cartolidar')
        filenameXLSotroDirCartolid = os.path.join(directorioActual, 'clidbase.xlsx')
        # filenameXML = os.path.join(directorioActual, 'clidbase.xml')
        if os.path.exists(filenameXLSotroDirCartolid):
            print(f'clidconfig-> Se usa un fichero de configuracion general del directorio: {directorioActual}')
            filenameXLS = filenameXLSotroDirCartolid
            print(f'{TB}-> Fichero de configuracion: {filenameXLS}')
        else:
            print(f'clidconfig-> ATENCION no se ha encontrando un fichero de configuracion. Se han intentado (por este orden):')
            print(f'{TB}-> {filenameXlsCfg1Clidbase}')
            print(f'{TB}-> {filenameXlsCwdClidbase}')
            print(f'{TB}-> {filenameXlsCwdModulo}')
            print(f'{TB}-> {filenameXlsCfg2Clidbase}')
            print(f'{TB}-> {filenameXlsProjClidbase}')
            sys.exit(0)
    # https://openpyxl.readthedocs.io/en/stable/
    # https://realpython.com/openpyxl-excel-spreadsheets-python/
    # https://www.pythonexcel.com/openpyxl.php
    if usarXLS:
        try:
            if os.path.exists(filenameXLS):
                from openpyxl import load_workbook
                usarXLS = True
            else:
                print(f'{TB}clidconfig-> ATENCION: no se encuentra {filenameXLS}')
                usarXLS = False
        except:
            print('clidconfig-> ATENCION: error al importar openpyxl')
            usarXLS = False

    if not usarXLS:
        print(f'clidconfig-> Instalar openpyxl para que clidbase pueda leer: {moduloInicialXls}')
        sys.exit(0)

    try:
        intIdProceso = int(idProceso)
    except:
        intIdProceso = 999999
    try:
        if len(sys.argv) == 0 or sys.argv[0] == '' or sys.argv[0] == '-m':
            # print('Se esta ejecutando fuera de un modulo, en el interprete interactivo')
            filenameXLSactual = 'clidbase.xlsx'
        else:
            # print('Se esta ejecutando desde un modulo')
            if intIdProceso:
                filenameXLSactual = sys.argv[0].replace('.py', '{:006}.xlsx'.format(intIdProceso))
                filenameXLSactual = sys.argv[0].replace('.py', '{:006}.xlsx'.format(intIdProceso))
            else:
                filenameXLSactual = sys.argv[0].replace('.py', '.xlsx')
        if CONFIGverbose:
            print(f'{TB}clidconfig-> Hago una copia del fichero de configuracion xlsx: {filenameXLSactual}')
        shutil.copy(filenameXLS, filenameXLSactual)
    except:
        print('clidconfig-> No se ha podido copiar {} a {} -> revisar.'.format(
            filenameXLS,
            filenameXLSactual,
            )
        )

    myWorkbookOrigin = load_workbook(filename=filenameXLS, read_only=True, data_only=True)
    # myWorkbookActual = load_workbook(filename=filenameXLSactual)
    hojas = myWorkbookOrigin.sheetnames
    nombreHojaCartolidConfig = 'cartolidConfig'
    if nombreHojaCartolidConfig in hojas:
        #hojaActiva = myWorkbookOrigin.active
        hojaCartolidConfigOrigin = myWorkbookOrigin[nombreHojaCartolidConfig]
        # hojaCartolidConfigActual = myWorkbookActual[nombreHojaCartolidConfig]
        verbose = False
        if verbose:
            print('clidconfig-> Mostrando variables globales del fichero de configuracion xlsx:')
            for nRow in range(1, 1000):
                if (
                    hojaCartolidConfigOrigin.cell(row=nRow, column=3).value is None
                    or str(hojaCartolidConfigOrigin.cell(row=nRow, column=3).value).startswith('#')
                    or str(hojaCartolidConfigOrigin.cell(row=nRow, column=3).value).startswith('_')
                ):
                    continue
                print('{:>50}: {:<30} ({})'.format(
                    str(hojaCartolidConfigOrigin.cell(row=nRow, column=3).value),
                    str(hojaCartolidConfigOrigin.cell(row=nRow, column=4).value),
                    str(type(hojaCartolidConfigOrigin.cell(row=nRow, column=4).value)),
                    )
                )
            print()

        configTextDict = {}
        # La funcion iter_rows() considera la columna A como 1, pero dentro de la lista generada (tupleRow), esa columna tiene el index 0.
        for nRow, tupleRow in enumerate(hojaCartolidConfigOrigin.iter_rows(min_col= 1, max_col=25, values_only=True)):
            try:
                nombreParametroDeConfiguracion = tupleRow[2]
                if (
                    nombreParametroDeConfiguracion is None or 
                    nombreParametroDeConfiguracion == 'name'
                ):
                    continue
                if (
                    not nombreParametroDeConfiguracion.startswith('MAIN')
                    and not nombreParametroDeConfiguracion.startswith('GLBL')
                    and not nombreParametroDeConfiguracion.startswith('#')
                ):
                    if verbose:
                        print(
                            'clidconfig-> ATENCION: se ignoran parametros que no empiecen por MAIN o GLBL: {}'.format(
                                nombreParametroDeConfiguracion
                            )
                        )
                    continue
                tipoVariable = tupleRow[5] # Columna F
                valorPrincipal = tupleRow[3] # Columna D
                valorAlternativo = valorPrincipal
                valorCREA_TILES = tupleRow[8] # Columna I
                if valorCREA_TILES is None or valorCREA_TILES == '':
                    valorCREA_TILES = valorPrincipal
                valorPREPROCESADO_EN_CALENDULA = tupleRow[9] # Columna J
                if valorPREPROCESADO_EN_CALENDULA is None or valorPREPROCESADO_EN_CALENDULA == '':
                    valorPREPROCESADO_EN_CALENDULA = valorPrincipal
                valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = tupleRow[10] # Columna K
                if valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ is None or valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ == '':
                    valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = valorPrincipal
                valorCREAR_LAZ = tupleRow[11] # Columna L
                if valorCREAR_LAZ is None or valorCREAR_LAZ == '':
                    valorCREAR_LAZ = valorPrincipal
                valorEXTRA2 = tupleRow[12] # Columna M
                if valorEXTRA2 is None or valorEXTRA2 == '':
                    valorEXTRA2 = valorPrincipal
                valorEXTRA3 = tupleRow[13] # Columna N
                if valorEXTRA3 is None or valorEXTRA3 == '':
                    valorEXTRA3 = valorPrincipal

                grupoParametros = tupleRow[1] # Columna B
                descripcionParametro = tupleRow[6]
                usoParametro = tupleRow[4]
                configTextDict[nombreParametroDeConfiguracion] = [
                    tipoVariable,
                    valorPrincipal,
                    valorAlternativo,
                    grupoParametros,
                    descripcionParametro,
                    usoParametro,
                    valorCREA_TILES,
                    valorPREPROCESADO_EN_CALENDULA,
                    valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ,
                    valorCREAR_LAZ,
                    valorEXTRA2,
                    valorEXTRA3,
                ]
                # hojaCartolidConfigActual.cell(row=nRow+1, column=1).value = nRow
            except:
                print('clidconfig-> Ha habido un problema al leer la hoja', nombreHojaCartolidConfig, 'del fichero de configuracion:', filenameXLS)
                print(f'{TB}-> Fila: {nRow} Columnas leidas (tupleRow): {type(tupleRow)} {tupleRow}')
        # myWorkbookActual.save(filename=filenameXLSactual)
    else:
        usarXLS = False
        print('clidconfig-> ATENCION: Revisar contenido del fichero de configuracion', filenameXLS)
        print(f'{TB}No tiene la hoja "cartolidConfig"')
        sys.exit(0)

    # Creo otro dict, pero:
    #    -> Solo con los valores de los parametros de configuracion
    #    -> Guardados en su formato especifico (str, int, float o bool)
    #    -> Eligiendo el valor principal o alternativo
    GLOBALconfigDict = {}
    if CONFIGverbose and __verbose__ > 3 and not moduloPreviamenteCargado:
        print('clidconfig-> Mostrando parametros de configuracion del xls:')
    for nombreParametroDeConfiguracion in configTextDict.keys():
        # configTextDict[nombreParametroDeConfiguracion] -> [tipoVariable, valorPrincipal, valorAlternativo, grupoParametros, descripcionParametro, usoParametro]
        tipoVariable = configTextDict[nombreParametroDeConfiguracion][0]
        valorPrincipal = configTextDict[nombreParametroDeConfiguracion][1]
        valorAlternativo = configTextDict[nombreParametroDeConfiguracion][2] # El valor alternativo lo he abandonada al pasar de xml a xlsx
        grupoParametros = configTextDict[nombreParametroDeConfiguracion][3]
        descripcionParametro = configTextDict[nombreParametroDeConfiguracion][4]
        usoParametro = configTextDict[nombreParametroDeConfiguracion][5]
        valorCREA_TILES = configTextDict[nombreParametroDeConfiguracion][6]
        valorPREPROCESADO_EN_CALENDULA = configTextDict[nombreParametroDeConfiguracion][7]
        valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = configTextDict[nombreParametroDeConfiguracion][8]
        valorCREAR_LAZ = configTextDict[nombreParametroDeConfiguracion][9]
        valorEXTRA2 = configTextDict[nombreParametroDeConfiguracion][10]
        valorEXTRA3 = configTextDict[nombreParametroDeConfiguracion][11]

        valorParametroDeConfiguracionPrincipal = valorConfig(
            valorPrincipal,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionCREA_TILES = valorConfig(
            valorCREA_TILES,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionPREPROCESADO_EN_CALENDULA = valorConfig(
            valorPREPROCESADO_EN_CALENDULA,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ = valorConfig(
            valorCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionCREAR_LAZ = valorConfig(
            valorCREAR_LAZ,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionEXTRA2 = valorConfig(
            valorEXTRA2,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )
        valorParametroDeConfiguracionEXTRA3 = valorConfig(
            valorEXTRA3,
            valorAlternativo,
            False,
            nombreParametro=nombreParametroDeConfiguracion,
            tipoVariable=tipoVariable,
        )

        GLOBALconfigDict[nombreParametroDeConfiguracion] = [
            valorParametroDeConfiguracionPrincipal,
            grupoParametros,
            descripcionParametro,
            tipoVariable,
            valorParametroDeConfiguracionCREA_TILES,
            valorParametroDeConfiguracionPREPROCESADO_EN_CALENDULA,
            valorParametroDeConfiguracionCREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ,
            valorParametroDeConfiguracionCREAR_LAZ,
            valorParametroDeConfiguracionEXTRA2,
            valorParametroDeConfiguracionEXTRA3,
        ]
        if CONFIGverbose and __verbose__ > 3 and not moduloPreviamenteCargado:
            print(f'{TB}{nombreParametroDeConfiguracion}-> {GLOBALconfigDict[nombreParametroDeConfiguracion]}')


    return GLOBALconfigDict


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
class VariablesGlobales(object):
    def __init__(self, LOCALconfigDict={}, LCLverbose=0):
        """
        Clase cuyas propiedades son las variables de configuracion
        que utilizo en otros modulos como variables globales
        Utiliza un dict (LOCALconfigDict) con los nombres y valores de esas variables
        En definitiva, esta clase sirve para convertir valores de un dict
        en propiedades de un objeto
        Ademas crea algunas variables adicionales que se crean en tiempo de ejecucion:
            -> Variables derivadas
            -> Nombres de los ficheros de control
        #TODO: Lo siguiente seria guardar todas esas variables en un fichero de configuracion
        """

        # sys.argv tiene como primer valor el nombre del modulo que se ejecuta inicialmente sys.argv[0]
        # Ademas he agregado dos valores adicionales (van despues de los argumentos en linea de comandos):
        #  sys.argv.append('--idProceso')  -> sys.argv[-2]
        #  sys.argv.append(MAIN_idProceso) -> sys.argv[-1]

        # if len(sys.argv) > 3:
        #     self.ARGSobjetivoEjecucion = sys.argv[1]
        # else:
        #     self.ARGSobjetivoEjecucion = ''

        self.LCLverbose = LCLverbose
        if CONFIGverbose:
            print(f'{TB}clidconfig-> VariablesGlobales<>-> Creando la clase VariablesGlobales')
        self.configVarsDict = LOCALconfigDict

        if CONFIGverbose:
            print(f'{TB}clidconfig-> VariablesGlobales<>-> Cargando los paramConfig como propiedades de esta clase (GLO va a ser un objeto de esta clase)')
            # print(f'{TB}{TV}-> self.configVarsDict: {self.configVarsDict}')
        self.cargaVariablesDelConfigVarsDict()

        # Casos espaciales (objetivos de ejecucion no reglados):
        #  Ese objetivo de ejecucion puede llegar en linea de comandos o en fichero de configuracion xls.
        #  En cargaVariablesDelConfigVarsDict<>, el MAINobjetivoEjecucion se ha revertido a un valor reglado
        #  y el valor no reglado se guarda en self.MAINobjetivoNoReglado
        # En la siguiente funcion se retoca, adicionalmente, lo que haga falta.

        # print(f'clidconfig.VariablesGlobales-> callingModuleInicial: <{callingModuleInicial}>')
        if (
            hasattr(self, 'MAINprocedimiento')
            and callingModuleInicial != 'runpy'
            and callingModuleInicial != '__init__'
            and callingModuleInicial != '__main__'
            and not callingModuleInicial.startswith('test_')
            and callingModuleInicial != 'clidtwins' and callingModuleInicial != 'qlidtwins'
            and callingModuleInicial != 'clidmerge' and callingModuleInicial != 'qlidmerge'
        ):
            if CONFIGverbose:
                print(f'{TB}clidconfig-> Adaptando algunas variables a casos especiales de objetivoEjecucion no reglado.')
            self.revisarSiHayObjetivoEjecucionEspecial()

        if CONFIGverbose:
            print(f'{TB}clidconfig-> #3a LOCALconfigDict["GLBLverbose"]:     {LOCALconfigDict["GLBLverbose"]}')
            print(f'{TB}clidconfig-> #3b self.configVarsDict["GLBLverbose"]: {self.configVarsDict["GLBLverbose"]}')


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def cargaVariablesDelConfigVarsDict(self):
        # Solo se muestra el contenido de configVarsDict la primera vez que
        # se importa clidconfig.py (normalmente desde clidbase.py) y no cuando
        # se importa desde el resto de modulos (controlado con self.LCLverbose).

        # ======================================================================
        # Por si acaso...
        if not 'GLBLmostrarVariablesDeConfiguracion' in self.configVarsDict.keys():
            self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'] = [self.configVarsDict['GLBLverbose'][0], 'Main', 'str', 'Herencias de cartolidar']
        # ======================================================================

        # ======================================================================
        # ========================= objetivoEjecucion ==========================
        # ======================================================================
        if 'MAINobjetivoEjecucion' in self.configVarsDict.keys():
            if (
                callingModuleInicial != 'runpy'
                and callingModuleInicial != '__init__'
                and callingModuleInicial != '__main__'
                and callingModuleInicial != 'clidtwins' and callingModuleInicial != 'qlidtwins'
                and callingModuleInicial != 'clidmerge' and callingModuleInicial != 'qlidmerge'
                and not callingModuleInicial.startswith('test_')
            ):
                if self.LCLverbose:
                    print(f'clidconfig-> Uso cargaVariablesDelConfigVarsDict<> para poner los parametros leidos del cfg')
                    print(f'{TB}como propiedades de esta clase (GLO va a ser un objeto de esta clase)')
                    print(f'{TB}Uso el parametro MAINobjetivoEjecucion ({self.configVarsDict["MAINobjetivoEjecucion"][0]})')
                    print(f'{TB}para decidir si uso el valor general o el especifico del objetivoEjecucion (tengo todos en el cfg).')
        else:
            self.configVarsDict['MAINobjetivoEjecucion'] = ['GENERAL', 'Main', 'str', 'Herencias de cartolidar']
        # ======================================================================
        if not 'ARGSobjetivoEjecucion' in self.configVarsDict.keys():
            if '-o' in sys.argv or '--objetivo' in sys.argv:
                self.configVarsDict['ARGSobjetivoEjecucion'] = self.configVarsDict['MAINobjetivoEjecucion']
            else:
                self.configVarsDict['ARGSobjetivoEjecucion'] = ['', 'str', '', 'GrupoMAIN', '']
        if not hasattr(self, 'ARGSobjetivoEjecucion'):
            self.ARGSobjetivoEjecucion = self.configVarsDict['ARGSobjetivoEjecucion'][0]
        if self.ARGSobjetivoEjecucion == '':
            self.MAINobjetivoEjecucion = self.configVarsDict['MAINobjetivoEjecucion'][0]
            if self.LCLverbose:
                print(f'\n{"":_^80}')
                print(f'clidconfig-> Uso cargaVariablesDelConfigVarsDict<> para incorporar el parametro objetivoEjecucion como propiedad de GLO:')
                print(f'{TB}clidconfig-> *MAINobjetivoEjecucion en fichero de configuracion (xlsx o cfg): {self.MAINobjetivoEjecucion}')
        else:
            self.MAINobjetivoEjecucion = self.ARGSobjetivoEjecucion
            if self.LCLverbose:
                print(f'\n{"":_^80}')
                print(f'clidconfig-> Uso cargaVariablesDelConfigVarsDict<> para incorporar el argument -o (objetivoEjecucion) leido en linea de comandos:')
                print(f'{TB}*ARGSobjetivoEjecucion en linea de comandos: {self.ARGSobjetivoEjecucion}')
        # ======================================================================

        # ======================================================================
        # Casos espaciales (objetivos de ejecucion no reglados):
        if self.MAINobjetivoEjecucion == 'CREAR_PUNTOS_TRAIN_ROQUEDOS':
            self.MAINobjetivoSiReglado = 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ'
            self.configVarsDict['MAINobjetivoSiReglado'] = [
                'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'GrupoMAIN', '', 'str', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
                'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
                'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ', 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ',
            ]
            self.MAINobjetivoNoReglado = 'CREAR_PUNTOS_TRAIN_ROQUEDOS'
            self.configVarsDict['MAINobjetivoNoReglado'] = [
                'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'GrupoMAIN', '', 'str', 'CREAR_PUNTOS_TRAIN_ROQUEDOS',
                'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS',
                'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS', 'CREAR_PUNTOS_TRAIN_ROQUEDOS',
            ]
        else:
            self.MAINobjetivoSiReglado = self.MAINobjetivoEjecucion
            self.configVarsDict['MAINobjetivoSiReglado'] = [
                self.MAINobjetivoEjecucion, 'GrupoMAIN', '', 'str', self.MAINobjetivoEjecucion,
                self.MAINobjetivoEjecucion, self.MAINobjetivoEjecucion, self.MAINobjetivoEjecucion, 
                self.MAINobjetivoEjecucion, self.MAINobjetivoEjecucion, self.MAINobjetivoEjecucion,
            ]
            self.MAINobjetivoNoReglado = 'NINGUNO'
            self.configVarsDict['MAINobjetivoNoReglado'] = [
                'NINGUNO', 'GrupoMAIN', '', 'str', 'NINGUNO',
                'NINGUNO', 'NINGUNO', 'NINGUNO',
                'NINGUNO', 'NINGUNO', 'NINGUNO',
            ]

        if self.LCLverbose:
            print(f'{TB}-> *Se elige la columna correspondiente al MAINobjetivoSiReglado: {self.configVarsDict["MAINobjetivoSiReglado"][0]}')
            print(f'{TB}-> *callingModulePrevio: {callingModulePrevio}; callingModuleInicial: {callingModuleInicial}')
            print(f'{"":=^80}')

        # Tengo objetivos de ejecucion preconfigurados, distinto del objetivo generico:
        if self.configVarsDict['MAINobjetivoSiReglado'][0] == 'GENERAL':
            # Usa la columna D del fichero de configuracion
            objetivoEjecucion = 0
        elif self.configVarsDict['MAINobjetivoSiReglado'][0] == 'CREAR_TILES_TRAIN':
            # Para crear tiles destinados al entrenamiento convolucional
            objetivoEjecucion = 4
        elif self.configVarsDict['MAINobjetivoSiReglado'][0] == 'PREPROCESADO_EN_CALENDULA':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            objetivoEjecucion = 5
        elif self.configVarsDict['MAINobjetivoSiReglado'][0] == 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            objetivoEjecucion = 6
        elif self.configVarsDict['MAINobjetivoSiReglado'][0] == 'CREAR_LAZ':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            objetivoEjecucion = 7
        elif self.configVarsDict['MAINobjetivoSiReglado'][0] == 'EXTRA2':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            objetivoEjecucion = 8
        elif self.configVarsDict['MAINobjetivoSiReglado'][0] == 'EXTRA3':
            # Para adelantar el procesado de ficheros laz, generando ficheros npz y tiles (tb genera ASC).
            objetivoEjecucion = 9
        else:
            # Objetivos de ejecucion general:
            objetivoEjecucion = 0

        if self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'][0]:
            if self.LCLverbose:
                print(f'\n{"":_^80}')
                print(f'clidconfig-> Se cargan como propiedades de GLO los parametros globales de configuracion, guardados en configVarsDict')
                print(f'{TB} {self.configVarsDict.keys()}')

        for nombreParametroDeConfiguracion in self.configVarsDict.keys():
            if self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'][0]:
                if self.LCLverbose:
                    print(
                        'clidconfig->-------------------------->',
                        nombreParametroDeConfiguracion,
                        type(nombreParametroDeConfiguracion),
                        self.configVarsDict[nombreParametroDeConfiguracion][objetivoEjecucion]
                    )
            # print('clidconfig-> Parametro', nombreParametroDeConfiguracion, '>> self.configVarsDict:', len(self.configVarsDict[nombreParametroDeConfiguracion]), self.configVarsDict[nombreParametroDeConfiguracion])
            if objetivoEjecucion < len(self.configVarsDict[nombreParametroDeConfiguracion]):
                setattr(self, nombreParametroDeConfiguracion, self.configVarsDict[nombreParametroDeConfiguracion][objetivoEjecucion])
            else:
                if self.LCLverbose and nombreParametroDeConfiguracion != 'configFileNameCfg':
                    print(f'{TB}-> El parametro {nombreParametroDeConfiguracion} no tiene valor especifico para el objetivo de ejecucion {objetivoEjecucion}')
                    print(f'{TB}-> Se adopta el valor correspondiente al objetivo general: {self.configVarsDict[nombreParametroDeConfiguracion][0]}') 
                setattr(self, nombreParametroDeConfiguracion, self.configVarsDict[nombreParametroDeConfiguracion][0])
        if self.configVarsDict['GLBLmostrarVariablesDeConfiguracion'][0]:
            if self.LCLverbose:
                print(f'{"":=^80}')


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def revisarSiHayObjetivoEjecucionEspecial(self):
        ARGScodCuadrante = self.configVarsDict['MAINcuadrante'][0]
        # ======================================================================
        if self.configVarsDict['MAINobjetivoEjecucion'][0] != 'CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ':
            callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
            if callingModulePrevio == 'clidbase':
                if self.LCLverbose:
                    print(f'\n{"":_^80}')
                    print(f'clidconfig-> Uso revisarSiHayObjetivoEjecucionEspecial<>:')
                    print(f'{TB}-> ATENCION: ESTO ES PROVISIONAL Y SOLO ACTUA CUANDO SE USAN ARGUMENTOS EN LINEA DE COMANDOS')
                    print(f'{TB}-> Esta pensado solo para CALENDULA, pero funciona tb en windows.')
                    print(f'{TB}-> Es para permitir la coexistencia de dos ejecuciones con un mismo clidbase.xls (en calendula).')
                    print(f'{TB}-> Se fuerza un MAINprocedimiento diferente en funcion de las mayus/minusc del cuadrante:')
                    print(f'{TB}{TV}-> Ejecucion completa de cuadrante SE/CE/NE y Se/Ce/Ne-> AUTOMATICO_EN_CALENDULA_SCRATCH')
                    print(f'{TB}{TV}-> Ejecucion de chequeo de cuadrante se/ce/ne y sE/cE/nE-> AUTOMATICO_EN_CALENDULA_SELECT')
                    print(f'{TB}-> Parametros de configuracion actuales:')
                    print(f'{TB}{TV}-> ARGScodCuadrante en linea de comandos?: {ARGScodCuadrante}')
                    print(f'{TB}{TV}-> self.MAINprocedimiento antes del retoque: {self.MAINprocedimiento}')
                    print(f'{TB}-> Esto no tiene efecto para la ejecucion destinada a generar puntos de entrenamiento')
                    print(f'{TB}{TV}-> MAINobjetivoEjecucion = CREAR_PUNTOS_TRAIN_ACUMULATIVO_NPZ')
                    print(f'{TB}{TV}-> Esa ejecucion la hago normalmente con Se/Ce/Ne -antes SE/CE/**- y AUTOMATICO_EN_CALENDULA_SELECT.')
            # print('clidconfig->> self.MAINprocedimiento:', type(self.MAINprocedimiento), self.MAINprocedimiento)
            if self.MAINprocedimiento.startswith('AUTOMATICO_EN_CALENDULA'):
                if ARGScodCuadrante == 'SE' or ARGScodCuadrante == 'CE' or ARGScodCuadrante == 'NE':
                    # Desde 10/2021 esto no conlleva seleccionar lasFiles, sino normalmente procesar todos los lasFiles (no solo _SELECT)
                    self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SCRATCH'
                    self.configVarsDict['MAINprocedimiento'] = ['AUTOMATICO_EN_CALENDULA_SCRATCH', 'GrupoMAIN', '', 'str', 'AUTOMATICO_EN_CALENDULA_SCRATCH']
                elif ARGScodCuadrante == 'Se' or ARGScodCuadrante == 'Ce' or ARGScodCuadrante == 'Ne':
                    self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SELECT'
                    self.configVarsDict['MAINprocedimiento'] = ['AUTOMATICO_EN_CALENDULA_SELECT', 'GrupoMAIN', '', 'str', 'AUTOMATICO_EN_CALENDULA_SELECT']
                elif (
                    ARGScodCuadrante == 'se' or ARGScodCuadrante == 'ce' or ARGScodCuadrante == 'ne'
                    or ARGScodCuadrante == 'sE' or ARGScodCuadrante == 'cE' or ARGScodCuadrante == 'nE'
                ):
                    self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SELECT'
                    self.configVarsDict['MAINprocedimiento'] = ['AUTOMATICO_EN_CALENDULA_SELECT', 'GrupoMAIN', '', 'str', 'AUTOMATICO_EN_CALENDULA_SELECT']
                else:
                    # No hay argumentos en linea de comandos -> la configuracion es la de clidbase.xls
                    pass
            if callingModulePrevio == 'clidbase':
                if self.LCLverbose:
                    print(f'{TB}-> self.MAINprocedimiento desp. del retoque: {self.MAINprocedimiento}')
                    print(f'{"":=^80}')
        # ======================================================================
        if self.configVarsDict['MAINobjetivoNoReglado'][0] != 'NINGUNO':
            if self.LCLverbose:
                print(f'\n{"":_^80}')
                print(f'clidconfig-> Revisando posible objetivo no reglado-: {self.configVarsDict["MAINobjetivoNoReglado"][0]}')

            if self.MAINobjetivoNoReglado == 'CREAR_PUNTOS_TRAIN_ROQUEDOS':
                if self.LCLverbose:
                    print(f'{TB}-> Se adapta la ejecucion a la recopilacion de puntos roquedo para entrenamiento.')
                self.MAINprocedimiento = 'AUTOMATICO_EN_CALENDULA_SELECT'
                # self.MAINcuadrante = 'AUTOMATICO_EN_CALENDULA_SELECT'
                # self.GLBLusarVectorNucleosUrbanosVector = True
                # self.GLBLusarVectorNucleosUrbanosRaster = True
                self.GLBLusarVectorGeologicoVector = True
                self.GLBLusarVectorGeologicoRaster = True
                self.GLBLsoloRoquedosParaEntrenamiento = True
                self.GLBLfraccionDeMuestreoGeneral = 1
                self.GLBLmuestreoEspecificoDeClase = False
                self.GLBLsobreMuestrearClasesSubRepresentadas = False
                self.GLBLsubMuestrearClasesSobreRepresentadas = False
                # Esto condiciona la lectura de una u otra columna de parece que no tiene efecto:
                self.configVarsDict['MAINprocedimiento'] = ['AUTOMATICO_EN_CALENDULA_SELECT', 'GrupoMAIN', '', 'str', 'AUTOMATICO_EN_CALENDULA_SELECT']
                self.configVarsDict['MAINcuadrante'][0] = 'sE'
                self.configVarsDict['GLBLusarVectorGeologicoVector'][0] = True
                self.configVarsDict['GLBLusarVectorGeologicoRaster'][0] = True
                self.configVarsDict['GLBLsoloRoquedosParaEntrenamiento'][0] = True
                self.configVarsDict['GLBLfraccionDeMuestreoGeneral'][0] = 1
                self.configVarsDict['GLBLmuestreoEspecificoDeClase'][0] = False
                self.configVarsDict['GLBLsobreMuestrearClasesSubRepresentadas'][0] = False
                self.configVarsDict['GLBLsubMuestrearClasesSobreRepresentadas'][0] = False
            else:
                if self.LCLverbose:
                    print(f'{TB}-> ATENCION: Objetivo no reglado no implementado; se cambia a "NINGUNO".')
                self.MAINobjetivoNoReglado = 'NINGUNO'
                self.GLBLsoloRoquedosParaEntrenamiento = False
                self.configVarsDict['MAINobjetivoNoReglado'] = ['NINGUNO', 'GrupoMAIN', '', 'str', 'NINGUNO']
                self.configVarsDict['GLBLsoloRoquedosParaEntrenamiento'][0] = False

            if self.LCLverbose:
                print(f'{"":=^80}')


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def revisarCompletarVariablesMAINdelConfigVarsDict(self):
        # if not 'MAINSoloProcesarPendientes' in self.configVarsDict.keys():
        #    self.configVarsDict['MAINSoloProcesarPendientes'] = [False, 'GrupoMAIN', '', 'bool']


        # ======================================================================
        # Adaptacion a/de MAINprocedimiento
        if (
            (self.configVarsDict['MAINprocedimiento'][0]).startswith('AUTOMATICO_EN_CALENDULA_SCRATCH')
            or (self.configVarsDict['MAINprocedimiento'][0]).startswith('AUTOMATICO_EN_CALENDULA_SELECT')
        ):
            # Si MAIN_ENTORNO es calendula, lasFiles (GLO.MAINrutaLaz) y resultados (GLO.MAINrutaOutput) en scratch.
            # En caso contrario, es se reconvierte a AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA 
            if (self.configVarsDict['MAINprocedimiento'][0]).endswith('_UNIFICAR_RGBI'):
                tareaEspecial = '_UNIFICAR_RGBI'
            elif (self.configVarsDict['MAINprocedimiento'][0]).endswith('_H29_COLOREAR_RGBI'):
                tareaEspecial = '_H29_COLOREAR_RGBI'
            elif (self.configVarsDict['MAINprocedimiento'][0]).endswith('_COLOREAR_RGBI'):
                tareaEspecial = '_COLOREAR_RGBI'
            else:
                tareaEspecial = ''
            if MAIN_ENTORNO != 'calendula':
                # Rutas por defecto
                if (self.configVarsDict['MAINprocedimiento'][0]).startswith('AUTOMATICO_EN_CALENDULA_SCRATCH'):
                    self.configVarsDict['MAINprocedimiento'][0] = 'AUTOMATICO_CON_RUTA_LAZ_PREDETERMINADA' + tareaEspecial
                elif (self.configVarsDict['MAINprocedimiento'][0]).startswith('AUTOMATICO_EN_CALENDULA_SELECT'):
                    self.configVarsDict['MAINprocedimiento'][0] = (self.configVarsDict['MAINprocedimiento'][0]).replace('_EN_CALENDULA', '_CON_RUTA_LAZ_PREDETERMINADA') + tareaEspecial
        # ======================================================================



        if not 'MAINbuscarPendientes' in self.configVarsDict.keys():
            self.configVarsDict['MAINbuscarPendientes'] = [False, 'GrupoMAIN', '', 'bool', False]
        MAINbuscarPendientes = self.configVarsDict['MAINbuscarPendientes'][0]
        MAINlistaRepetir = []
        if MAINbuscarPendientes:
            if self.configVarsDict['MAINusuario'][0] == 'JB':
                DBF_DIR = r'../data/dbf/'
                elArchivo = DBF_DIR + 'MallaLidar2x2km.dbf'
            elif self.configVarsDict['MAINusuario'][0] == 'benmarjo':
                DBF_DIR = r'O:/Sigmena/usuarios/COMUNES/Bengoa/Lidar/Lidas/'
                elArchivo = DBF_DIR + 'MallaLidar2x2km.dbf'
            else:
                MAINbuscarPendientes = False
        if MAINbuscarPendientes:
            # MAINlistaRepetir = ['216-4656', '262-4510', '264-4510', '266-4510', '266_4510', '268-4510','268_4510']
            # MAINlistaRepetir = ['262-4510', '264-4510', '266-4510', '268-4510']
            # nfilas = len(MAINlistaRepetir)
            nfilas, MAINlistaRepetir = leerTablaDBF(elArchivo, TRNSseleccionadosParaRepetir, MAINlistaRepetir)
            if MAINlistaRepetir != []:
                if self.LCLverbose:
                    print(
                        'Se van a procesar %i ficheros laz de un total de %i que hay que repetir (ver TRNSseleccionadosParaRepetir)' % len(MAINlistaRepetir, nfilas)
                    )
                    print(MAINlistaRepetir)

        self.configVarsDict['MAINbuscarPendientes'][0] = MAINbuscarPendientes
        self.configVarsDict['MAINlistaRepetir'] = [MAINlistaRepetir, 'GrupoMAIN', '', 'list', MAINlistaRepetir]


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def revisarCompletarVariablesGLBLdelConfigVarsDict(self):
        # Se recalculan a la vista del entorno de trabajo
        if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
            print(f'\n{"":_^80}')
            print(f'clidconfig-> Chequeando la configuracion...')

        # Si mi objetivo es CREAR_LAZ y mi modelo no usa las hiperformas, me ahorro ese paso.
        # Para saber si el modelo usa las hioerformas, me baso en:
        #   1. El numero de inputs del nombre del modelo: GLBLnombreFicheroConModeloParaInferencia
        #   2. En su defecto, en GLBLincluirVarDeHiperformasEnElModeloAcumulativo
        if self.configVarsDict['MAINobjetivoEjecucion'][0] == 'CREAR_LAZ':
            if (
                not self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][0] is None
                and self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][0] != ''
                and '_i' in self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][0]
            ):
                posicionNumInputVars = (self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][0]).index('_i')
                # Asumo un maximo de 999 variables input
                nInputVars = int((self.configVarsDict['GLBLnombreFicheroConModeloParaInferencia'][0])[posicionNumInputVars + 2 : posicionNumInputVars + 5])
                if self.LCLverbose:
                    print(f'clidconfig-> AVISO: si mi objetivo es CREAR_LAZ y mi modelo no usa las hiperformas, me ahorro ese paso.')
                    print(f'{TB}-> Saco el numero de inputs del nombre del modelo:')
                    print(f'{TB}{TV}-> GLBLnombreFicheroConModeloParaInferencia: {self.configVarsDict["GLBLnombreFicheroConModeloParaInferencia"][0]}')
                    print(f'{TB}{TV}-> nInputVars: {nInputVars}')
            else:
                nInputVars = 0
            if (
                (nInputVars > 0 and not nInputVars in [71, 72, 73, 83, 84, 85])
                or (nInputVars == 0 and not self.configVarsDict['GLBLincluirVarDeHiperformasEnElModeloAcumulativo'][0])
            ):
                self.configVarsDict['GLBLcalcularHiperFormas'][0] = False
                if self.LCLverbose:
                    print(f'{TB}-> Como mi objetivo es CREAR_LAZ y mi modelo no usa las hiperformas')
                    print(f'{TB}{TV}Cambio GLBLcalcularHiperFormas a False')
            else:
                if self.LCLverbose:
                    print(f'{TB}-> No necesito cambiar GLBLcalcularHiperFormas a False')

        if (
            GLO.GLBLpredecirClasificaMiniSubCelConvolucional
            or GLO.GLBLpredecirCubiertasSingularesConvolucional
        ):
            GLO.GLBLcrearTilesTargetMiniSubCelSoloSiHayNoSueloSuficientes = False
            GLO.GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes = False

        if (
            not GLO.GLBLhacerInferenciaEntrenandoModeloAcumulativoTree
            and not GLO.GLBLentrenarNeuronalNetworkConTF
            and not GLO.GLBLcalcularAciertoEntrenandoModeloAcumulativoNeuronal
        ):
            if self.configVarsDict['GLBLcalcularMdk2mConPuntosClasificados'][0]:
                print('clidconfig-> ATENCION: Para generar plano Mdk se requieren puntos reclasificados (hay que hacer inferencia).')
                sys.exit(0)
            if self.configVarsDict['GLBLgenerarNuevoLaxReclasificado'][0]:
                print('clidconfig-> ATENCION: Para guardar puntos reclasificados las hay que hacer inferencia (tree o nln).')
                sys.exit(0)
            if self.configVarsDict['GLBLguardarTrainPointsEnShape'][0]:
                print('clidconfig-> ATENCION: Para guardar puntos reclasificados shape hay que hacer inferencia (tree o nln).')
                sys.exit(0)

        if not self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][0]:
            print(
                '\nclidconfig-> AVISO: he quitado la opcion de guardar como string.',
                '\nSe cambia GLBLalmacenarPuntosComoNumpyDtype a True (GLBLalmacenarPuntosComoByteString = False).'
            )
            self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][0] = True
            self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0] = False


        if self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][0]:
            print(
                '\nclidconfig-> AVISO: he quitado la opcion de guardar version mini string.',
                '\nSe cambia GLBLalmacenarPuntosComoCompactNpDtype a False.'
            )
            self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][0] = False

        if (
            self.configVarsDict['GLBLsoloCrearTilesNoGuardarAsc'][0]
            or self.configVarsDict['GLBLsoloGuardarArraysNpzSinCrearOutputFiles'][0]
        ) and (
            self.configVarsDict['GLBLcrearTilesPostVuelta2'][0]
            or self.configVarsDict['GLBLguardarArraysVuelta2a9EnNpz'][0]
            or self.configVarsDict['GLBLreDepurarMiniSubCelEnVueltaAjustesMdp'][0]
        ):
            if self.LCLverbose:
                print(
                    '\nclidconfig-> AVISO: Si no se crean output files (asc), no tienen validez estos parametros:',
                    '\nGLBLcrearTilesPostVuelta2',
                    '\nGLBLguardarArraysVuelta2a9EnNpz',
                    '\nGLBLreDepurarMiniSubCelEnVueltaAjustesMdp'
                )

        if not self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][0]:
            if self.LCLverbose:
                print('\nclidconfig-> AVISO: revisar si se quiere usar formato de punto de texto (en desuso)')
            if (
                GLO.GLBLacumularPuntosEnNpzParaEntrenamientoFuturo
                or GLO.GLBLentrenarNeuronalNetworkConTF
                or GLO.GLBLhacerInferenciaParaTodosLosPuntos
            ):
                print('\nclidconfig-> ATENCION: no se puede haer entrenamiento o inferencia con formato de punto que no sea dtype')
                print(f'{TB}-> Cambiar la configuracion')
                sys.exit(0)

        if self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][0] != 0:
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                if self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][0] in [1, 7, 12, 18] and (
                    self.configVarsDict['GLBLentrenarConTarget_1lasOrig_2LasAsig_3lasPred'][0] == 2
                ):
                    if self.LCLverbose:
                        print(
                            '\nclidconfig-> AVISO: Se va a entrenar con puntos acumulativos con una sola clase seleccionada',
                            '\nLa clase seleccionada es {} (no interesa): Revisar GLBLentrenarBinarioCategoriaSeleccionada'.format(
                                self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][0]
                            )
                        )
                elif self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][0] > 100:
                    if self.LCLverbose:
                        print(
                            '\nclidconfig-> AVISO: No hay categorias de lasClass superiores a 100. GLBLentrenarBinarioCategoriaSeleccionada: {}'.format(
                                self.configVarsDict['GLBLentrenarBinarioCategoriaSeleccionada'][0]
                            )
                        )

        if self.configVarsDict['GLBLcalcularHiperFormas'][0]:
            if (
                not self.configVarsDict['GLBLcalcularMds'][0]
                and not self.configVarsDict['GLBLcalcularMdb'][0]
                and (
                    not self.configVarsDict['GLBLcalcularMdp'][0] or not (
                        self.configVarsDict['GLBLcalcularMdfConMiniSubCelValidadosConMetodoManualPuro'][0]
                        or self.configVarsDict['GLBLcalcularMdfConMiniSubCelValidadosConModeloConvolucional'][0]
                        or self.configVarsDict['GLBLcalcularMdfConMiniSubCelValidadosConModConvoManualizado'][0]
                    )
                )
            ):
                self.configVarsDict['GLBLcalcularHiperFormas'][0] = False
                if self.configVarsDict['GLBLmostrarAvisos'][0]:
                    if self.LCLverbose:
                        print(
                            '\nclidconfig-> AVISO: Se desactiva el calculo de hiperformas porque no se calculan el plano pleno ni basal ni suelo.',
                            '\n  Revisar GLBLcalcularHiperFormas, GLBLcalcularMdp, GLBLcalcularMdfConMiniSubCelValidados***',
                            '\n  Revisar configuracion.'
                        )

        if self.configVarsDict['GLBLacumularPuntosEnNpzParaEntrenamientoFuturo'][0]:
            if (
                not self.configVarsDict['GLBLcrearTilesPostVuelta1'][0]
                or not self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][0]
                or not self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][0]
            ):
                print('\nclidconfig-> ATENCION: Para acumular puntos para entrenamiento, es necesario predecir convolucionalmente la clase de los miniSubCel y los usos singulares')
                print('\t-> GLBLcrearTilesPostVuelta1:                    {}'.format(self.configVarsDict['GLBLcrearTilesPostVuelta1'][0]))
                print('\t-> GLBLpredecirCubiertasSingularesConvolucional: {}'.format(self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][0]))
                print('\t-> GLBLpredecirClasificaMiniSubCelConvolucional: {}'.format(self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][0]))
                print('Revisar configuracion.')
                sys.exit(0)


        if self.configVarsDict['GLBLsoloGuardarArraysNpzSinCrearOutputFiles'][0] and not self.configVarsDict['GLBLcrearTilesPostVuelta2'][0] and self.configVarsDict['GLBLmostrarAvisos'][0]:
            if self.LCLverbose:
                print(
                    '\nclidconfig-> AVISO: si GLBLsoloGuardarArraysNpzSinCrearOutputFiles,',
                    '\n  no tiene sentido que este activado GLBLcrearTilesPostVuelta2.',
                    '\n  Se puede cambiar en el fichero de configuracion.'
                )
        
        if (
            (
                self.configVarsDict['GLBLcrearTilesTargetDeCartoRefMdt'][0]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefNucleos'][0]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefLandCover'][0]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefSingUse'][0]
                or self.configVarsDict['GLBLcrearTilesTargetDeCartoRefPixel1m'][0]
            )
            and not self.configVarsDict['GLBLformatoTilesAscRasterRef'][0]
            and self.configVarsDict['GLBLmostrarAvisos'][0]
        ):
            if self.LCLverbose:
                print(f'\nclidconfig-> AVISO: si GLBLcrearTilesTargetDeCartoRefSingUse u otros TargetDeCartoRef,')
                print(f'{TB}-> es recomendable GLBLformatoTilesAscRasterRef para visualizar en Qgis los tiles que se generan.')
                print(f'{TB}-> Se puede cambiar en el fichero de configuracion.')

        if (
            self.configVarsDict['GLBLeliminarTilesTrasProcesado'][0]
            and (
                not self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][0]
                or not self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][0]
                or self.configVarsDict['GLBLsoloCrearTilesNoGuardarAsc'][0]
            )
            and self.configVarsDict['GLBLmostrarAvisos'][0]
        ):
            print(
                '\nclidconfig-> ATENCION: si GLBLeliminarTilesTrasProcesado,',
                '\n  es porque se usan para hacer predicciones, (GLBLpredecirCubiertasSingularesConvolucional or GLBLpredecirClasificaMiniSubCelConvolucional).',
                '\n  y porque no es una ejecucion para GLBLsoloCrearTilesNoGuardarAsc.'
                '\n  Revisar configuracion.'
            )
            return False

        if (
            self.configVarsDict['GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes'][0]
            and (
                not self.configVarsDict['GLBLsoloCrearTilesNoGuardarAsc'][0]
                or self.configVarsDict['GLBLpredecirCubiertasSingularesConvolucional'][0]
                or self.configVarsDict['GLBLpredecirClasificaMiniSubCelConvolucional'][0]
            )
            and self.configVarsDict['GLBLmostrarAvisos'][0]
        ):
            print(
                '\nclidconfig-> ATENCION: si GLBLcrearTilesTargetDeCartoRefSoloSiHaySingUseSuficientes,',
                '\n  es porque se trata de una ejecucion destinada solo a crear tiles (GLBLsoloCrearTilesNoGuardarAsc)',
                '\n  y no se va a hacer predicciones, (not GLBLpredecirCubiertasSingularesConvolucional and not GLBLpredecirClasificaMiniSubCelConvolucional).',
                '\n  Para hacer predicciones se deberian generar todos los tiles.'
                '\n  Se puede cambiar en el fichero de configuracion.'
            )
            return False

        minimoDeMemoriaRAMrecomendable = asignarMinimoDeMemoriaRAM(self.configVarsDict['GLBLminimoDeMemoriaRAM'][0])
        if minimoDeMemoriaRAMrecomendable != self.configVarsDict['GLBLminimoDeMemoriaRAM'][0]:
            self.configVarsDict['GLBLminimoDeMemoriaRAM'] = [minimoDeMemoriaRAMrecomendable, 'GrupoManejoMemoria', '', 'int']
        if self.configVarsDict['GLBLusarNumba'][0] and self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][0] < 100:
            self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][0] = 100
        if not self.configVarsDict['GLBLcalcularMdb'][0]:
            self.configVarsDict['GLBLproyectarPuntosSobreMdb'][0] = False
        if self.configVarsDict['GLBLmetrosBloc'][0] < self.configVarsDict['GLBLmetrosCelda'][0]:
            self.configVarsDict['GLBLmetrosBloc'][0] = self.configVarsDict['GLBLmetrosCelda'][0]
        if self.configVarsDict['GLBLtipoIndice'][0] == 1:
            self.configVarsDict['GLBLtipoIndice'][0] = int(
                math.ceil(1.0 * self.configVarsDict['GLBLmetrosBloque'][0] / self.configVarsDict['GLBLmetrosBloc'][0])
            )  # malla regular de blocs
            self.configVarsDict['GLBLtipoIndice'][0] = 100 if self.configVarsDict['GLBLtipoIndice'][0] > 100 else self.configVarsDict['GLBLtipoIndice'][0]
        if self.configVarsDict['GLBLmetrosCelda'][0] < 5:
            self.configVarsDict['GLBLminimoDePuntosTodosRetornosPasadaSeleccionada'][0] = int(
                self.configVarsDict['GLBLminimoDePuntosTodosRetornosPasadaSeleccionada'][0] / 2
            )
            self.configVarsDict['GLBLminimoDePuntosTotales'][0] = int(self.configVarsDict['GLBLminimoDePuntosTotales'][0] / 2)
            self.configVarsDict['GLBLminimoDePuntosSueloParaAjustarPlano'][0] = int(self.configVarsDict['GLBLminimoDePuntosSueloParaAjustarPlano'][0] / 2)
            self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][0] = int(self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][0] / 2)
        if self.configVarsDict['GLBLminDePtosParaAjustarPlanoBasalCielo'][0] < 3 or self.configVarsDict['GLBLminDePtosParaAjustarPlanoMajor'][0] < 3:
            if self.LCLverbose:
                print(f'clidconfig-> Corregir GLBLminDePtosParaAjustarPlanoBasalCielo o GLBLminDePtosParaAjustarPlanoMajor')
                print(f'clidconfig-> El numero minimo de puntos para ajustar debe ser mayor de 3')
            return False

        if (
            self.configVarsDict['GLBLusarNumba'][0]
            and not self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][0]
            and not self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0]
        ):
            if self.LCLverbose:
                print(f'clidconfig-> Si se usa numba y no se usa Dtype, solo se puede guardar con GLBLalmacenarPuntosComoByteString = True -> Se cambia a True')
            try:
                selec = input('\tclidconfig-> Confirmar el cambio de GLBLalmacenarPuntosComoByteString a True (S/n)')
                self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0] = False if selec.upper() == 'N' else True
            except:
                self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0] = True
            if self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0] == False:
                if self.LCLverbose:
                    print(
                        '\tclidconfig-> self.configVarsDict["GLBLalmacenarPuntosComoByteString"] = False -> Opcion no permitida con GLBLusarNumba=True y self.configVarsDict["GLBLalmacenarPuntosComoNumpyDtype"]=False'
                    )
                return False
            if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                print('GLBLalmacenarPuntosComoByteString:', self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0])

        if self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][0] > 0 and not self.configVarsDict['GLBLselecPasadasConClasificacion'][0]:
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                if self.LCLverbose:
                    print(f'clidconfig-> -> la pasada elegida debe tener puntos suelo -> se cambia la opcion GLBLselecPasadasConClasificacion a True')
            try:
                selec = input('\tclidconfig-> Confirmar el cambio de GLBLselecPasadasConClasificacion a True (S/n)')
                self.configVarsDict['GLBLselecPasadasConClasificacion'][0] = False if selec.upper() == 'N' else True
            except:
                self.configVarsDict['GLBLselecPasadasConClasificacion'][0] = True
            if self.configVarsDict['GLBLselecPasadasConClasificacion'][0] == False:
                if self.LCLverbose:
                    print(
                        '\tclidconfig-> self.configVarsDict["GLBLselecPasadasConClasificacion"][0] = False -> Opcion no permitida con GLBLminimoDePuntosSueloParaElegirPasada > 0'
                    )
                return False
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                if self.LCLverbose:
                    print('GLBLselecPasadasConClasificacion:', self.configVarsDict['GLBLselecPasadasConClasificacion'][0])

        if self.configVarsDict['GLBLusarNumba'][0] and (self.configVarsDict['GLBLusarSklearn'][0] or self.configVarsDict['GLBLusarStatsmodels'][0]):
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                if self.LCLverbose:
                    print(f'clidconfig-> Cuando se usa Numba los ajustes se hacen con algebra matricial y no con SkLearn o Statsmodels.')
            try:
                selec = input('\tclidconfig-> Confirmar el cambio de GLBLusarSklearn y GLBLusarStatsmodels a False (S/n)')
                rpta = True if selec.upper() == 'N' else False
                self.configVarsDict['GLBLusarSklearn'][0] = rpta
                self.configVarsDict['GLBLusarStatsmodels'][0] = rpta
            except:
                self.configVarsDict['GLBLusarSklearn'][0] = False
                self.configVarsDict['GLBLusarStatsmodels'][0] = False
            if rpta == False:
                if self.LCLverbose:
                    print(f'clidconfig-> GLBLusarSklearn o GLBLusarStatsmodels = True -> Opcion no permitida con GLBLusarNumba True')
                return False
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                if self.LCLverbose:
                    print('GLBLusarSklearn:    ', self.configVarsDict['GLBLusarSklearn'][0])
                    print('GLBLusarStatsmodels:', self.configVarsDict['GLBLusarStatsmodels'][0])


        if (
            self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0]
            or self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][0]
        ):
            if self.configVarsDict['GLBLcalcularMds'][0]:
                if (
                    self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0]
                    and self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][0]
                ):
                    print(f'clidconfig-> AVISO:')
                    print(f'{TB}-> GLBLguardarPuntosSueloEnArrayPredimensionada = True')
                    print(f'{TB}-> GLBLguardarPuntosSueloAlFinalDelArrayPralAll = True')
                    print(f'{TB}-> Esto guarda los puntos duplicados y consume extra de RAM')
                    print(f'{TB}-> Se desactiva la opcion GLBLguardarPuntosSueloAlFinalDelArrayPralAll')
                    selec = 'N'
                    self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][0] = False if selec.upper() == 'N' else True
            else:
                print(f'clidconfig-> AVISO:')
                print(f'{TB}-> GLBLguardarPuntosSueloEnArrayPredimensionada = {self.configVarsDict["GLBLguardarPuntosSueloEnArrayPredimensionada"][0]}')
                print(f'{TB}-> GLBLguardarPuntosSueloAlFinalDelArrayPralAll = {self.configVarsDict["GLBLguardarPuntosSueloAlFinalDelArrayPralAll"][0]}')
                print(f'{TB}-> GLBLcalcularMds = False')
                print(f'{TB}-> Esto guarda los puntos suelo pero no se usan para Mds con lo que consume RAM innecesariamente.')
                print(f'{TB}-> Se desactivan las opciones GLBLguardarPuntosSueloEnArrayPredimensionada y GLBLguardarPuntosSueloAlFinalDelArrayPralAll.')
                selec = 'N'
                self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0] = False if selec.upper() == 'N' else True
                self.configVarsDict['GLBLguardarPuntosSueloAlFinalDelArrayPralAll'][0] = False if selec.upper() == 'N' else True
            if self.configVarsDict['GLBLverbose'][0]:
                print(f'{TB}-> GLBLguardarPuntosSueloEnArrayPredimensionada: {self.configVarsDict["GLBLguardarPuntosSueloEnArrayPredimensionada"][0]}')
                print(f'{TB}-> GLBLguardarPuntosSueloAlFinalDelArrayPralAll: {self.configVarsDict["GLBLguardarPuntosSueloAlFinalDelArrayPralAll"][0]}')
        else:
            if self.configVarsDict['GLBLcalcularMds'][0]:
                try:
                    print(f'clidconfig-> AVISO:')
                    print(f'{TB}-> Se quiere calcular el plano suelo; pero eso requiere ')
                    print(f'{TB}{TV}almacenar puntos suelo en array predimensionada.')
                    print(f'{TB}{TV}-> Se activa esa opcion.')
                    selec = 'S'
                    # selec = input(f'{TB}{TV}almacenar puntos suelo en array predimensionada. Activar esa opcion? (S/n)')
                    confirmar = False if selec.upper() == 'N' else True
                except:
                    confirmar = True
                if confirmar:
                    self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0] = True
                    print(f'{TB}-> Se usa el array self_aCeldasListaDePtosSuePral. ')
                    print(f'{TB}-> Si se quiere GLBLguardarPuntosSueloAlFinalDelArrayPralAll, cambiar clidbase.xlsx')
                else:
                    self.configVarsDict['GLBLcalcularMds'][0] = False
                    print(f'{TB}-> No se calcula en plano suelo.')

        # Atencion: para calcular el punto suelo fuerzo que se use GLBLguardarPuntosSueloEnArrayPredimensionada
        # por lo que lo siguiente ya no tiene vigencia porque
        # no se si se puede calcular plano suelo usando la Psel,
        # sin almacenar los puntos suelo en su array/lugar especifico
        if (
            self.configVarsDict['GLBLcalcularMds'][0]
            and not self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0]
            and not self.configVarsDict['GLBLselecPasadasConClasificacion'][0]
            and self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][0] == 0
        ):
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                print(f'clidconfig-> -> Se ajusta plano a los puntos suelo pero:')
                print(f'{TB}No se guardan en array los puntos de la pasada con puntos suelo porque')
                print(f'{TB}ocupa demasiada memoria, por lo que se usan siempre los puntos de la pasada Psel.')
                print(f'{TB}La pasada Psel se selecciona por angulo de incidencia sin requerir que tenga puntos clasificados suelo')
                print(f'{TB}GLBLcalcularMds = True pero GLBLguardarPuntosSueloEnArrayPredimensionada = False ')
                print(f'{TB}y GLBLselecPasadasConClasificacion = False y GLBLminimoDePuntosSueloParaElegirPasada = 0')
            try:
                print(f'{TB}Confirmar que se quiere calcular el plano suelo usando la pasada seleccionada')
                selec = input('solo por angulo y que, por lo tanto, puede no tener puntos suelo (S/n)')
                # if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                #     print(
                #         f'{TB}Confirmar que se quiere calcular el plano suelo usando la pasada seleccionada solo por angulo y que, por lo tanto, puede no tener puntos suelo (S/n)'
                #     )
                #     selec = 'S'
                confirmar = False if selec.upper() == 'N' else True
            except:
                confirmar = True
            if not confirmar:
                try:
                    selec = input('\tclidconfig->     Calcular el plano suelo? (S/n)')
                    self.configVarsDict['GLBLcalcularMds'][0] = False if selec.upper() == 'N' else True
                except:
                    self.configVarsDict['GLBLcalcularMds'][0] = True
                if self.configVarsDict['GLBLcalcularMds'][0]:
                    try:
                        selec = input('\tclidconfig->     Requerir que la pasada seleccionada tenga puntos clasificados? (S/n)')
                        requerirPuntosClasificados = False if selec.upper() == 'N' else True
                    except:
                        requerirPuntosClasificados = True
                    if requerirPuntosClasificados:
                        self.configVarsDict['GLBLselecPasadasConClasificacion'][0] = True
                        try:
                            selec = input('\tclidconfig->     Requerir que la pasada seleccionada tenga al menos un punto clasificado suelo? (S/n)')
                            requerirPuntoClasificadoSuelo = False if selec.upper() == 'N' else True
                        except:
                            requerirPuntoClasificadoSuelo = True
                        if requerirPuntoClasificadoSuelo:
                            self.configVarsDict['GLBLselecPasadasConClasificacion'][0] = True
                            self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][0] = 1
            if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                print(
                    '\tclidconfig->     GLBLcalcularMds =',
                    self.configVarsDict['GLBLcalcularMds'][0],
                    'GLBLguardarPuntosSueloEnArrayPredimensionada =',
                    self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0],
                    'GLBLminimoDePuntosSueloParaElegirPasada =',
                    self.configVarsDict['GLBLminimoDePuntosSueloParaElegirPasada'][0],
                )

            if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                print(
                    '\tclidconfig-> GLBLselecPasadasConClasificacion:',
                    self.configVarsDict['GLBLselecPasadasConClasificacion'][0],
                    'GLBLselecPasadaConMasPuntosSuelo:',
                    self.configVarsDict['GLBLselecPasadaConMasPuntosSuelo'][0],
                )
            if self.configVarsDict['GLBLselecPasadaConMasPuntosSuelo'][0]:
                if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                    print(f'clidconfig-> Se selecciona la misma pasada para puntos suelo y para puntos basales')
            elif self.configVarsDict['GLBLselecPasadasConClasificacion'][0]:
                if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                    print(f'clidconfig-> Si solo hay una pasada con puntos clasificados, se selecciona la misma pasada para puntos suelo y para puntos basales')
                    print(
                        f'clidconfig-> Si hay varias, en las celdas con mas de una pasada, la pasada seleccionada (para puntos basales) puede ser distinta de la seleccionada para puntos suelo:'
                    )
                    print(f'clidconfig->     Para puntos suelo: la que teniendo puntos clasificados tenga mas puntos suelo')
                    print(f'clidconfig->     Para puntos basales: la que teniendo puntos clasificados tenga menor angulo medio')
            else:
                if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                    print(
                        f'clidconfig-> En las celdas con mas de una pasada, la pasada seleccionada para puntos basales puede ser distinta de la seleccionada para puntos suelo'
                    )
                    print(f'clidconfig->     Para puntos suelo: la que tenga mas puntos suelo')
                    print(f'clidconfig->     Para puntos basales: la que tenga menor angulo medio')
                    print(f'clidconfig-> AVISO: se trabaja solo con celdas seleccionadas para puntos basales: algunas pueden no tener puntos suelo')

        if self.configVarsDict['GLBLgrabarPropiedadTime'][0] and self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][0]:
            if self.configVarsDict['GLBLmostrarAvisos'][0]:
                print(
                    '\tclidconfig-> GLBLgrabarPropiedadTime = True-> Si se quiere guardar la propiedad RawTime no se puede usar GLBLalmacenarPuntosComoCompactNpDtype'
                )
            try:
                selec = input('\tclidconfig-> Quieres mantener GLBLalmacenarPuntosComoCompactNpDtype = True (no se graba RawTime)? (S/n)')
                self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][0] = False if selec.upper() == 'N' else True
            except:
                self.configVarsDict['GLBLalmacenarPuntosComoCompactNpDtype'][0] = True
            if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
                print('GLBLguardarPuntosSueloEnArrayPredimensionada:', self.configVarsDict['GLBLguardarPuntosSueloEnArrayPredimensionada'][0])

        if self.configVarsDict['GLBLgeigerMode'][0]:
            if self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0]:
                self.configVarsDict['GLBLalmacenarPuntosComoByteString'][0] = False
                self.configVarsDict['GLBLalmacenarPuntosComoNumpyDtype'][0] = True
            # Esto siguiente no hace falta porque mas adelante reviso el numero medio de puntos por celda y amplio GLBLnMaxPtosCeldaArrayPredimensionadaTodos si es necesario
            # if self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][0] < 20 * (self.configVarsDict['GLBLmetrosCelda'][0] ** 2):
            #    self.configVarsDict['GLBLnMaxPtosCeldaArrayPredimensionadaTodos'][0] = 20 * (self.configVarsDict['GLBLmetrosCelda'][0] ** 2)

        if self.configVarsDict['GLBLverbose'][0] and self.LCLverbose:
            print(f'{"":=^80}')

        # Esto no deberia ser necesario porque solo trabajo con self.configVarsDict[] o con GLO.paramConfig.
        global GLOBALconfigDict
        GLOBALconfigDict = self.configVarsDict

        return True


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    #     def nuevosParametrosGLBL(self):
    #         #Nuevas variables globales
    #         if GLOBALconfigDict['GLBLmetrosCelda'][0] >= 10:
    #             GLOBALconfigDict['nAnillosCheadosParaZonaUrbana'] = [int(40/GLOBALconfigDict['GLBLmetrosCelda'][0]), 'GrupoMAIN', '', 'str']
    #             #Para 10 m -> anillos 1, 2 y 3: 8+16+24=48 celdas
    #         else:
    #             GLOBALconfigDict['nAnillosCheadosParaZonaUrbana'] = [int(35/GLOBALconfigDict['GLBLmetrosCelda'][0]), 'GrupoMAIN', '', 'str']
    #             #Para 5 m -> anillos 1, 2, 3 y 4: 8+16+24+32=80 celdas
    #             #Para 5 m -> anillos 1, 2, 3, 4 y 5: 8+16+24+32+40=120 celdas
    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
    def registrarVariablesObjectControlFiles(self, ctrlFileLasName='', ctrlFileLas=None, GLBLficheroDeCtrlGral='', ctrlFileGral=None):
        self.controlFileLasName = ctrlFileLasName
        self.controlFileLas = ctrlFileLas
        self.GLBLficheroDeControlGral = GLBLficheroDeCtrlGral
        self.controlFileGral = ctrlFileGral
        self.controlFileGlobal = ctrlFileGral


    # ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
#     def registrarVariablesObjectControlFilesAuto(self):
#         # Tries to read global variables in this module defined previously: doesn't work
#         self.controlFileLasName = controlFileLasName
#         self.controlFileLas = controlFileLas
#         self.GLBLficheroDeControlGral = GLBLficheroDeControlGral
#         self.controlFileGral = controlFileGral
#         self.controlFileGlobal = controlFileGlobal


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# He modificado la funcion para que el tipo se deduzca del valor;
# Ya no es necesario poner el tipo de variable en el fichero de configuracion,
# salvo que quiera que un numero se interprete como texto, por ejemplo: 0000
# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def valorConfig(valorPrincipalTxt, valorAlternativoTxt='', usarAlternativo=False, nombreParametro='SinNombre', tipoVariable=''):
    # if nombreParametro == 'GLBLmetrosCelda':
    #     print('clidconfig->Borrar-> nombreParametro', nombreParametro, valorPrincipalTxt, tipoVariable)
    if valorPrincipalTxt == None:
        valorPrincipalOk = None
        tipoVariable = 'NoneType'
    elif valorPrincipalTxt == 'None':
        valorPrincipalOk = None
    elif tipoVariable == 'str':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. 0000 puede ser str
        valorPrincipalOk = valorPrincipalTxt
    elif tipoVariable == 'dict_str':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. '[0000, 1]' puede ser list_str
        try:
            valorPrincipalOk = {valorPrincipalTxt.split(',')[0]: valorPrincipalTxt.split(',')[1:]}
        except:
            valorPrincipalOk = {}
    elif tipoVariable == 'dict_int':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. '[0000, 1]' puede ser list_str
        try:
            valorPrincipalOk = {valorPrincipalTxt.split(',')[0]: [int(item) for item in valorPrincipalTxt.split(',')[1:]]}
        except:
            valorPrincipalOk = {}
    elif tipoVariable == 'list_str':
        # Manda sobre el aspecto que tiene el valor adoptado: p. ej. '[0000, 1]' puede ser list_str
        valorPrincipalOk = valorPrincipalTxt.split(',')
        # print('-->>', nombreParametro, tipoVariable, type(valorPrincipalOk), valorPrincipalOk)
    elif tipoVariable == 'list_int':
        valorPrincipalList = valorPrincipalTxt.split(',')
        valorPrincipalOk = []
        try:
            valorPrincipalOk = [int(item) for item in valorPrincipalList]
        except:
            print('clidconfig-> Error en parametro', nombreParametro, 'tiene formato distinto al previsto:', tipoVariable)
        # print('-->>', nombreParametro, tipoVariable, type(valorPrincipalOk), valorPrincipalOk)
    elif tipoVariable == 'list_float':
        valorPrincipalList = valorPrincipalTxt.split(',')
        valorPrincipalOk = []
        try:
            valorPrincipalOk = [float(item) for item in valorPrincipalList]
        except:
            print('clidconfig-> Error en parametro', nombreParametro, 'tiene formato distinto al previsto.')
    else:
        # Incluye el tipoVariable == 'desconocido'
        try:
            valorPrincipalFloat = float(valorPrincipalTxt)
            esNumero = True
        except:
            esNumero = False
        if esNumero:
            try:
                if valorPrincipalFloat == int(valorPrincipalFloat):
                    valorPrincipalOk = int(valorPrincipalFloat)
                    tipoVariable = 'int'
                else:
                    valorPrincipalOk = valorPrincipalFloat
                    tipoVariable = 'float'
            except:
                valorPrincipalOk = ''
                tipoVariable = 'error'
        else:
            if valorPrincipalTxt == 'True':
                valorPrincipalOk = True
                tipoVariable = 'bool'
            elif valorPrincipalTxt == 'False':
                valorPrincipalOk = False
                tipoVariable = 'bool'
            elif valorPrincipalTxt == 'None':
                valorPrincipalOk = None
                # No se el tipo de variable, pero todas las None de lisdas.xml son str
                tipoVariable = 'str'
            else:
                valorPrincipalOk = valorPrincipalTxt
                tipoVariable = 'str'
        # print('-->>', nombreParametro, tipoVariable, type(valorPrincipalOk), valorPrincipalOk)
        # if nombreParametro == 'GLBLmetrosCelda':
        #     print('clidconfig->Borrar-> int/float valorPrincipalOk', valorPrincipalOk, type(valorPrincipalOk))

    if usarAlternativo and valorAlternativoTxt != '':
        if tipoVariable == 'NoneType':
            valorAlternativoOk = None
        elif tipoVariable == 'str':
            valorAlternativoOk = valorAlternativoTxt
        elif tipoVariable == 'bool':
            if valorAlternativoTxt == 'True':
                valorAlternativoOk = True
            elif valorAlternativoTxt == 'False':
                valorAlternativoOk = False
            else:
                print('clidconfig-> Valor alternativo del parametro', nombreParametro, 'tiene formato distinto al principal; se usa el valor principal')
                valorAlternativoOk = valorPrincipalOk
        elif tipoVariable == 'int':
            try:
                valorAlternativoOk = int(valorAlternativoTxt)
            except:
                print('clidconfig-> Valor alternativo del parametro', nombreParametro, 'tiene formato distinto al principal; se usa el valor principal')
                valorAlternativoOk = valorPrincipalOk
        elif tipoVariable == 'float':
            try:
                valorAlternativoOk = float(valorAlternativoTxt)
            except:
                print('clidconfig-> Valor alternativo del parametro', nombreParametro, 'tiene formato distinto al principal; se usa el valor principal')
                valorAlternativoOk = valorPrincipalOk
        else:
            valorPrincipalOk = None
            valorAlternativoOk = None
        valorElegido = valorAlternativoOk
    else:
        valorElegido = valorPrincipalOk
    # if nombreParametro == 'GLBLmetrosCelda':
    #     print('clidconfig->Borrar-> valorElegido', nombreParametro, valorElegido, type(valorElegido))
    return valorElegido


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# Funcion sin uso
textoXML = ''
def leerConfigXML(miXML, esFile=True, verbose=False):
    if esFile:
        try:
            tree = ET.parse(miXML)
            root = tree.getroot()
        except:
            print('Error en el fichero de configuracion: se ignora')
            esFile = False
            miXML = textoXML
            root = ET.fromstring(miXML)
    else:
        root = ET.fromstring(miXML)

    # root [no] es una lista
    dictConfig = {}
    numSinNameXML = 1
    if verbose:
        print('ooooooooooooooooo Contenido del fichero XML de configuacion ooooooooooooooooooo')
    if root.tag != 'config':
        print('\nATENCION: Revisar el fichero de configuracion; el elemento principal debe llamarse: "config"')
        return dictConfig

    for item in root:
        if item.tag != 'param':
            print('Este item se ignora (solo se procesan los "param"):', item.tag)
            continue
        try:
            if 'name' in item.attrib.keys():
                nombreParametroDeConfiguracion = item.attrib['name']
                valorPrincipal = item.text
            else:
                nombreParametroDeConfiguracion = 'NoDisponible_%i' % numSinNameXML
                valorPrincipal = ''
                numSinNameXML += 1
                print('\nAVISO: Revisar el fichero de configuracion; hay entradas sin la propiedad "name"')
        except:
            print('\nAVISO: Revisar el fichero de configuracion; hay errores en el nombre de algun parametro')

        # Esta propiedad ya no es necesaria (se ignora)
        try:
            if 'tipo' in item.attrib.keys():
                tipoVariable = item.attrib['tipo']
            else:
                tipoVariable = 'str'
        except:
            print('\nATENCION: Revisar el fichero de configuracion; hay errores en la propiedad "tipo" del parametro', nombreParametroDeConfiguracion)

        # AVISO: El valor alternativo lo he abandonada al pasar de xml a xlsx
        # La propiedad 'ALTER' no es obligatoria
        # Si existe la propiedad 'ALTER' y tiene un valor distinto de '', ese es el valorAlternativo
        # En caso contrario, valorAlternativo = valorPrincipal
        try:
            if 'ALTER' in item.attrib.keys():
                valorAlternativo = item.attrib['ALTER'] if item.attrib['ALTER'] != '' else valorPrincipal
            else:
                valorAlternativo = valorPrincipal
        except:
            print('\nATENCION: Revisar el fichero de configuracion; hay errores en la propiedad "ALTER" del parametro', nombreParametroDeConfiguracion)

        # La propiedad 'grupo' no es obligatoria y solo sirve para tener los parametros organizados en clidbase.cfg
        try:
            if 'grupo' in item.attrib.keys():
                grupoParametros = item.attrib['grupo']
            else:
                grupoParametros = 'Miscelanea'
        except:
            print('\nATENCION: Revisar el fichero de configuracion; hay errores en la propiedad "tipo" del parametro', nombreParametroDeConfiguracion)

        # La propiedad 'desc' no es obligatoria y por el momento no la uso, es cosa interna
        try:
            if 'desc' in item.attrib.keys():
                descripcionParametro = item.attrib['desc']
            else:
                descripcionParametro = ''
        except:
            print('\nATENCION: Revisar el fichero de configuracion; hay errores en la propiedad "desc" del parametro', nombreParametroDeConfiguracion)

        # La propiedad 'uso' no es obligatoria y solo la uso como recordatorio de parametros que no uso pero puedo usar en el futuro
        try:
            if 'uso' in item.attrib.keys():
                usoParametro = item.attrib['uso']
                if usoParametro != 'Ok' and usoParametro[:1] != 'N' and usoParametro != 'FC':
                    print(
                        'Uso del parametro de configuracion %s: %s. Debe ser provisional: verificarlo y, si no se usa, eliminarlo del archivo de configuracion'
                        % (nombreParametroDeConfiguracion, usoParametro)
                    )
            else:
                usoParametro = ''
        except:
            print('\nATENCION: Revisar el fichero de configuracion; hay errores en la propiedad "uso" del parametro', nombreParametroDeConfiguracion)
            print(item.attrib['uso'])

        dictConfig[nombreParametroDeConfiguracion] = [tipoVariable, valorPrincipal, valorAlternativo, grupoParametros, descripcionParametro, usoParametro]
        if verbose:
            print(
                '->XML: %s: <%s>%s%s'
                % (
                    nombreParametroDeConfiguracion.rjust(50),
                    item.text,
                    ' Valor alternativo: ' if valorAlternativo != item.text else '',
                    valorAlternativo if valorAlternativo != item.text else '',
                )
            )

    if verbose:
        print('ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo')
        for clave in dictConfig.keys():
            print('clidconfig-> ->->', clave, dictConfig[clave])
    return dictConfig


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def normalize(c):
    return unicodedata.normalize("NFD", c)[0]


# ==============================================================================
def get_MAIN_CFG_DIR():
    # Orden de preferencia para el fichero de configuracion:
    #  1-> El del directorio actual (desde el que se lanza la aplicacion) si no incluye site-packages
    #  2-> El del modulo inicial si no incluye site-packages
    #  3-> El del modulo actual si no incluye site-packages
    #  4-> El user/documents
    MAIN_THIS_DIR = os.getcwd()
    if not 'site-packages' in MAIN_THIS_DIR:
        MAIN_CFG_DIR = MAIN_THIS_DIR
    else:
        MAIN_BASE_DIR = os.path.abspath(os.path.dirname(sys.argv[0]))
        if not 'site-packages' in MAIN_BASE_DIR:
            MAIN_CFG_DIR = MAIN_BASE_DIR
        else:
            try:
                MAIN_FILE_DIR = os.path.dirname(os.path.abspath(__file__))
                MAIN_FILE_DIR_ok = True
            except:
                MAIN_FILE_DIR = ''
                MAIN_FILE_DIR_ok = False
            if not 'site-packages' in MAIN_FILE_DIR and MAIN_FILE_DIR_ok:
                MAIN_CFG_DIR = MAIN_FILE_DIR
            else:
                MAIN_HOME_DIR = str(pathlib.Path.home())
                MAIN_CFG_DIR = os.path.join(MAIN_HOME_DIR, 'Documents')
    return MAIN_CFG_DIR


# ==============================================================================
def getConfigFileName():
    if len(sys.argv) == 0:
        print(f'\nclidconfig-> Revisar esta forma de ejecucion. sys.argv: <{sys.argv}>')
        sys.exit(0)
    elif (sys.argv[0].endswith('__main__.py') and 'cartolidar' in sys.argv[0]):
        # print('\nqlidtwins.py se ejecuta lanzando el paquete cartolidar desde linea de comandos:')
        # print('\t  python -m cartolidar')
        configFileNameSinExt = 'clidbase{:006}'.format(int(MAIN_idProceso))
    elif sys.argv[0].endswith('qlidtwins.py'):
        # print('\nqlidtwins.py se ha lanzado desde linea de comandos:')
        # print('\t  python qlidtwins.py')
        configFileNameSinExt = 'qlidtwins'
    elif sys.argv[0] == '':
        # print('\nqlidtwins se esta importando desde el interprete interactivo:')
        configFileNameSinExt = 'clidbase{:006}'.format(int(MAIN_idProceso))
    else:
        # print(f'\nqlidtwins.py se esta importando desde el modulo: {sys.argv[0]}')
        if MAIN_idProceso:
            if not type(MAIN_idProceso) == int and not type(MAIN_idProceso) == str:
                print('\nclidconfig-> AVISO: Revisar asignacion de idProceso (no es int ni str):')
                print('idProceso:   <{}> type: {}'.format(MAIN_idProceso, type(MAIN_idProceso)))
                print('sys.argv[0]: <{}>'.format(sys.argv[0]))
            try:
                if sys.argv[0].endswith('.py'):
                    configFileNameSinExt = os.path.basename(sys.argv[0]).replace('.py', '{:006}'.format(int(MAIN_idProceso)))
                elif sys.argv[0].endswith('pytest'):
                    configFileNameSinExt = 'cfgForTest'
                else:
                    configFileNameSinExt = 'unknownLaunch'
            except:
                print('\nclidconfig-> Revisar asignacion de idProceso (b):')
                print('idProceso:   <{}> type: {}'.format(MAIN_idProceso, type(MAIN_idProceso)))
                print('sys.argv[0]: <{}>'.format(sys.argv[0]))
                sys.exit(0)
        else:
            if sys.argv[0].endswith('.py'):
                configFileNameSinExt = os.path.basename(sys.argv[0]).replace('.py', '')
            elif sys.argv[0].endswith('pytest'):
                configFileNameSinExt = 'cfgForTest'
            else:
                configFileNameSinExt = 'unknownLaunch'

    return configFileNameSinExt


# ==============================================================================
def getConfigFileNameCfg(idProceso, LOCL_verbose=0):
    rutaAlternativa = False
    MAIN_CFG_DIR = get_MAIN_CFG_DIR()
    configFileNameSinExt = getConfigFileName()
    configFileNameSinPath = f'{configFileNameSinExt}.cfg'
    
    configFileNameCfg = os.path.join(MAIN_CFG_DIR, configFileNameSinPath)
    if LOCL_verbose:
        print(f'\n{"":_^80}')
        print(f'clidconfig-> Fichero de configuracion cfg: {configFileNameCfg}')

    try:
        if not os.path.exists(configFileNameCfg):
            controlConfigFile = open(configFileNameCfg, mode='w')
            controlConfigFile.close()
            os.remove(configFileNameCfg)
        else:
            controlConfigFile = open(configFileNameCfg, mode='r+')
            controlConfigFile.close()
        if LOCL_verbose:
            if os.path.exists(configFileNameCfg):
                print(f'{TB}-> Este fichero de configuracion ya existe previamente (se usan sus parametros).')
            print(f'{"":=^80}')
    except:
        if LOCL_verbose:
            # print(f'\n{"":_^80}')
            print(f'clidconfig-> AVISO:')
            print(f'{TB}-> No se puede guardar el fichero de configuracion:  {configFileNameCfg}')
            print(f'{TB}-> Es posible que no tenga permisos de escritura en: {MAIN_CFG_DIR}')
            print(f'{TB}-> O que exista el fichero {configFileNameCfg} y este bloqueado.')
            print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')
        rutaAlternativa = True
        try:
            MAIN_HOME_DIR = str(pathlib.Path.home())
            MAIN_CFG_DIR = os.path.join(MAIN_HOME_DIR, 'Documents')
            if LOCL_verbose:
                print(f'{TB}Se intenta la ruta alternativa: {MAIN_CFG_DIR}')
            configFileNameCfg = os.path.join(MAIN_CFG_DIR, configFileNameSinPath)
            if not os.path.exists(configFileNameCfg):
                controlConfigFile = open(configFileNameCfg, mode='w')
                controlConfigFile.close()
                os.remove(configFileNameCfg)
            else:
                controlConfigFile = open(configFileNameCfg, mode='r+')
                controlConfigFile.close()
            if LOCL_verbose:
                print(f'{TB}Ok log file: {configFileNameCfg}')
        except:
            if LOCL_verbose:
                print(f'{TB}Tampoco se puede escribir en la ruta {MAIN_CFG_DIR}')
            MAIN_HOME_DIR = str(pathlib.Path.home())
            configFileNameCfg = os.path.join(MAIN_HOME_DIR, os.path.basename(sys.argv[0]).replace('.py', '.cfg'))
            if LOCL_verbose:
                print(f'clidconfig-> Ok cfg file (b): {configFileNameCfg}')
        if LOCL_verbose:
            if os.path.exists(configFileNameCfg):
                print(f'{TB}-> Este fichero de configuracion ya existe previamente (se usan sus parametros en lugar de los del xls).')
            print(f'{"":=^80}')

    return configFileNameCfg


# ==============================================================================
def getConfigFileNameXls(configFileNameCfg, LOCL_verbose=0):
    MAIN_XLS_DIR = get_MAIN_CFG_DIR()
    directorioActual = (os.getcwd()).replace(os.sep, '/')
    directorioProyecto = os.path.dirname(sys.argv[0])
    directorioCfg = os.path.dirname(configFileNameCfg)
    configFileNameSinPath = 'clidbase.xlsx'
    moduloInicialXls = (os.path.basename(sys.argv[0])).replace('.py', '.xlsx')

    # Opcion 1: fichero de configuracion con el nombre clidbase.xls
    #           en el directorio inicialmente propuesto para el cfg:
    #           1-> El del directorio actual (desde el que se lanza la aplicacion) si no incluye site-packages
    #           2-> El del modulo inicial si no incluye site-packages
    #           3-> El del modulo actual si no incluye site-packages
    #           4-> El user/documents
    filenameXlsCfg1Clidbase = os.path.join(MAIN_XLS_DIR, configFileNameSinPath)
    # No verifico si hay derechos de escritura en este directorio (cosa que si hago para cfg)

    # Opcion 2: fichero de configuracion con el nombre clidbase.xls
    #           en el directorio actual (desde el que se lanza la aplicacion)
    #           Puede coincidir con el anterior si no incluye site-packages
    filenameXlsCwdClidbase = os.path.join(directorioActual, configFileNameSinPath)

    # Opcion 3: fichero de configuracion con el nombre del modulo que se lanza inicialmente (__init__.py, __main__.py, clidbase.py, etc.)
    #           en el directorio actual (desde el que se lanza la aplicacion)
    #           Puede coincidir con los anteriores si el modulo inicial es clidbase.py
    filenameXlsCwdModulo = os.path.join(directorioActual, moduloInicialXls)

    # Opcion 4: fichero de configuracion con el nombre clidbase.xls
    #           en el directorio finalmente propuesto para el cfg:
    filenameXlsCfg2Clidbase = os.path.join(directorioCfg, configFileNameSinPath)

    # Opcion 5: fichero de configuracion instalado junto al modulo que lanza la aplicacion (clidbase.py)
    filenameXlsProjClidbase = os.path.abspath(os.path.join(directorioProyecto, configFileNameSinPath))


    return (
        filenameXlsCfg1Clidbase,
        filenameXlsCwdClidbase,
        filenameXlsCwdModulo,
        filenameXlsCfg2Clidbase,
        filenameXlsProjClidbase,
    )


# ==============================================================================
def guardarVariablesGlobales(
        LOCALconfigDict,
        LCL_idProceso=MAIN_idProceso,
        inspect_stack=inspect.stack(),
    ):
    # runnigFileName = sys.argv[0].replace('.py', '.running')
    # if not os.path.exists(runnigFileName):
    #     open(runnigFileName, mode='w+')
    #     time.sleep(2)
    # else:
    #     time.sleep(5)

    configFileNameCfg = getConfigFileNameCfg(LCL_idProceso, LOCL_verbose=0)
    if os.path.exists(configFileNameCfg):
        try:
            os.remove(configFileNameCfg)
        except:
            print('No se elimina el fichero', configFileNameCfg, 'porque debe haberse eliminado por otro proceso')
            time.sleep(5)

    if CONFIGverbose or LCLverbose:
        print(f'\n{"":_^80}')
        print(f'clidconfig-> Guardo los paramConfig en fichero cfg (inicial) con guardarVariablesGlobales:')
        print(f'{TB}{configFileNameCfg}')
        print(f'{TB}-> Reviso la pila de llamadas por si llamara a esta funcion de nuevo:')
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect_stack, verbose=CONFIGverbose)
        print(f'{TB}{TV}-> callingModulePrevio: {callingModulePrevio}; callingModuleInicial: {callingModuleInicial}')

    config = RawConfigParser()
    config.optionxform = str  # Avoid change to lowercase

    # When adding sections or items, add them in the reverse order of
    # how you want them to be displayed in the actual file.
    # In addition, please note that using RawConfigParser's and the raw
    # mode of ConfigParser's respective set functions, you can assign
    # non-string values to keys internally, but will receive an error
    # when attempting to write to a file or when you get it in non-raw
    # mode. SafeConfigParser does not allow such assignments to take place.

    # Creo los grupos de variables
    for nombreParametroDeConfiguracion in LOCALconfigDict.keys():
        grupoParametroConfiguracion = LOCALconfigDict[nombreParametroDeConfiguracion][1]
        if not grupoParametroConfiguracion in config.sections():
            # print('clidconfig-> ->grupoParametros nuevo', grupoParametroConfiguracion)
            config.add_section(grupoParametroConfiguracion)
    config.add_section('ChangedConfig')
    config.add_section('UserConfig')
    config.add_section('ControlFiles')

    numObjetivosExtraMax = 0
    for nombreParametroDeConfiguracion in LOCALconfigDict.keys():
        # LOCALconfigDict[nombreParametroDeConfiguracion] = [valorParametroDeConfiguracion, grupoParametros, tipoVariable, descripcionParametro]
        listaParametroConfiguracion = LOCALconfigDict[nombreParametroDeConfiguracion]
        valorParametroConfiguracion = listaParametroConfiguracion[0]
        grupoParametroConfiguracion = listaParametroConfiguracion[1]
        descripcionParametroConfiguracion = listaParametroConfiguracion[2]
        tipoParametroConfiguracion = listaParametroConfiguracion[3]

        # config.set(grupoParametroConfiguracion, nombreParametroDeConfiguracion, [str(valorParametroConfiguracion), tipoParametroConfiguracion])
        if not descripcionParametroConfiguracion is None:
            if (
                'á' in descripcionParametroConfiguracion
                or 'é' in descripcionParametroConfiguracion
                or 'í' in descripcionParametroConfiguracion
                or 'ó' in descripcionParametroConfiguracion
                or 'ú' in descripcionParametroConfiguracion
                or 'ñ' in descripcionParametroConfiguracion
                or 'ç' in descripcionParametroConfiguracion
            ):
                descripcionParametroConfiguracion = ''.join(normalize(c) for c in str(descripcionParametroConfiguracion))
            if (descripcionParametroConfiguracion.encode('utf-8')).decode('cp1252') != descripcionParametroConfiguracion:
                descripcionParametroConfiguracion = ''

        listaConcatenada = '{}|+|{}|+|{}'.format(
            str(valorParametroConfiguracion),
            str(tipoParametroConfiguracion),
            str(descripcionParametroConfiguracion)
            )

        numObjetivosExtra = max(len(listaParametroConfiguracion) - 4, 0)
        numObjetivosExtraMax = max(numObjetivosExtra, numObjetivosExtraMax)
        if numObjetivosExtraMax:
            valObjetivosExtra = []
            for nObjetivoExtra in range(numObjetivosExtraMax):
                if len(listaParametroConfiguracion) > 4 + nObjetivoExtra:
                    valObjetivosExtra.append(listaParametroConfiguracion[4 + nObjetivoExtra])
                    listaConcatenada += '|+|{}'.format(listaParametroConfiguracion[4 + nObjetivoExtra])
                else:
                    valObjetivosExtra.append(listaParametroConfiguracion[0])
                    listaConcatenada += '|+|{}'.format(listaParametroConfiguracion[0])
        config.set(
            grupoParametroConfiguracion,
            nombreParametroDeConfiguracion,
            listaConcatenada
        )
        # print(str(valorParametroConfiguracion), tipoParametroConfiguracion, descripcionParametroConfiguracion)

    # print('clidconfig-> configFileNameCfg:', configFileNameCfg)
    # with open(configFileNameCfg, 'wb') as configfile:
    try:
        with open(configFileNameCfg, mode='w+') as configfile:
            config.write(configfile)
    except:
        print(f'\nclidconfig-> ATENCION, revisar caracteres no admitidos en el fichero de configuracion: {configFileNameCfg}')
        print(f'{TB}Ejemplos: vocales acentuadas, ennes, cedillas, flecha dchea (->), etc.')

    if CONFIGverbose or LCLverbose:
        print(f'{TB}-> Ok creado fichero: {configFileNameCfg}')
        print(f'{"":=^80}')


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
def leerCambiarVariablesGlobales(
        nuevosParametroConfiguracion={},
        LCL_idProceso=MAIN_idProceso,
        inspect_stack=inspect.stack(),
        verbose=False
    ):
    # Lectura del config (cfg) generado especificamente para esta ejecucion.
    # print('LCL_idProceso:', type(LCL_idProceso), LCL_idProceso)
    # print('sys.argv:', sys.argv)
    configFileNameCfg = getConfigFileNameCfg(LCL_idProceso, LOCL_verbose=0)

    if CONFIGverbose or verbose:
        print(f'clidconfig-> Leo los paramConfig del cfg (lo actualizo si tengo nuevosParametroConfiguracion) con leerCambiarVariablesGlobales<>')
        print(f'{TB}Reviso la pila de llamadas para ver desde que modulo estoy cargando los paramConfig del cfg')
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
        print(f'{TB}-> callingModulePrevio: {callingModulePrevio}')

    config = RawConfigParser()
    config.optionxform = str  # Avoid change to lowercase

    if not os.path.exists(configFileNameCfg):
        print(f'\nclidconfig-> ATENCION: fichero de configuracion no encontrado: {configFileNameCfg}')
        print(f'{TB}-> Revisar la linea ~2523 de clidconfig.py, para que se cree el .cfg si callingModuleInicial == "clidbase" or ...')
        callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=True)
        print(f'{TB}-> callingModulePrevio: {callingModulePrevio} callingModuleInicial: {callingModuleInicial}')
        sys.exit(0)
        # return False

    if CONFIGverbose or verbose:
        print(f'clidconfig-> Leyendo cfg: {configFileNameCfg}')
    numObjetivosExtraMax = 0
    LOCALconfigDict = {}
    # if True:
    try:
        config.read(configFileNameCfg)
        mostrarFrupoMAIN = False
        if mostrarFrupoMAIN:
            print(f'clidconfig-> Paramtros del GrupoMAIN del fichero de configuracion ({configFileNameCfg}):')
        for grupoParametroConfiguracion in config.sections():
            for nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion):
                strParametroConfiguracion = config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion)
                listaParametroConfiguracion = strParametroConfiguracion.split('|+|')
                valorParametroConfiguracion = valorConfig(
                    listaParametroConfiguracion[0],
                    nombreParametro=nombreParametroDeConfiguracion,
                    tipoVariable=listaParametroConfiguracion[1]
                )
                if len(listaParametroConfiguracion) > 1:
                    tipoParametroConfiguracion = listaParametroConfiguracion[1]
                else:
                    tipoParametroConfiguracion = 'str'
                if len(listaParametroConfiguracion) > 2:
                    descripcionParametroConfiguracion = listaParametroConfiguracion[2]
                else:
                    descripcionParametroConfiguracion = ''
                if nombreParametroDeConfiguracion[:1] == '_':
                    grupoParametroConfiguracion_new = '_%s' % grupoParametroConfiguracion
                else:
                    grupoParametroConfiguracion_new = grupoParametroConfiguracion
                LOCALconfigDict[nombreParametroDeConfiguracion] = [
                    valorParametroConfiguracion,
                    grupoParametroConfiguracion_new,
                    descripcionParametroConfiguracion,
                    tipoParametroConfiguracion,
                ]
                numObjetivosExtra = max(len(listaParametroConfiguracion) - 3, 0)
                numObjetivosExtraMax = max(numObjetivosExtra, numObjetivosExtraMax)
                if numObjetivosExtraMax:
                    valObjetivosExtra = []
                    for nObjetivoExtra in range(numObjetivosExtraMax):
                        if 3 + nObjetivoExtra < len(listaParametroConfiguracion):
                            valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[3 + nObjetivoExtra], tipoVariable=listaParametroConfiguracion[1]))
                        else:
                            valObjetivosExtra.append(valorConfig(listaParametroConfiguracion[0], tipoVariable=listaParametroConfiguracion[1]))
                    LOCALconfigDict[nombreParametroDeConfiguracion].extend(valObjetivosExtra)

                if mostrarFrupoMAIN and grupoParametroConfiguracion == 'GrupoMAIN':
                    print(f'{TB}-> >>>5 numObjetivosExtra: {numObjetivosExtra}, Max: {numObjetivosExtraMax}, >>> {nombreParametroDeConfiguracion}, {LOCALconfigDict[nombreParametroDeConfiguracion]}')

        LOCALconfigDict['configFileNameCfg'] = [
            configFileNameCfg,
            'GrupoMAIN',
            'Nombre del fichero de configuracion (este fichero).',
            'str',
        ]
        if CONFIGverbose or verbose:
            print(f'clidconfig-> Parametros leidos ok del fichero cfg: {LOCALconfigDict["configFileNameCfg"][0]}')

        # configLeidoDelCfgOk = True
    except Exception as excpt:
        program_name = 'clidconfig.py'
        print(f'\n{program_name}-> Error Exception en clidconfig-> {excpt}\n')
        # https://stackoverflow.com/questions/1278705/when-i-catch-an-exception-how-do-i-get-the-type-file-and-line-number
        exc_type, exc_obj, exc_tb = sys.exc_info()
        if verbose > 1:
            print()
            # print(f'exc_obj ({type(exc_obj)}): <<{str(exc_obj)}>>')
            # print(dir(exc_obj)) # 'args', 'characters_written', 'errno', 'filename', 'filename2', 'strerror', 'winerror', 'with_traceback'
            try:
                numeroError = exc_obj.errno
            except:
                numeroError = -1
            print(f'numError:  {numeroError}')      # 13
            print(f'filename:  {exc_obj.filename}')
            print(f'filename2: {exc_obj.filename2}')
            try:
                print(f'strerror:  {exc_obj.strerror}')
            except:
                print(f'strerror_: {exc_obj}')
            # print(f'with_traceback: {exc_obj.with_traceback}')  # <built-in method
            print()
            print(f'filename {exc_tb.tb_frame.f_code.co_filename}')
            print(f'lineno   {exc_tb.tb_lineno}')
            print(f'function {exc_tb.tb_frame.f_code.co_name}')
            print(f'type     {exc_type.__name__}')
            # print(f'message  {exc_obj.message}')  # or see traceback._some_str()
            print()
    
        fileNameError = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        funcError = os.path.split(exc_tb.tb_frame.f_code.co_name)[1]
        lineError = exc_tb.tb_lineno
        typeError = exc_type.__name__
        try:
            descError = exc_obj.strerror
        except:
            descError = exc_obj
        sys.stderr.write(f'\nPuede contribuir a este programa remitiendo este error a cartolidar@gmail.com:\n')
        sys.stderr.write(f'{TB}Error en     {fileNameError}\n')
        sys.stderr.write(f'{TB}Funcion:     {funcError}\n')
        sys.stderr.write(f'{TB}Linea:       {lineError}\n')
        sys.stderr.write(f'{TB}Descripcion: {descError}\n') # = {exc_obj}
        sys.stderr.write(f'{TB}Tipo:        {typeError}\n')
    
        sys.stderr.write(f'clidconfig-> Error al leer la configuracion del fichero: {configFileNameCfg}\n')
        # configLeidoDelCfgOk = False
    
        sys.exit(0)


    if CONFIGverbose or verbose:
        print(f'clidconfig-> nuevosParametroConfiguracion: {nuevosParametroConfiguracion}')
    # Estos parametros llegan como dict de listas de valores (no como listas de textos, que es lo que ocurre con la listaParametroConfiguracion leida del cfg)        
    if nuevosParametroConfiguracion != {}:
        for nombreParametroDeConfiguracion in nuevosParametroConfiguracion.keys():
            listaNuevosParametroConfiguracion = nuevosParametroConfiguracion[nombreParametroDeConfiguracion]
            valorParametroConfiguracion = listaNuevosParametroConfiguracion[0]
            if valorParametroConfiguracion is None or valorParametroConfiguracion == 'None':
                continue
            try:
                tipoParametroConfiguracion = listaNuevosParametroConfiguracion[1]
            except:
                tipoParametroConfiguracion = 'str'
            try:
                descripcionParametroConfiguracion = listaNuevosParametroConfiguracion[2]
            except:
                descripcionParametroConfiguracion = ''
            try:
                grupoParametroConfiguracion = listaNuevosParametroConfiguracion[3]
            except:
                grupoParametroConfiguracion = 'GrupoDESC'
            if not grupoParametroConfiguracion in config.sections():
                # print('clidconfig-> ->grupoParametros nuevo', grupoParametroConfiguracion)
                config.add_section(grupoParametroConfiguracion)

            # Si el parametro ya tiene descripcion en clidbase.xml, se mantiene esa descripcion
            # print('clidconfig-> Revisando valores previos-> grupo: {} parametro: {}'.format(grupoParametroConfiguracion, nombreParametroDeConfiguracion))
            # print('\t', nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion))
            if nombreParametroDeConfiguracion in config.options(grupoParametroConfiguracion):
                listaParametroConfiguracionOriginal = config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion).split('|+|')
                descParametroConfiguracion = listaParametroConfiguracionOriginal[2]
                if descParametroConfiguracion != '':
                    descripcionParametroConfiguracion = descParametroConfiguracion
                # print('clidconfig-> valor previo de: {} -> {}'.format(nombreParametroDeConfiguracion, listaParametroConfiguracionOriginal))

            listaConcatenada = '{}|+|{}|+|{}'.format(
                str(valorParametroConfiguracion),
                str(tipoParametroConfiguracion),
                str(descripcionParametroConfiguracion)
                )

            if numObjetivosExtraMax:
                valObjetivosExtra = []
                for nObjetivoExtra in range(numObjetivosExtraMax):
                    if 4 + nObjetivoExtra < len(listaNuevosParametroConfiguracion):
                        valObjetivosExtra.append(listaNuevosParametroConfiguracion[4 + nObjetivoExtra])
                        listaConcatenada += '|+|{}'.format(listaNuevosParametroConfiguracion[4 + nObjetivoExtra])
                    else:
                        valObjetivosExtra.append(listaNuevosParametroConfiguracion[0])
                        listaConcatenada += '|+|{}'.format(listaNuevosParametroConfiguracion[0])

            config.remove_option(
                grupoParametroConfiguracion,
                nombreParametroDeConfiguracion,
            )
            config.set(
                grupoParametroConfiguracion,
                nombreParametroDeConfiguracion,
                listaConcatenada
            )

            LOCALconfigDict[nombreParametroDeConfiguracion] = [
                valorParametroConfiguracion,
                grupoParametroConfiguracion,
                descripcionParametroConfiguracion,
                tipoParametroConfiguracion,
            ]
            if numObjetivosExtra:
                LOCALconfigDict[nombreParametroDeConfiguracion].extend(valObjetivosExtra)

            if verbose:
                # print(f'{TB}Nuevo Valor de: {nombreParametroDeConfiguracion} -> {listaConcatenada}')
                print(f'{TB}-> Nuevo Valor ok: {nombreParametroDeConfiguracion} -> {config.get(grupoParametroConfiguracion, nombreParametroDeConfiguracion).split("|+|")}')

            if (CONFIGverbose or verbose) and grupoParametroConfiguracion == 'GrupoMAIN':
                # print(f'{TB}{TV}clidconfig-> >>>6 {configFileNameCfg}, {nombreParametroDeConfiguracion}, {listaNuevosParametroConfiguracion}')
                print(
                    f'{TB}-> >>>6 Nuevo parametro del grupo',
                    grupoParametroConfiguracion,
                    '->',
                    nombreParametroDeConfiguracion,
                    'con valor:',
                    valorParametroConfiguracion,
                    'y valores extras:',
                    valObjetivosExtra,
                    'listaNuevosParametroConfiguracion',
                    listaNuevosParametroConfiguracion
                )

        if CONFIGverbose or verbose:
            print(f'clidconfig-> Guardando los nuevos parametros en: {configFileNameCfg}')

        os.remove(configFileNameCfg)
        with open(configFileNameCfg, mode='w+') as configfile:
            config.write(configfile)

        if CONFIGverbose or verbose:
            if os.path.exists(configFileNameCfg):
                print(f'{TB}-> Ok fichero cfg actualizado con los nuevos parametros.')
            else:
                print(f'{TB}-> AVISO: error al crear {configFileNameCfg} con los nuevos parametros.')

    return LOCALconfigDict


# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
# A este modulo se le llama desde clidbase.py
def registrarVariablesModuleControlFiles(ctrlFileLasName='', ctrlFileLas='', GLBLficheroDeCtrlGral='', ctrlFileGral=''):
    # global controlsInitiated
    # controlsInitiated = True
    global controlFileLasName
    controlFileLasName = ctrlFileLasName
    global controlFileLas
    controlFileLas = ctrlFileLas
    global GLBLficheroDeControlGral
    GLBLficheroDeControlGral = GLBLficheroDeCtrlGral
    global controlFileGral
    controlFileGral = ctrlFileGral
    global controlFileGlobal
    controlFileGlobal = ctrlFileGral


# ==============================================================================
def cerrarFicheroLasPorErrorDeLectura(ficheroDeControlGral):
    print('Error de lectura del fichero las: Muchos puntos con coordenadas no validas')
    controlFileLas.close()

    controlFileGral = open(ficheroDeControlGral, mode='a+')
    controlFileGral.write('Error de lectura de fichero las. Ver fichero de control correspondiente.\n')
    controlFileGral.close()

    # liberaArraysVuelta1(interrumpidoPorError=True, completo=False)
    return


# ==============================================================================
def crearListaVariablesEntrenamientoAcumulativo(nInputVars=0):
    dictVariablesEndogenasTodas = {

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Requieren RGBI
        # miPto['red']..miPto['nir'] extraidos de self_aCeldasListaDePtosTlcPralPF99[] -> clidnv0->numbaMainVuelta0{}
        # RGBI de cada punto.
        # ATENCION: se usan distintas escalas en distintos ficheros. Ver clidflow -> #RangoDeValoresRGBI
        0: 'red',
        1: 'green',
        2: 'blue',
        3: 'nir',

        # miPto['intensity'] extraido de self_aCeldasListaDePtosTlcPralPF99[] -> clidnv0->numbaMainVuelta0{}
        # Intensidad de cada punto.
        # ATENCION: se usan distintas escalas en distintos ficheros. Ver clidflow -> #RangoDeValoresRGBI
        4: 'intensity',

        # self_aMetricoIntSRet[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # Valor medio de la intensidad de los puntos con un solo retorno en cada metro cuadrado (mas buffer).
        # ATENCION: se usan distintas escalas en distintos ficheros.  Ver clidflow -> #RangoDeValoresRGBI
        5: 'metroIntensitySRet',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> No uso estas variables
        # Indices calculadas con las propiedades RGBI de miPto[] extraidas de self_aCeldasListaDePtosTlcPralPF99[] -> clidnv0->numbaMainVuelta0{}
        # Re-escalado de puntoEVI2calculado = 2.5 * (miPtoNewNir - miPtoNewRed) / np.float32(miPtoNewNir + (2.4 * miPtoNewRed) + 1)
        6: 'puntoEVI2ReEscalado',
        # Re-escalado de puntoNDVIcalculado = (miPtoNewNir255 - miPtoNewRed255) / sumaNirRed
        7: 'puntoNDVIReEscalado',
        # Re-escalado de puntoNDWIcalculado = (miPtoNewNir255 - miPtoNewRed255) / sumaGreenNir
        8: 'puntoNDWIReEscalado',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # self_aMetricoEVI2Med[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # self_aMetricoEVI2Med[nMetX, nMetY] += int(100 * (2.5 * (float(miPtoNewNir - miPtoNewRed) / (float(miPtoNewNir) + (2.4 * float(miPtoNewRed)) + 1))))
        9: 'metroEVI2ReEscalado',
        # self_aMetricoNDVIMed[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # self_aMetricoNDVIMed[nMetX, nMetY] += int(100 * (float(miPtoNewNir - miPtoNewRed) / (float(miPtoNewNir) + float(miPtoNewRed))))
        10: 'metroNDVIReEscalado',
        # self_aMetricoNDWIMed[metrX, metrY] -> clidnv0->numbaMainVuelta0{}
        # self_aMetricoNDWIMed[nMetX, nMetY] += int(100 * ((miPtoNewGreen - miPtoNewNir) / (float(miPtoNewGreen) + float(miPtoNewNir))))
        11: 'metroNDWIReEscalado',

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> No uso estas variables
        # Esta variable es menor de 100 cuando no es lastReturn (% rpto a un total de retornos del punto) y
        # superior a 100 si es last return, tanto mayor cuantos mas retornos haya (120->~200)
        # miPtoRetNumRel = (100.0 * miPtoReturnNumber / miPtoReturnTotal) + (20 * numLastReturn)
        # Multiplico el numero total de retornos x 50 para ampliar el rango hasta 250 (en caso de 5 retornos)
        12: 'retNumRel250',
        # Numero total de retornos x 50: miPtoReturnTotal * 50
        13: 'retNumTotX50',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        14: 'miPtoZdecametros', # Cota del punto en decametros

        # Estas variables se calculan en clidnv1.procesaCeldasVuelta1b<>->clidnv1.buscarGuardarPuntosMiniSubCel<> 
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        # mseMx50MicroPlanoNubePuntual = miPtoNpExtrVar['mseMx50MicroPlanoNubePuntual']
        # 15: 'mseMx50MicroPlanoNubePuntual',
        # mseCmMicroPlanoNubePuntual = miPtoNpExtrVar['mseMx50MicroPlanoNubePuntual']
        15: 'mseCmMicroPlanoNubePuntual',
        # ptePctjMicroPlanoNubePuntual = miPtoNpExtrVar['ptePctjMicroPlanoNubePuntual']
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        16: 'ptePctjMicroPlanoNubePuntual',
        # cotaRelDmMinNubePuntual = miPtoNpExtrVar['cotaRelDmMinNubePuntual']
        # Es la diferencia (abs) de cotas entre este punto y el mas bajo de la nube puntual calculado
        #  en el mini enjambre de puntos con radio: GLBLradioNubeDeCadaPunto y un maximo de puntos: GLBLmaxPtNubeDeCadaPunto)
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        17: 'cotaRelDmMinNubePuntual',
        # cotaRelDmPlanoNubePuntual = miPtoNpExtrVar['cotaRelDmPlanoNubePuntual']
        # Es la diferencia (abs) de cotas entre este punto y ?el mas alto? de la nube puntual calculado
        #  en el mini enjambre de puntos con radio: GLBLradioNubeDeCadaPunto y un maximo de puntos: GLBLmaxPtNubeDeCadaPunto)
        # Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv1.buscarGuardarPuntosMiniSubCel{}
        18: 'cotaRelDmPlanoNubePuntual',

        19: 'random0a255_1',
        20: 'random0a255_2',

        # ///////////////////////////////////////////////// variables opcionales-> si GLBLcalcularHiperFormas
        # Fuente: self_aMetricoRugosidadMacroInterCeldillas[metrX, metrY] -> calculado en clidnv4.procesaCeldasVuelta4{}
        21: 'metricoRugoMacroReEscalado',
        # self_aMetricoRugosidadMesosInterCeldillas[metrX, metrY] -> clidnv4.procesaCeldasVuelta4{}
        22: 'metricoRugoMesosReEscalado',
        # self_aMetricoRugosidadMicroInterCeldillas[metrX, metrY] -> clidnv4.procesaCeldasVuelta4{}
        23: 'metricoRugoMicroReEscalado',
        # self_aSubCeldasRugosidadMacroInterCeldillas -> clidnv4.procesaCeldasVuelta4{}
        24: 'subCeldasRugoMacroReEscalado',
        # self_aSubCeldasRugosidadMesosInterCeldillas -> clidnv4.procesaCeldasVuelta4{}
        25: 'subCeldasRugoMesosReEscalado',
        # self_aSubCeldasRugosidadMicroInterCeldillas -> clidnv4.procesaCeldasVuelta4{}
        26: 'subCeldasRugoMicroReEscalado',
        # self_aMetricoPlanoTejado[metrX, metrY] -> clidnv4.procesaCeldasVuelta4{}
        27: 'metricoPlanoTejadoReEscalado',
        # //////////////////////////////////////////////////////////////////////

        # >>>>>>>>>>>>>>>>>>>>> Calculadas a partir de MDS cartolid aproximativo
        # Esta variable la uso si GLBLcalcularMdb and GLBLincluirVarDeMdbEnElModeloAcumulativo
        # CotaEnDmPlus20SobreMdb calculada a partir de cotaSobreMdbEnMetros, calculada a partir de cotaCmSobreMdb32bits
        # cotaCmSobreMdb32bits -> calculado a partir de self_aCeldasCoeficientesMdb_[] con las coordenadas de punto de aCeldasListaDePtosTlcPralPF99[]
        28: 'CotaEnDmPlus20SobreMdb',
        # Esta variable la uso si GLBLcalcularMdp and GLBLincluirVarDeMdpEnElModeloAcumulativo
        # CotaEnDmPlus20SobreMdf calculada a partir de cotaSobreMdfEnMetrosElegida, calculada a partir de cotaCmSobreMdf16bitsElegida
        # cotaCmSobreMdf16bitsElegida = aCeldasListaDePtosTlcPralPF99['cotaCmSobreMdfManual16bits'] o aCeldasListaDePtosTlcPralPF99['cotaCmSobreMdfConvol16bits'] o aCeldasListaDePtosTlcPralPF99['cotaCmSobreMdfConual16bits']
        # Extraido de aCeldasListaDePtosTlcPralPF99 = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        29: 'CotaEnDmPlus20SobreMdf',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> No uso estas variables
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Dependientes del modelo convolucional
        # El miniSubcel de la celda en la que cae el punto esta clasificado como suelo segun el modelo convolucional 
        30: 'esSueloMiniSubCel',
        # El miniSubcel de la celda en la que cae el punto esta clasificado como vegetacion segun el modelo convolucional 
        31: 'esVegetaMiniSubCel',
        # El miniSubcel de la celda en la que cae el punto esta clasificado como edificio segun el modelo convolucional 
        32: 'esEdifiMiniSubCel',

        # Esta en una subcelda clasificada como edificio segun el modelo convolucional entrenado con cartoSingu
        33: 'esEdiSubCeldaUsoSingularPredicho',
        # Esta en una subcelda clasificada como via segun el modelo convolucional entrenado con cartoSingu
        34: 'esViaSubCeldaUsoSingularPredicho',
        # Esta en una subcelda clasificada como puente segun el modelo convolucional entrenado con cartoSingu
        35: 'esPteSubCeldaUsoSingularPredicho',

        # Esta en una subcelda clasificada como agua segun el modelo convolucional entrenado con cartoSingu
        100: 'esAguSubCeldaUsoSingularPredicho',
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        # Distancia en x2cm hasta el punto miniSubCel mas cercano (en dm x5 y truncada a 255)
        36: 'distDmX5HastaMiniSubCel0',
        # Cota en dm sobre el punto miniSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        37: 'cotaDmPlus20SobreMiniSubCel0',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        38: 'pteReEscaladaSobreMiniSubCel0',
        # Es suelo el punto miniSubCel mas cercano
        39: 'esSueMiniSubCel0',
        # Es vegetacion el punto miniSubCel mas cercano
        40: 'esVegMiniSubCel0',
        # Es edificio el punto miniSubCel mas cercano
        41: 'esEdiMiniSubCel0',
        # Es agua el punto miniSubCel mas cercano
        42: 'esAguMiniSubCel0',
        # Es via de comunicacion el punto miniSubCel mas cercano
        43: 'esViaMiniSubCel0',

        # Distancia en x2cm hasta el segundo punto miniSubCel mas cercano (en dm x5 y truncada a 255)
        44: 'distDmX5HastaMiniSubCel1',
        # Cota en dm sobre el segundo punto miniSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        45: 'cotaDmPlus20SobreMiniSubCel1',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        46: 'pteReEscaladaSobreMiniSubCel1',
        # LasClass original del segundo punto miniSubCel mas cercano
        47: 'esSueMiniSubCel1',
        # LasClass predicha convolucionalmente del segundo punto miniSubCel mas cercano
        48: 'esVegMiniSubCel1',
        # Es edificio el punto miniSubCel mas cercano
        49: 'esEdiMiniSubCel1',
        # Es agua el punto miniSubCel mas cercano
        50: 'esAguMiniSubCel1',
        # Es via de comunicacion el punto miniSubCel mas cercano
        51: 'esViaMiniSubCel1',

        # Distancia en x2cm hasta el tercer punto miniSubCel mas cercano (en dm x5 y truncada a 255)
        52: 'distDmX5HastaMiniSubCel2',
        # Cota en dm sobre el tercer punto miniSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        53: 'cotaDmPlus20SobreMiniSubCel2',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        54: 'pteReEscaladaSobreMiniSubCel2',
        # LasClass original del tercer punto miniSubCel mas cercano
        55: 'esSueMiniSubCel2',
        # LasClass predicha convolucionalmente del tercer punto miniSubCel mas cercano
        56: 'esVegMiniSubCel2',
        # Es edificio el punto miniSubCel mas cercano
        57: 'esEdiMiniSubCel2',
        # Es agua el punto miniSubCel mas cercano
        58: 'esAguMiniSubCel2',
        # Es via de comunicacion el punto miniSubCel mas cercano
        59: 'esViaMiniSubCel2',

        # Distancia en x2cm hasta el punto maxiSubCel mas cercano (en dm x5 y truncada a 255)
        60: 'distDmX5HastaMaxiSubCel0',
        # Cota en dm sobre el punto maxiSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        61: 'cotaDmPlus50BajoMaxiSubCel0',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        62: 'pteReEscaladaBajoMaxiSubCel0',
        # Distancia en x2cm hasta el punto maxiSubCel mas cercano (en dm x5 y truncada a 255)
        63: 'distDmX5HastaMaxiSubCel1',
        # Cota en dm sobre el punto maxiSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        64: 'cotaDmPlus50BajoMaxiSubCel1',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        65: 'pteReEscaladaBajoMaxiSubCel1',
        # Distancia en x2cm hasta el punto maxiSubCel mas cercano (en dm x5 y truncada a 255)
        66: 'distDmX5HastaMaxiSubCel2',
        # Cota en dm sobre el punto maxiSubCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        67: 'cotaDmPlus50BajoMaxiSubCel2',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        68: 'pteReEscaladaBajoMaxiSubCel2',

        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        69: 'distMtX2HastaMiniCel0',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        70: 'cotaDmPlus20SobreMiniCel0',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        71: 'pteReEscaladaSobreMiniCel0',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        72: 'distMtX2HastaMiniCel1',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        73: 'cotaDmPlus20SobreMiniCel1',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        74: 'pteReEscaladaSobreMiniCel1',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        75: 'distMtX2HastaMiniCel2',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        76: 'cotaDmPlus20SobreMiniCel2',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        77: 'pteReEscaladaSobreMiniCel2',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        78: 'distMtX2HastaMiniCel3',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        79: 'cotaDmPlus20SobreMiniCel3',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        80: 'pteReEscaladaSobreMiniCel3',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        81: 'distMtX2HastaMiniCel4',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        82: 'cotaDmPlus20SobreMiniCel4',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        83: 'pteReEscaladaSobreMiniCel4',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        84: 'distMtX2HastaMiniCel5',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        85: 'cotaDmPlus20SobreMiniCel5',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        86: 'pteReEscaladaSobreMiniCel5',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        87: 'distMtX2HastaMiniCel6',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        88: 'cotaDmPlus20SobreMiniCel6',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        89: 'pteReEscaladaSobreMiniCel6',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        90: 'distMtX2HastaMiniCel7',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        91: 'cotaDmPlus20SobreMiniCel7',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        92: 'pteReEscaladaSobreMiniCel7',
        # Distancia en dm hasta el punto miniCel mas cercano (en dm x5 y truncada a 255)
        93: 'distMtX2HastaMiniCel8',
        # Cota en dm sobre el punto miniCel mas cercano (mayorada en 20 dm y truncada al rango 0-255)
        94: 'cotaDmPlus20SobreMiniCel8',
        # 50 + ( 100 x sign( log(ptex10 + 1) ) ); (pte: cota/dist)
        95: 'pteReEscaladaSobreMiniCel8',

        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minAnisotropy = miPtoNpExtrVar['minAnisotropy']
        96: 'minAnisotropy',
        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minPlanarity = miPtoNpExtrVar['minPlanarity']
        97: 'minPlanarity',
        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minSphericity = miPtoNpExtrVar['minSphericity']
        98: 'minSphericity',
        # No lo calculo. Extraido de miPtoNpExtrVar = self_aCeldasListaDePtosExtrVar[] -> clidnv2y.asignaCotaSobreMdfYautoValores{}
        # minLinearity = miPtoNpExtrVar['minLinearity']
        99: 'minLinearity',

        101: 'tSNEx',
        102: 'tSNEy',
        103: 'tSNEz',
        104: 'tSNElabel',
    }

    # ==========================================================================
    # Review of Input Variable Selection Methods for Artificial Neural Networks 
    # https://www.researchgate.net/profile/Mah-Parsa/post/What-is-the-best-approach-to-select-input-variables-in-Artificial-Neural-Networks-ANNs/attachment/59d63c46c49f478072ea7c27/AS%3A273748995837952%401442278254924/download/input+selection.pdf
    # ==========================================================================
    
    # Variables obligadas (0-1-3-4-8)
    if nInputVars == 0 or nInputVars == 1 or nInputVars == 3:
        listaNumerosVariablesRGBI = [] # +0 variables
    elif nInputVars == 2:
        listaNumerosVariablesRGBI = [
            10, # NDVI metrico
        ] # +1 variables
    elif nInputVars in [4, 5, 6, 7, 8]:
        listaNumerosVariablesRGBI = [
            5, # Intensity Single Return
            10, 11, # NDVI, NDWI metricos
        ] # +3 variables
    elif nInputVars == 9:
        listaNumerosVariablesRGBI = [
            0, 1, 2, 3, # RGBI
        ] # +4 variables
    elif nInputVars == 58:
        listaNumerosVariablesRGBI = [
            5, # Intensity Single Return
        ] # +1 variables
    else:
        listaNumerosVariablesRGBI = [
            0, 1, 2, 3, # RGBI
            5, # Intensity Single Return
            9, 10, 11, # EVI2, NDVI, NDWI metricos
        ] # +8 variables

    # Variables obligadas (0-1-2-3-4-7)
    if nInputVars <= 2 or nInputVars == 4:
        listaNumerosVariablesGeometriaBasica = [] # +0 variables
    elif nInputVars == 5:
        listaNumerosVariablesGeometriaBasica = [
            15, # Geometria de la nube puntual
        ] # +1 variables
    elif nInputVars == 3 or nInputVars == 6:
        listaNumerosVariablesGeometriaBasica = [
            15, 18, # Geometria de la nube puntual
        ] # +2 variables
    elif nInputVars == 7:
        listaNumerosVariablesGeometriaBasica = [
            15, 17, 18, # Geometria de la nube puntual
        ] # +3 variables
    elif nInputVars == 8 or nInputVars == 9:
        listaNumerosVariablesGeometriaBasica = [
            15, 16, 17, 18, # Geometria de la nube puntual
        ] # +4 variables
    else:
        listaNumerosVariablesGeometriaBasica = [
            12, 13, # Posicion relativa del retorno
            14, # miPtoZdecametros
            15, 16, 17, 18, # Geometria de la nube puntual
        ] # +7 variables

    if nInputVars <= 9:
        GLO.GLBLincluirVarUsoSingPredichoEnElModeloAcumulativo = False
        GLO.GLBLincluirVarClaseMiniSubCelPredichaEnElModeloAcumulativo = False
        GLO.GLBLincluirVarGeometriaMiniMaxiSubCelEnElModeloAcumulativo = False
        GLO.GLBLincluirVarGeometriaMiniCelEnElModeloAcumulativo = False
        GLO.GLBLincluirVarDeMdbEnElModeloAcumulativo = False
        GLO.GLBLincluirVarDeHiperformasEnElModeloAcumulativo = False

    # Variables opcionales
    if nInputVars == 0:
        # Si entreno un modelo nuevo, lo hago conforme a las variables:
        # GLBLincluirVarUsoSingPredichoEnElModeloAcumulativo
        # GLBLincluirVarClaseMiniSubCelPredichaEnElModeloAcumulativo
        # GLBLincluirVarGeometriaMiniMaxiSubCelEnElModeloAcumulativo
        # GLBLincluirVarGeometriaMiniCelEnElModeloAcumulativo
        # GLBLincluirVarDeMdbEnElModeloAcumulativo
        # GLBLincluirVarDeMdpEnElModeloAcumulativo
        # GLBLincluirVarDeHiperformasEnElModeloAcumulativo
        if GLO.GLBLpredecirCubiertasSingularesConvolucional and GLO.GLBLincluirVarUsoSingPredichoEnElModeloAcumulativo:
            listaNumerosVariablesUsoSingularPredichoConvolucional = [
                33, 34, 35, 100 # Uso singular predicho de la subCel (esEdi, esVia, esPte, esAgu)
            ] # +4 variables
        else:
            listaNumerosVariablesUsoSingularPredichoConvolucional = []
        if GLO.GLBLpredecirClasificaMiniSubCelConvolucional and GLO.GLBLincluirVarClaseMiniSubCelPredichaEnElModeloAcumulativo:
            listaNumerosVariablesMiniSubCelClasePredicha = [
                30, 31, 32, # Uso predicho del miniSubCel (esSue, esVeg, esEdi)
                39, 40, 41, # Uso predicho del miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                47, 48, 49, # Uso predicho del segundo miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                55, 56, 57, # Uso predicho del tercer miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
            ]
        else: # +12 variables
            listaNumerosVariablesMiniSubCelClasePredicha = []
        if GLO.GLBLincluirVarGeometriaMiniMaxiSubCelEnElModeloAcumulativo:
            listaNumerosVariablesMiniMaxiSubCelGeometria = [
                36, 37, 38, # Distancia, cota y pte hasta el punto miniSubCel mas cercano
                44, 45, 46, # Distancia, cota y pte hasta el segundo punto miniSubCel mas cercano
                52, 53, 54, # Distancia, cota y pte hasta el tercer punto miniSubCel mas cercano
                60, 61, 62, # Distancia, cota y pte hasta el punto maxiSubCel mas cercano
                63, 64, 65, # Distancia, cota y pte hasta el segundo punto maxiSubCel mas cercano
                66, 67, 68, # Distancia, cota y pte hasta el tercer punto maxiSubCel mas cercano
            ]
        else: # +18 variables
            listaNumerosVariablesMiniMaxiSubCelGeometria = []
        if GLO.GLBLincluirVarGeometriaMiniCelEnElModeloAcumulativo:
            listaNumerosVariablesMiniCelGeometria = [
                69, 70, 71, # Distancia, cota y pte hasta el punto miniCel 0 mas cercano
                72, 73, 74, # Distancia, cota y pte hasta el punto miniCel 1 mas cercano
                75, 76, 77, # Distancia, cota y pte hasta el punto miniCel 2 mas cercano
                78, 79, 80, # Distancia, cota y pte hasta el punto miniCel 3 mas cercano
                81, 82, 83, # Distancia, cota y pte hasta el punto miniCel 4 mas cercano
                84, 85, 86, # Distancia, cota y pte hasta el punto miniCel 5 mas cercano
                87, 88, 89, # Distancia, cota y pte hasta el punto miniCel 6 mas cercano
                90, 91, 92, # Distancia, cota y pte hasta el punto miniCel 7 mas cercano
                93, 94, 95, # Distancia, cota y pte hasta el punto miniCel 8 mas cercano
            ]
        else: # +27 variables
            listaNumerosVariablesMiniCelGeometria = []
        if GLO.GLBLcalcularMdb and GLO.GLBLincluirVarDeMdbEnElModeloAcumulativo:
            listaNumerosVariablesMdb = [
                28, # Cota sobre Mdb
            ] # +1 variables
        else:
            listaNumerosVariablesMdb = []
        if GLO.GLBLcalcularMdp and GLO.GLBLincluirVarDeMdpEnElModeloAcumulativo:
            listaNumerosVariablesMdp = [
                29, # Cota sobre Mdf
            ] # +1 variables
        else:
            listaNumerosVariablesMdp = []
        if GLO.GLBLcalcularHiperFormas and GLO.GLBLincluirVarDeHiperformasEnElModeloAcumulativo:
            listaNumerosVariablesHiperFormas = [
                21, 22, 23, 24, 25, 26, 27, # Usarlos si los calculo -> Inconveniente: lo que tarda en calcular el hipercubo
            ] # +7 variables
        else:
            listaNumerosVariablesHiperFormas = []

    else:
        # Si entreno un modelo preentrenado, leo el numero de inputVars del nombre
        #  y elijo la lista de variables a la vista del numero
        # Opcion elegida por el momento: 65 variables, que incluye:
        #  Obligatorias RGBI (0-4-8)
        #  Obligatorias GeometriaBasica (2-3-7)
        #  CotaSobreMdb (1)
        #  UsoSingularPredichoConvolucional (4)
        #  MiniMaxiSubCelGeometria (18)
        #  MiniCelGeometria (27)
        # Y excluye:
        #  CotaSobreMdp (1)
        #  HiperFormas (7)
        #  MiniSubCelClasePredicha (12)
        if nInputVars in [
            1, 2, 3, 4, 5, 6, 7, 8, 9,
            20, 21, 38, 39, 58, 65, 66, 72, 73, 84, 85
        ]:
            listaNumerosVariablesMdb = [
                28, # Cota sobre Mdb
            ] # +1 variables
        else:
            listaNumerosVariablesMdb = []
        if nInputVars in [19, 20, 38, 58, 65, 72, 84, 85]:
            listaNumerosVariablesUsoSingularPredichoConvolucional = [
                33, 34, 35, 100 # Uso singular predicho de la subCel (esEdi, esVia, esPte, esAgu)
            ] # +4 variables
        else:
            listaNumerosVariablesUsoSingularPredichoConvolucional = []
        if nInputVars in [21, 39, 66, 73, 85]:
            listaNumerosVariablesMdp = [
                29, # Cota sobre Mdf
            ] # +1 variables
        else:
            listaNumerosVariablesMdp = []
        # Para las siguientes, de cada 3 numeros consecutivos:
        #  El primero excluye cotaSobreMdb y cotaSobreMdp
        #  El segundo incluye cotaSobreMdb y excluye cotaSobreMdp
        #  El tercero incluye cotaSobreMdb y cotaSobreMdp
        # Por el momento uso el segundo para evitar calcular Mdp
        if nInputVars in [37, 38, 39, 58, 64, 65, 66, 71, 72, 73, 83, 84, 85]:
            listaNumerosVariablesMiniMaxiSubCelGeometria = [
                36, 37, 38, # Distancia, cota y pte hasta el punto miniSubCel mas cercano
                44, 45, 46, # Distancia, cota y pte hasta el segundo punto miniSubCel mas cercano
                52, 53, 54, # Distancia, cota y pte hasta el tercer punto miniSubCel mas cercano
                60, 61, 62, # Distancia, cota y pte hasta el punto maxiSubCel mas cercano
                63, 64, 65, # Distancia, cota y pte hasta el segundo punto maxiSubCel mas cercano
                66, 67, 68, # Distancia, cota y pte hasta el tercer punto maxiSubCel mas cercano
            ]
        else: # +18 variables
            listaNumerosVariablesMiniMaxiSubCelGeometria = []
        if nInputVars in [58, 64, 65, 66, 71, 72, 73, 83, 84, 85]:
            listaNumerosVariablesMiniCelGeometria = [
                69, 70, 71, # Distancia, cota y pte hasta el punto miniCel 0 mas cercano
                72, 73, 74, # Distancia, cota y pte hasta el punto miniCel 1 mas cercano
                75, 76, 77, # Distancia, cota y pte hasta el punto miniCel 2 mas cercano
                78, 79, 80, # Distancia, cota y pte hasta el punto miniCel 3 mas cercano
                81, 82, 83, # Distancia, cota y pte hasta el punto miniCel 4 mas cercano
                84, 85, 86, # Distancia, cota y pte hasta el punto miniCel 5 mas cercano
                87, 88, 89, # Distancia, cota y pte hasta el punto miniCel 6 mas cercano
                90, 91, 92, # Distancia, cota y pte hasta el punto miniCel 7 mas cercano
                93, 94, 95, # Distancia, cota y pte hasta el punto miniCel 8 mas cercano
            ]
        else: # +27 variables
            listaNumerosVariablesMiniCelGeometria = []
        if nInputVars in [71, 72, 73, 83, 84, 85]:
            listaNumerosVariablesHiperFormas = [
                21, 22, 23, 24, 25, 26, 27, # Usarlos si los calculo -> Inconveniente: lo que tarda en calcular el hipercubo
            ] # +7 variables
        else:
            listaNumerosVariablesHiperFormas = []
        if nInputVars in [83, 84, 85]:
            listaNumerosVariablesMiniSubCelClasePredicha = [
                30, 31, 32, # Uso predicho del miniSubCel (esSue, esVeg, esEdi)
                39, 40, 41, # Uso predicho del miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                47, 48, 49, # Uso predicho del segundo miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
                55, 56, 57, # Uso predicho del tercer miniSubCel mas cercano (esSue, esVeg, esEdi, esAgu, esVia)
            ]
        else: # +12 variables
            listaNumerosVariablesMiniSubCelClasePredicha = []

    # Variables obligatorias: 15

    # Variables descartadas (20 variables):
    # listaNumerosVariablesMiniSubCelClaseNoPredicha = [
    #     42, 43, # Uso predicho del miniSubCel mas cercano (esAgu, esVia)
    #     50, 51, # Uso predicho del segundo miniSubCel mas cercano (esAgu, esVia)
    #     58, 59, # Uso predicho del tercer miniSubCel mas cercano (esAgu, esVia)
    # ] # +6 variables
    # listaNumerosVariablesDescartados = [
    #     4, # intensity
    #     6, 7, 8, # puntoEVI2ReEscalado, puntoNDVIReEscalado, puntoNDWIReEscalado
    #     19, 20, # Disponibles (random)
    #     96, 97, 98, 99, # anisotropy
    #     101, 102, 103, 104, # tSNE
    # ] # +14 variables
    listaNumerosVariablesForTrain = (
        listaNumerosVariablesRGBI # (0-3-4-8) variables
        + listaNumerosVariablesGeometriaBasica # (1-2-3-4-7) variables
        + listaNumerosVariablesMdb # 1 variables
        + listaNumerosVariablesUsoSingularPredichoConvolucional # 0-4 variables
        + listaNumerosVariablesMiniMaxiSubCelGeometria # 0-18 variables
        + listaNumerosVariablesMiniCelGeometria # 0-27 variables
        + listaNumerosVariablesHiperFormas # 0-7 variables
        + listaNumerosVariablesMiniSubCelClasePredicha # 0-12 variables
        + listaNumerosVariablesMdp # 0-1 variables
    ) # maximo: 85 variables
    listaNumerosVariablesForTrain.sort()
    print('clidconfig-> Numero de Variables seleccionadas forTrain: {}'.format(nInputVars))
    print('\t-> listaNumerosVariablesForTrain: {}->{}'.format(len(listaNumerosVariablesForTrain), listaNumerosVariablesForTrain))
    print(f'{"":=^80}')

    return dictVariablesEndogenasTodas, listaNumerosVariablesForTrain


def foo1():
    pass

# ==============================================================================
# Dejo esto por si quiero ver como funciona la pila de llamadas segun desde donde se cargue este modulo
if False:
    mostrarModuloInicial = True
    if mostrarModuloInicial:
        if len(inspect.stack()) > 1:
            try:
                callingModuleActual = inspect.getmodulename(inspect.stack()[0][1])
                callingModuleInicial = inspect.getmodulename(inspect.stack()[-1][1])
                callingModulePrevio = 'desconocido'
                for nOrden, llamada in enumerate(inspect.stack()[1:]):
                    #print('\t->', nOrden, llamada[1])
                    if not llamada[1].startswith('<frozen'):
                        callingModulePrevio = inspect.getmodulename(llamada[1])
                        break
            except:
                print('Error al identificar el modulo')
        else:
            callingModuleActual = 'Este modulo'
            callingModuleInicial = 'Este modulo'
            callingModulePrevio = 'Modulo no importado desde otro modulo'
        # print('clidconfig->  Este modulo:                   ', callingModuleActual)
        # print('clidconfig->  Modulo desde el que se importa:', callingModulePrevio)
        # print('clidconfig->  Modulo ejecutado inicialmente: ', callingModuleInicial)
# ==============================================================================


def foo2():
    pass

# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
if callingModulePrevio == 'clidbase' or callingModuleInicial == 'clidclas': # or callingModuleInicial == 'clidtry':
    LCLverbose = True
    print(f'\nclidconfig-> Este modulo se importa desde todos los modulos,')
    print(f'{TB}pero solo se muestra la carga en pantalla cuando:')
    print(f'{TB}-> Se importa desde clidbase.')
    print(f'{TB}-> El modulo lanzado inicialmente es clidclas.')
else:
    LCLverbose = False



# ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo
if (
    callingModuleInicial == 'clidbase'
    or callingModuleInicial == 'clidflow'
    or callingModuleInicial == 'clidtools'
    or callingModuleInicial == 'clidclas'
    or callingModuleInicial == 'clidtry'
    or callingModuleInicial == 'clidgis'
):
    TRNSpreguntarMemoria = False
    TRNSseleccionadosParaRepetir = 2

    controlFileLasName = None
    controlFileLas = None
    controlFileGral = None
    controlFileGlobal = None

    callingModulePrevio, callingModuleInicial = showCallingModules(inspect_stack=inspect.stack(), verbose=False)
    if CONFIGverbose:
        print(f'clidconfig-> Modulo desde el que se importa clidconfig: {callingModulePrevio}')
        print(f'{TB}-> callingModuleInicial: {callingModuleInicial}')

    if CONFIGverbose:
        print('''
clidconfig-> Secuencia de carga de variables de configuracion:
             1. Se lanza initConfigDicts<> para:
                 a. Buscar y, si existe, leer el fichero cfg
                    -> Si hay fichero cfg de anteriores ejecucuiones
                       Se leen los parametros del cfg y se guardan en GLOBALconfigDict
                    -> Si no hay fichero cfg se pasa a la opcion b
                       (se tira del fichero xlsx de configuracion).
                 b. Buscar y leer el fichero xlsx
                    Esta funcion devuelve la variable GLOBALconfigDict (globales de clidconfig)
                    Ademas hace una copia del xlsx que dejo solo como testimonio (no la uso)
                    AVISO: aqui se eligen las columnas a usar para las configuraciones extra
                      Por el momento solo la columna I, con valor CREA_TILES (eso queda guardado en GLOBALconfigDict)
             2. Creo el fichero de configuracion cfg correspondiente con guardarVariablesGlobales<> a partir de GLOBALconfigDict.
                De esta forma esta disponible para todos los modulos 
                  Lo puedo actualizar si hace falta y esta disponible para todos,
                  cosa que es mas dificil con GLO salvo que fuerce la recarga de clidconfig desde cada modulo.
                  En principio solo se carga la primera vez que lo llamo
                  (cosa que ocurre desde clidaux y no desde clidbase, que llama primero a clidaux).
                Incluyo todas las configuraciones extra, ademas de la general.''')
    LCL_idProceso=MAIN_idProceso
    # Leo los parametros de configuracion de clidbase.xml y los cargo en diccionarios
    if CONFIGverbose:
        print('clidconfig-> Para leer el fichero de configuracion xlsx lanzo initConfigDicts<> (se hace copia con LCL_idProceso: {})'.format(LCL_idProceso))
    GLOBALconfigDict = initConfigDicts(LCL_idProceso, LOCL_verbose=__verbose__)
    if CONFIGverbose:
        print(f'clidconfig-> Resultado de lanzar initConfigDicts<> GLOBALconfigDict["GLBLverbose"]: {len(GLOBALconfigDict["GLBLverbose"])}, {GLOBALconfigDict["GLBLverbose"]}')

    if CONFIGverbose:
        print(f'\nclidconfig-> Valores tras leer el fichero de configuracion xls')
        print(f'{TB}-> MAINprocedimiento1: {GLOBALconfigDict["MAINprocedimiento"][0]}')
        print(f'{TB}-> MAINcuadrante1: {GLOBALconfigDict["MAINcuadrante"][0]}')

    # Creo el objeto GLO (de la clase VariablesGlobales) y le asigno los paramConfig como propiedades
    #  Uso el parametro MAINobjetivoEjecucion (CREAR_TILES_TRAIN) para elegir la columna del excel que le corresponde (objetivoEjecucion)
    # Ademas reviso los paramConfig cambiando o ampliando los que proceda
    if CONFIGverbose:
        print(f'clidconfig-> Voy a crear el objeto GLO (objeto global de clidconfig de la clase VariablesGlobales<>)')
        print(f'\t-> al que puedo acceder desde el resto de los modulos importando este modulo')

    # ==========================================================================
    GLO = VariablesGlobales(GLOBALconfigDict, LCLverbose=__verbose__)
    # ==========================================================================

    if CONFIGverbose:
        print(f'clidconfig-> #4a GLO.GLBLverbose:                   {GLO.GLBLverbose}')

    # Esto solo es necesario la primera vez que creo esta clase (desde clidaux)
    # GLO.adaptarVariablesGlobales()
    if CONFIGverbose:
        print(f'clidconfig-> Revisando la coherencia y validez de la configuracion')
    GLO.revisarCompletarVariablesMAINdelConfigVarsDict()
    ok = GLO.revisarCompletarVariablesGLBLdelConfigVarsDict()
    if not ok:
        print(f'clidconfig-> Configuracion no permitida: cambiarla')
        sys.exit(0)

    if CONFIGverbose:
        print(f'clidconfig-> MAINprocedimiento2: {GLOBALconfigDict["MAINprocedimiento"][0]}')
        print(f'clidconfig-> MAINprocedimiento2: {GLO.MAINprocedimiento}')

    # Vuelvo a guardarlos en el fichero clidbase.cfg:
    if CONFIGverbose:
        print(f'clidconfig-> paramConfig tras cargarlos en GLO-> GLOBALconfigDict[GLBLverbose]: {len(GLOBALconfigDict["GLBLverbose"])}, {GLOBALconfigDict["GLBLverbose"]}')
        # print('clidconfig-> sys.argv:', len(sys.argv), '->', sys.argv)
        print('clidconfig-> Guardo los paramConfig en el fichero cfg')
    guardarVariablesGlobales(
        GLOBALconfigDict,
        LCL_idProceso=MAIN_idProceso,
    )

    # Puedo leer los paramtros de configuracion desde cualquier funcion con:
    # GLOBALconfigDict = leerCambiarVariablesGlobales(LCL_idProceso=MAIN_idProceso)
    # Tambien puedo cambiar los paramtros que quiera de forma permanente o agregar nuevos.
    # Ejemplo de como cambiar o agregar parametros de configuracion
    # print('clidconfig->\n\t1.- GLOBALconfigDict', GLOBALconfigDict)
    # nuevosParametroConfiguracion={}
    # nuevosParametroConfiguracion['MAINusarValorALTER'] = [True, 'GrupoMAIN', 'Nueva descripcion 1', 'bool']
    # GLOBALconfigDict = leerCambiarVariablesGlobales(LCL_idProceso=MAIN_idProceso, nuevosParametroConfiguracion=nuevosParametroConfiguracion)
    # print('clidconfig->\n\t2.- GLOBALconfigDict', GLOBALconfigDict)

else:
    # print('clidconfig-> No se cargan las variables globales de clidbase')
    pass

