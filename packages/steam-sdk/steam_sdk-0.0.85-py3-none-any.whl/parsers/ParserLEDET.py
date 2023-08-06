import numpy as np
import datetime
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from builders.BuilderLEDET import BuilderLEDET


class ParserLEDET:
    """
        Class with methods to read/write LEDET information from/to other programs
    """

    def __init__(self, builder_ledet: BuilderLEDET = BuilderLEDET(flag_build=False)):
        """
            Initialization using a BuilderLEDET object containing LEDET parameter structure
        """

        self.builder_ledet: BuilderLEDET = builder_ledet

    def readFromExcel(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads an Excel file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        ##File must be whole eos string
        workbookVariables = openpyxl.load_workbook(file_name, data_only=True)

        #Inputs
        worksheetInputs = workbookVariables['Inputs']
        lastAttribute = worksheetInputs.cell(1, 2).value
        for i in range(1, worksheetInputs.max_row+1):
            # builder_ledet.variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            attribute = worksheetInputs.cell(i, 2).value
            try:
                if (attribute == None):
                    if worksheetInputs.cell(i, 3).value is not None:
                        values = self.read_row(worksheetInputs, i)
                        values = np.array([k for k in values if(str(k))])
                        current = builder_ledet.getAttribute(builder_ledet.Inputs, lastAttribute)
                        current = np.vstack((current, values))
                        builder_ledet.setAttribute(builder_ledet.Inputs, lastAttribute, current)
                    else:
                        continue
                elif (type(builder_ledet.getAttribute(builder_ledet.Inputs, attribute)) == np.ndarray):
                    lastAttribute = attribute
                    values = self.read_row(worksheetInputs, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Inputs, attribute, values)
                else:
                    value = worksheetInputs.cell(i, 3).value
                    builder_ledet.setAttribute(builder_ledet.Inputs, attribute, value)
            except TypeError as e:
                if attribute in builder_ledet.Inputs.__annotations__: raise e
                if attribute=='None' or attribute==None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        #Options
        worksheetOptions = workbookVariables['Options']
        for i in range(1, worksheetOptions.max_row+1):
            # builder_ledet.variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            attribute = worksheetOptions.cell(i, 2).value
            try:
                if (type(builder_ledet.getAttribute(builder_ledet.Options, attribute)) == np.ndarray):
                    values = self.read_row(worksheetOptions, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Options, attribute, values)
                else:
                    value = worksheetOptions.cell(i, 3).value
                    builder_ledet.setAttribute(builder_ledet.Options, attribute, value)
            except TypeError as e:
                if attribute in builder_ledet.Options.__annotations__: raise e
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        #Plots
        worksheetPlots = workbookVariables['Plots']
        for i in range(1, worksheetPlots.max_row+1):
            # builder_ledet.variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            attribute = worksheetPlots.cell(i, 2).value
            try:
                if (type(builder_ledet.getAttribute(builder_ledet.Plots, attribute)) == np.ndarray):
                    values = self.read_row(worksheetPlots, i, St=True)[2:]
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Plots, attribute, values)
                else:
                    try:
                        value = worksheetPlots.cell(i, 3).value
                    except:
                        value = ''
                    builder_ledet.setAttribute(builder_ledet.Plots, attribute, value)
            except TypeError as e:
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        # Variables
        try:
            worksheetVariables = workbookVariables['Variables']
            for i in range(1, worksheetVariables.max_row+1):
                # builder_ledet.variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
                attribute = worksheetVariables.cell(i, 2).value
                try:
                    if (type(builder_ledet.getAttribute(builder_ledet.Variables, attribute)) == np.ndarray):
                        if attribute != 'typeVariableToSaveTxt':  values = self.read_row(worksheetVariables, i, St = True)[2:]
                        else:  values = self.read_row(worksheetVariables, i)
                        values = np.array([k for k in values if(str(k))])
                        builder_ledet.setAttribute(builder_ledet.Variables, attribute, values)
                    else:
                        value = worksheetVariables.cell(i, 3).value
                        builder_ledet.setAttribute(builder_ledet.Variables, attribute, value)
                except TypeError as e:
                    if attribute in builder_ledet.Variables.__annotations__: raise e
                    if attribute == 'None' or attribute == None: continue
                    if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        except:
            pass
            print("Error while reading Variables. Please check!")

        return builder_ledet

    def read_row(self, workSheet, Nrow, St = False):
        '''
            *** Function that reads a row of a worksheet of an Excel file ***

            :param workSheet: Name of the worksheet
            :type workSheet: str
            :param Nrow: Number of the row
            :type Nrow: int
            :param St: Flag indicating whether the entry to read is a string
            :type St: bool

            :return: np.array([])
        '''
        rowValues = np.array([])
        row = workSheet[Nrow]
        for cell in row:
            if not St:
                if isinstance(cell.value, str): continue
            rowValues = np.append(rowValues, cell.value)
        rowValues = rowValues[rowValues != None]
        return rowValues

    def write2Excel(self, nameFileLEDET: str, verbose: bool = True, SkipConsistencyCheck: bool = False):
        '''
            *** Function that writes a LEDET input file as an Excel file ***

            Function to write a LEDET input file composed of "Inputs", "Options", "Plots", and "Variables" sheets

            :param nameFileLEDET: String defining the name of the LEDET input file to be written
            :type nameFileLEDET: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool
            :param SkipConsistencyCheck: flag that determines, whether the parameters shall be checked for consistency or not [False = Apply checks, True = Skip checks]
            :type SkipConsistencyCheck: bool

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        if not SkipConsistencyCheck:
            if self._consistencyCheckLEDET():
                print("Variables are not consistent! Writing aborted - ", nameFileLEDET)
                return
            else:
                if verbose: print("Preliminary consistency check was successful! - ", nameFileLEDET)
        else:
            print("Skipped consistency checks.")

        workbook = openpyxl.Workbook()
        workbook.properties.creator = 'STEAM-Team'

        if verbose:
            print('')
            print('### Write "Variables" sheet ###')
        self.writeWorkbook(workbook, "Variables", builder_ledet.Variables, builder_ledet.descriptionsVariables)

        if verbose:
            print('')
            print('### Write "Plots" sheet ###')
        self.writeWorkbook(workbook, "Plots", builder_ledet.Plots, builder_ledet.descriptionsPlots)

        if verbose:
            print('')
            print('### Write "Options" sheet ###')
        self.writeWorkbook(workbook, "Options", builder_ledet.Options, builder_ledet.descriptionsOptions)

        # Special case: if cooling 0 int, expand 0 to all half-turns in the written file. If done, will revert afterwards
        self._expandCoolingScalarToArray(verbose)
        if verbose:
            print('')
            print('### Write "Inputs" sheet ###')
        self.writeWorkbook(workbook, "Inputs", builder_ledet.Inputs, builder_ledet.descriptionsInputs)

        # Save the workbook
        std = workbook['Sheet']
        workbook.remove(std)

        for s in range(len(workbook.sheetnames)):
            if workbook.sheetnames[s] == 'Inputs': break
        workbook.active = s
        workbook.save(nameFileLEDET)

        # Display time stamp and end run
        currentDT = datetime.datetime.now()
        if verbose:
            print(' ')
            print('Time stamp: ' + str(currentDT))
            print('New file ' + nameFileLEDET + ' generated.')
        return

    def writeWorkbook(self, book, sheet, variableGroup, variableLabels):
        """
            **Write one sheet of a LEDET input file**

            Function writes one sheet of a LEDET input file

            :param book: workbook object to write
            :type book: openpyxl.Workbook
            :param sheet: name of the sheet to write
            :type sheet: string
            :param variableGroup: dataclass containing the attributes and values to be written into sheet
            :type variableGroup: dataclass
            :param variableLabels: dictionary assigning a description to each variable name
            :type variableLabels: dict
            :return:
        """

        # Unpack variables
        builder_ledet = self.builder_ledet

        # Define optional variables, which will not be written in the LEDET input file in case their value is not defined (empty list or array)
        optional_variables = (
            'alpha_Nb3Sn0_inGroup',
            'f_scaling_Jc_BSCCO2212_inGroup'
            )

        # Define cell style
        def styled_cells(data):
            for c in data:
                c = Cell(sheet1, column="A", row=1, value=c)
                c.font = Font(size=14,bold=True)
                yield c

        # Loop through sheets to activate the correct sheet to write to
        book.create_sheet(index = 1 , title = sheet)
        for s in range(len(book.sheetnames)):
            if book.sheetnames[s] == sheet: break
        book.active = s
        sheet1 = book.active

        # Correct section titles, if Helium option is present
        if "overwrite_f_externalVoids_inGroup" in variableGroup.__annotations__:
            ofiVg = builder_ledet.getAttribute(variableGroup, "overwrite_f_externalVoids_inGroup")
            if len(ofiVg) != 0 and "fitParameters_inGroup" in builder_ledet.sectionTitles.keys():
                builder_ledet.sectionTitles["overwrite_f_externalVoids_inGroup"] = builder_ledet.sectionTitles["fitParameters_inGroup"]
                del builder_ledet.sectionTitles["fitParameters_inGroup"]

        # Write to the sheet of the workbook
        # Loop through all attributes in the given variableGroup
        for attribute in variableGroup.__annotations__:
            # Only write the helium options if they are set, otherwise skip
            if (attribute == "overwrite_f_internalVoids_inGroup"):
                ofiVg = builder_ledet.getAttribute(variableGroup, attribute)
                if len(ofiVg) == 0: continue
            if (attribute == "overwrite_f_externalVoids_inGroup"):
                ofiVg = builder_ledet.getAttribute(variableGroup, attribute)
                if len(ofiVg) == 0: continue

            # Skip optional variables if they have 0 elements
            if attribute in optional_variables:
                if isinstance(builder_ledet.getAttribute(variableGroup, attribute), list) and len(builder_ledet.getAttribute(variableGroup, attribute)) == 0:
                    print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(attribute))
                    continue
                elif isinstance(builder_ledet.getAttribute(variableGroup, attribute), np.ndarray) and builder_ledet.getAttribute(variableGroup, attribute).shape[0] == 0:
                    print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(attribute))
                    continue

            # Skip writing conductor resistance fraction if not enabled
            lookup = ['f_RRR1_Cu_inGroup', 'f_RRR2_Cu_inGroup', 'f_RRR3_Cu_inGroup',
                      'RRR1_Cu_inGroup', 'RRR2_Cu_inGroup', 'RRR3_Cu_inGroup']
            if attribute in lookup and not builder_ledet.enableConductorResistanceFraction:
                continue

            # Check if size of list is < 16382 (max size of xlsx - 2 rows for descriptions), if so: convert to np.ndarray
            if isinstance(builder_ledet.getAttribute(variableGroup, attribute), list):
                if len(builder_ledet.getAttribute(variableGroup, attribute)) > 16382:
                    builder_ledet.setAttribute(variableGroup, attribute, np.array(builder_ledet.getAttribute(variableGroup, attribute)).reshape(-1, 1))
            if isinstance(builder_ledet.getAttribute(variableGroup, attribute), np.ndarray):
                if builder_ledet.getAttribute(variableGroup, attribute).shape[0] > 16382:
                    builder_ledet.setAttribute(variableGroup, attribute, np.array(builder_ledet.getAttribute(variableGroup, attribute)).reshape(-1, 1))

            # Actual writing process. Check which datatype the attribute is and append it to the sheet
            # If datatype is matrix, a for loop executes all rows/columns before continuing to next attribute
            varDesc = variableLabels.get(str(attribute))
            if isinstance(builder_ledet.getAttribute(variableGroup, attribute), np.ndarray):
                if builder_ledet.getAttribute(variableGroup, attribute).ndim > 1:
                    for i in range(builder_ledet.getAttribute(variableGroup, attribute).shape[0]):
                        values = builder_ledet.getAttribute(variableGroup, attribute)[i,:].tolist()
                        if i == 0: sheet1.append([varDesc, attribute] + values)
                        else: sheet1.append([None, None] + values)
                else:
                    values = np.array(builder_ledet.getAttribute(variableGroup, attribute)).tolist()
                    sheet1.append([varDesc, attribute] + values)
            elif isinstance(builder_ledet.getAttribute(variableGroup, attribute), list):
                values = builder_ledet.getAttribute(variableGroup, attribute)
                sheet1.append([varDesc, attribute] + values)
            else:
                values = [builder_ledet.getAttribute(variableGroup, attribute)]
                sheet1.append([varDesc, attribute]+ values)

            if attribute in builder_ledet.sectionTitles.keys():
                sheet1.append([None])
                sheet1.append(styled_cells([builder_ledet.sectionTitles[attribute]]))

            # Setting the width of each cells in the workbook [only for good view]
            width = [80.7109375, 40.7109375, 20.7109375]
            sheet1.column_dimensions['A'].width = width[0]
            sheet1.column_dimensions['B'].width = width[1]

            if sheet1.max_column+1 > 18278:
                smc = 18278
            else:
                smc = sheet1.max_column+1
            for i in range(3,smc):
                cl = get_column_letter(i)
                sheet1.column_dimensions[cl].width = width[2]

    def _expandCoolingScalarToArray(self, verbose: bool = False):
        ''' Method adds as many elements as the number of half-turns to selected variables '''
        # TODO: Method can be made more general

        if isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_down, int) or isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_down, float) or \
                len(self.builder_ledet.Inputs.sim3D_f_cooling_down) == 1:
            if verbose: print('sim3D_f_cooling_down is a scalar. Expanding value to all half-turns')
            self.builder_ledet.Inputs.sim3D_f_cooling_down = np.array([self.builder_ledet.Inputs.sim3D_f_cooling_down] * int(np.sum(self.builder_ledet.Inputs.nT)))

        if isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_up, int) or isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_up, float) or \
                len(self.builder_ledet.Inputs.sim3D_f_cooling_up) == 1:
            if verbose: print('sim3D_f_cooling_up is a scalar. Expanding value to all half-turns')
            self.builder_ledet.Inputs.sim3D_f_cooling_up = np.array([self.builder_ledet.Inputs.sim3D_f_cooling_up] * int(np.sum(self.builder_ledet.Inputs.nT)))

        if isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_left, int) or isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_left, float) or \
                len(self.builder_ledet.Inputs.sim3D_f_cooling_left) == 1:
            if verbose: print('sim3D_f_cooling_left is a scalar. Expanding value to all half-turns')
            self.builder_ledet.Inputs.sim3D_f_cooling_left = np.array([self.builder_ledet.Inputs.sim3D_f_cooling_left] * int(np.sum(self.builder_ledet.Inputs.nT)))

        if isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_right, int) or isinstance(self.builder_ledet.Inputs.sim3D_f_cooling_right, float) or \
                len(self.builder_ledet.Inputs.sim3D_f_cooling_right) == 1:
            if verbose: print('sim3D_f_cooling_right is a scalar. Expanding value to all half-turns')
            self.builder_ledet.Inputs.sim3D_f_cooling_right = np.array([self.builder_ledet.Inputs.sim3D_f_cooling_right] * int(np.sum(self.builder_ledet.Inputs.nT)))

    def copy_map2d(self, verbose: bool = False):
        ''' #TODO '''
        # rename and copy
        # self.builder_ledet.Options.flagIron
        # self.builder_ledet.Options.flagSelfField
        # add suffix like LEDET-Manual like in OptionsLEDET of .yaml-file
        pass

    def copy_modified_map2d_ribbon_cable(self, geoArr, verbose: bool = False):
        ''' #TODO '''
        # call ParserMap2d.modify_map2d_ribbon_cable()
        # self.builder_ledet.Options.flagIron
        # self.builder_ledet.Options.flagSelfField
        # pass values to modify...
        # get new values from modfiy... and write correct file
        # then rename and copy
        pass


###############################################################
# START - Methods for consistency checks of LEDET input files #
# TODO: Update consistency checks of LEDET input files to the latest LEDET features

    def __getNumberOfCoilSections(self):
        '''
            **Consistency check of LEDET Inputs - Helper function**

            Returns the number of CoilSections from Mutual-inductance matrix

            :return: int
        '''
        k = self.Inputs.M_m
        if k.shape == (1,): return k.shape[0]
        try:
            if k.shape[0] != k.shape[1]: print("M_m is not square")
        except:
            print("M_m is not square")
            return -1
        k2 = max(self.Inputs.GroupToCoilSection)
        if k.shape[0] != k2:
            print('M_m matrix does have size: ',k.shape[0], ' but you assign a Coil-Section: ',k2)
            return -1
        return k.shape[0]

    def __checkM_InductanceBlock_m(self, Turns):
        '''
            **Consistency check of LEDET Inputs - Inductance Matrix **

            Check if Inductance matrix is squared. Issues a warning if it is not the size of the number of turns. Returns result as bool

            :param Turns: Number of turns of the LEDET object
            :type arr: int
            :return: bool
        '''
        if type(self.Inputs.M_InductanceBlock_m) != np.ndarray:
            k = np.array(self.Inputs.M_InductanceBlock_m)
        else:
            k = self.Inputs.M_InductanceBlock_m
        ## Account for the option to set the matrix to 0
        if k.shape == (1,):
            return True
        try:
            if k.shape[0] == k.shape[1]:
                if k.shape[0] != Turns:
                    print("M_InductanceBlock_m is squared, but its size unequal to the number of turns")
                return True
        except:
            print("M_InductanceBlock_m is not correct!")
            return False
        print("M_InductanceBlock_m is not correct!")
        return False

    def __checkHeFraction(self, Groups):
        '''
            **Consistency check of LEDET Inputs - Helium options check **

            Check if Helium options are both set and have the correct and same size, returns bool of result

            :param Groups: Number of groups of the LEDET object
            :type arr: int
            :return: bool
        '''
        k = self.Inputs.overwrite_f_externalVoids_inGroup
        k2 = self.Inputs.overwrite_f_internalVoids_inGroup
        if len(k) > 0:
            if len(k) != len(k2):
                print("Helium section was set but is corrupted.")
                return False
            if len(k) != Groups:
                print("Helium section was set but is wrong length.")
                return False
        elif len(k2) > 0:
            print("Helium section was set but is corrupted.")
            return False
        return True

    def __checkMonotony(self, arr, Invert=False):
        '''
            **Consistency check of LEDET Inputs - Monotony check **

            Check if given array is monotone or not. Returns bool of result.

            :param arr: Given array to be checked
            :type arr: np.ndarray()
            :param Invert: flag that determines which direction the array should be interpreted [True= from the last towards the first, False= from the first towards the last]
            :type Invert: bool
            :return: bool
        '''
        if Invert:
            arr = np.flip(arr)
        b = all(x <= y for x, y in zip(arr, arr[1:]))
        return b

    def __checkTimes(self):
        '''
            **Consistency check of LEDET Inputs - check Times **

            Function that checks if times in LEDET are all set accordingly, otherwise adjusts the time_vector
            :return: none
        '''

        ### Check start of time_vector
        # obtain all times from LEDET object
        try: t1 = min(self.Inputs.tQuench)
        except: t1 = min(np.array([self.Inputs.tQuench]))
        t2 = self.Inputs.t_PC_LUT[0]
        try: t3 = min(self.Inputs.tStartQuench)
        except: t3 = min(np.array([self.Inputs.tStartQuench]))
        t4 = self.Options.time_vector_params[0]
        try:  t5 = min(self.Inputs.tQH)
        except: t5 = min(np.array([self.Inputs.tQH]))
        t6 = self.Inputs.tCLIQ
        t7 = self.Inputs.tEE

        # Check if times are all after the beginning of the simulation, otherwise set the beginning to the earliest time used
        if any(x < t4 for x in [t1,t2,t3, t5, t6, t7]):
            print("You're using a time, that is before the start of the simulation. Corrected Time-Vector.")
            self.Options.time_vector_params[0] = np.min([t1, t2, t3, t5, t6, t7])-0.01
            self.Inputs.t_PC_LUT[0] = np.min([t1, t2, t3, t5, t6, t7])-0.01
            self.Inputs.tQuench = np.zeros((len(self.Inputs.tQuench),))+np.min([t1, t2, t3, t5, t6, t7])

        ### Check end of time vector
        # obtain all times from LEDET object [if times are above > 999, they are interpreted as not set]
        try: t1 = max(self.Inputs.tQuench)
        except: t1 = max(np.array([self.Inputs.tQuench]))
        if t1 >= 999: t1 = 0
        t2 = self.Inputs.t_PC_LUT[-1]
        if t2 >= 999: t2 = 0
        t3 = self.Inputs.tStartQuench
        try: t3 = [0 if x>=999 else x for x in t3]
        except:
            t3 = [t3]
            t3 = [0 if x >= 999 else x for x in t3]
        t3 = max(t3)
        t4 = self.Options.time_vector_params[-1]
        t5 = self.Inputs.tQH
        try: t5 = [0 if x>=999 else x for x in t5]
        except:
            t5 = [t5]
            t5 = t5 = [0 if x>=999 else x for x in t5]
        t5 = max(t5)
        t6 = self.Inputs.tCLIQ
        if t6 >= 999: t6 = 0
        t7 = self.Inputs.tEE
        if t7 >= 999: t7 = 0

        # Check if times are all before the end of the simulation, otherwise extend the time_vector
        if any(x > t4 for x in [t1, t2, t3, t5, t6, t7]):
            print("You're using a time, that is after the end of the simulation. Corrected Time-Vector.")
            self.Options.time_vector_params[-1] = np.max([t1, t2, t3, t5, t6, t7]) + 1
            self.Inputs.t_PC_LUT = np.append(self.Inputs.t_PC_LUT, np.max([t1, t2, t3, t5, t6, t7]) + 1)
            self.Inputs.I_PC_LUT = np.append(self.Inputs.I_PC_LUT, self.Inputs.I_PC_LUT[-1])
        return 1

    def __checkPersistentCurrents(self):
        if len(self.Inputs.df_inGroup)==1:
            if self.Options.flag_persistentCurrents == 0: return True
            else:
                print('Persistent current parameters flag is set, but parameters not. I set the flag to 0.')
                self.Options.flag_persistentCurrents = 0
                return True
        else:
            if self.Options.flag_persistentCurrents==0:
                print('Persistent current parameters are set but flag is not. Continuing.')
            maxFit = np.max(self.Inputs.selectedFit_inGroup)
            shp = self.Inputs.fitParameters_inGroup.shape
            if maxFit == 1:
                if not shp[0]>0 or not shp[1]==len(self.Inputs.nT):
                    print('You selected constant Jc, but fitParameters are not set. Abort.')
                    return False
                else: return True
            elif maxFit == 2:
                if not shp[0]>1 or not shp[1]==len(self.Inputs.nT):
                    print('You selected Botturas fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            elif maxFit == 3:
                if not shp[0] > 7 or not shp[1] == len(self.Inputs.nT):
                    print('You selected CUDI fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            elif maxFit == 4:
                if not shp[0] > 2 or not shp[1] == len(self.Inputs.nT):
                    print('You selected Summers fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            else:
                print('Unknown fit. Please check! Abort.')
                return False

    def __checkThermalConnections(self):
        if len(self.Inputs.iContactAlongHeight_From) == 0 or len(self.Inputs.iContactAlongHeight_To) == 0:
            self.Inputs.iContactAlongHeight_From = np.array([1])
            self.Inputs.iContactAlongHeight_To = np.array([1])
            print('No thermal connections in height directions set. I added at least 1.')
        if len(self.Inputs.iContactAlongWidth_From) == 0 or len(self.Inputs.iContactAlongWidth_To) == 0:
            self.Inputs.iContactAlongWidth_From = np.array([1])
            self.Inputs.iContactAlongWidth_To = np.array([1])
            print('No thermal connections in width directions set. I added at least 1.')

    def _consistencyCheckLEDET(self):
        '''
            **Consistency check of LEDET Inputs - Main function **

            Function applies different consistency checks on LEDET inputs to see if the values are set properly
            Applied checks:
                - Length checks [checking Inputs that require the same size]
                - checkM_InductanceBlock_m [checking if inductance matrix is squared]
                - checkHeFraction [checking if both Helium options are set]
                - checkMonotony [check Inputs that require monotony in themselves]
                - checkTimes [check if TimeVector fits to other times in the Inputs]

            :return Break: flag, showing if Inputs are consistent or not
            :type break: bool
        '''
        # Define groups that require the same size, number in each list contains the row-number of the attribute
        ## 0 Single - 1 CoilSections - 2 Groups - 3 Half-Turns - 4 doesn't matter - 5 iContactAlongWidth - 6 iContactAlongHeight - 7 vQlength
        ## 8 Quench Heater, 9 QH_QuenchToFrom, 10 CLIQ, 11 Persistent currents
        VarsSameInput = [['T00','l_magnet','I00','R_circuit','R_crowbar','Ud_crowbar','t_PC', 't_EE', 'R_EE_triggered', 'sim3D_uThreshold',
                          'sim3D_f_cooling_down','sim3D_f_cooling_up','sim3D_f_cooling_left','sim3D_f_cooling_right','sim3D_fExToIns',
                          'sim3D_fExUD','sim3D_fExLR','sim3D_min_ds_coarse','sim3D_min_ds_fine','sim3D_min_nodesPerStraightPart',
                          'sim3D_min_nodesPerEndsPart','sim3D_Tpulse_sPosition','sim3D_Tpulse_peakT',
                          'sim3D_Tpulse_width','sim3D_durationGIF','sim3D_flag_saveFigures','sim3D_flag_saveGIF',
                          'sim3D_flag_VisualizeGeometry3D','sim3D_flag_SaveGeometry3D'],
                         ['M_m', 'directionCurrentCLIQ', 'tQuench', 'initialQuenchTemp'],
                         ['GroupToCoilSection','polarities_inGroup','nT','nStrands_inGroup','l_mag_inGroup','ds_inGroup','f_SC_strand_inGroup',\
                         'f_ro_eff_inGroup','Lp_f_inGroup','RRR_Cu_inGroup','SCtype_inGroup','STtype_inGroup','insulationType_inGroup', \
                         'internalVoidsType_inGroup', 'externalVoidsType_inGroup', 'wBare_inGroup', 'hBare_inGroup','wIns_inGroup','hIns_inGroup',\
                         'Lp_s_inGroup', 'R_c_inGroup', 'Tc0_NbTi_ht_inGroup', 'Bc2_NbTi_ht_inGroup', 'c1_Ic_NbTi_inGroup','c2_Ic_NbTi_inGroup',\
                         'Tc0_Nb3Sn_inGroup','Bc2_Nb3Sn_inGroup','Jc_Nb3Sn0_inGroup','el_order_half_turns'],
                        ['el_order_half_turns', 'alphasDEG', 'rotation_block', 'mirror_block', 'mirrorY_block','HalfTurnToInductanceBlock'],
                        ['fL_I', 'fL_L', 'overwrite_f_internalVoids_inGroup', 'overwrite_f_externalVoids_inGroup','t_PC_LUT','I_PC_LUT', 'sim3D_idxFinerMeshHalfTurn'],
                        ['iContactAlongWidth_From', 'iContactAlongWidth_To'],
                        ['iContactAlongHeight_From', 'iContactAlongHeight_To'],
                        ['iStartQuench', 'tStartQuench', 'lengthHotSpot_iStartQuench', 'fScaling_vQ_iStartQuench'],
                        ['tQH', 'U0_QH', 'C_QH', 'R_warm_QH', 'w_QH', 'h_QH', 's_ins_QH', 'type_ins_QH', 's_ins_QH_He', 'type_ins_QH_He',\
                         'l_QH', 'f_QH'],
                        ['iQH_toHalfTurn_From','iQH_toHalfTurn_To'],
                        ['tCLIQ','nCLIQ','U0', 'C', 'Rcapa'],
                        ['df_inGroup', 'selectedFit_inGroup', 'fitParameters_inGroup']]


        slicesSameInput = []
        for i in range(len(VarsSameInput)):
            slicesSameInput.append([])

        counter = 1
        for l in self.Inputs.__annotations__:
            for i in range(len(VarsSameInput)):
                try:
                    _ = VarsSameInput[i].index(l)
                    slicesSameInput[i].append(counter-1)
                except:
                    pass
            counter = counter + 1

        # Acquire representative sizes for the defined groups
        lengthInputs = len(self.Inputs.__annotations__)
        sizeInputs = np.zeros((lengthInputs,1))
        sizeInputs[slicesSameInput[0]] = 1 #single Valued
        sizeInputs[slicesSameInput[1]] = self.__getNumberOfCoilSections() #Number of CoilSections

        if sizeInputs[slicesSameInput[1][0]] == -1:
            print("M_m or Number of Coilsections is corrupted. please check.")
            return True
        sizeInputs[slicesSameInput[2]] = len(self.Inputs.nT) #Number of Groups
        sizeInputs[slicesSameInput[3]] = sum(self.Inputs.nT) #Number of Turns
        sizeInputs[slicesSameInput[4]] = 0 #Unchecked
        sizeInputs[slicesSameInput[5]] = len(self.Inputs.iContactAlongWidth_From)
        sizeInputs[slicesSameInput[6]] = len(self.Inputs.iContactAlongHeight_From)
        sizeInputs[slicesSameInput[7]] = len(self.Inputs.iStartQuench)
        sizeInputs[slicesSameInput[8]] = len(self.Inputs.tQH)
        sizeInputs[slicesSameInput[9]] = len(self.Inputs.iQH_toHalfTurn_From)
        try:
            sizeInputs[slicesSameInput[10]] = len(self.Inputs.tCLIQ)
        except:
            sizeInputs[slicesSameInput[10]] = 1
        if len(self.Inputs.df_inGroup)>1:
            sizeInputs[slicesSameInput[11]] = len(self.Inputs.nT)
        else:
            sizeInputs[slicesSameInput[11]] = len(self.Inputs.df_inGroup)

        # Checks for types and lengths/sizes
        Count = 0
        Break = 0
        for k in self.Inputs.__annotations__:
            if sizeInputs[Count] == 0:
                Count = Count + 1
                continue
            cC = self.getAttribute(self.Inputs, k)
            if type(cC) == list:
                if not len(cC)==sizeInputs[Count]:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                    Break = 1
            elif type(cC) == np.ndarray:
                if len(cC.shape)==1:
                    if not len(cC)==sizeInputs[Count]:
                        print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                        Break = 1
                else:
                    if not cC.shape[1]==sizeInputs[Count]:
                        print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",cC.shape[1],"! Please check.")
                        Break = 1
            elif type(cC) == float or type(cC) == int or type(cC) == np.float64:
                if not sizeInputs[Count]== 1:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is", type(cC),". Please check.")
                    Break = 1
            else:
                print("Variable ", k, " has the wrong data-type set! Please check.")
                Break = 1
            Count = Count + 1

        ## Remaining checks in functions
        if not self.__checkHeFraction(len(self.Inputs.nT)):
            Break = 1
        if not self.__checkMonotony(self.Inputs.t_PC_LUT):
            print("t_PC_LUT is not monotonic")
            Break = 1
        if not self.__checkMonotony(self.Inputs.fL_I):
            print("fL_I is not monotonic")
            Break = 1
        if not self.__checkTimes():
            Break = 1
        if not self.__checkPersistentCurrents():
            Break = 1
        if not self.__checkM_InductanceBlock_m(int(sum(self.Inputs.nT)/2)):
            Break = 1
        return Break

# END - Methods for consistency checks of LEDET input files #
###############################################################
# START - Helper functions #

def CompareLEDETParameters(FileA, FileB, Precision=1E-5, showIndices=0):
    '''
        Compare all the variables imported from two LEDET Excel input files
    '''

    Diff = 0

    pl_a = ParserLEDET(BuilderLEDET(flag_build=False))
    pl_a.readFromExcel(FileA, verbose=False)
    pl_b = ParserLEDET(BuilderLEDET(flag_build=False))
    pl_b.readFromExcel(FileB, verbose=False)
    print("Starting Comparison of A: ({}) and B: ({})".format(FileA, FileB))

    ## Check Inputs
    for attribute in pl_a.builder_ledet.Inputs.__annotations__:
        Diff = _compare_two_parameters(pl_a.builder_ledet.getAttribute("Inputs", attribute),
                                       pl_b.builder_ledet.getAttribute("Inputs", attribute),
                                       Diff, attribute, Precision, showIndices)

    ## Check Options
    for attribute in pl_a.builder_ledet.Options.__annotations__:
        Diff = _compare_two_parameters(pl_a.builder_ledet.getAttribute("Options", attribute),
                                       pl_b.builder_ledet.getAttribute("Options", attribute),
                                       Diff, attribute, Precision, showIndices)

    ## Check Plots
    for attribute in pl_a.builder_ledet.Plots.__annotations__:
        Diff = _compare_two_parameters(pl_a.builder_ledet.getAttribute("Plots", attribute),
                                       pl_b.builder_ledet.getAttribute("Plots", attribute),
                                       Diff, attribute, Precision, showIndices)

    ## Check Variables
    for attribute in pl_a.builder_ledet.Variables.__annotations__:
        Diff = _compare_two_parameters(pl_a.builder_ledet.getAttribute("Variables", attribute),
                                       pl_b.builder_ledet.getAttribute("Variables", attribute),
                                       Diff, attribute, Precision, showIndices)

    if Diff == 0:
        print("Files {} and {} are equal.".format(FileA,FileB))
        return True
    else:
        return False


def _compare_two_parameters(var_1, var_2, Diff, attribute, Precision, showIndices):
    '''
        Helper-function for BuilderLEDET() to compare to parameters
        max_relative_error is a relative value
    '''
    Block = 1
    CC = 0

    if isinstance(var_1, float) or isinstance(var_1, int):
        if abs(var_1 - var_2) > Precision * abs(var_1):
            if Block: print(
                "Found difference in scalar Parameter {}, A: {}, B: {}".format(attribute, var_1, var_2))
            Block = 0
            Diff = True

    elif var_1 is None or var_2 is None:
        if var_1 is not None or var_2 is not None:
            Diff = 1
            print('Parameter A is {} while B is {}'.format(var_1, var_2))

    elif len(var_1) != len(var_2):
        Diff = 1
        if Block:
            Block = 0
            print('Parameter {} of A, {} has not the same length as Parameter of B, {}'.format(attribute,
                                                                                               len(var_1),
                                                                                               len(var_2)))
    else:
        Pos = []
        for k in range(len(var_1)):
            try:
                if var_1[k] != var_2[k]:
                    if isinstance(var_1[k], str):
                        if var_1[k] != var_2[k]:
                            Diff = 1
                            if Block:
                                print("Found difference in vector Parameter {}".format(attribute))
                                Block = 0
                            Pos.append(k)

                    elif abs(var_1[k] - var_2[k]) > Precision * abs(var_1[k]):
                        Diff = 1
                        if Block:
                            print("Found difference in vector Parameter {}".format(attribute))
                            Block = 0
                        Pos.append(k)
            except:
                for j in range(var_1.shape[1]):
                    if var_1[k, j] != var_2[k, j]:
                        if isinstance(var_1[k], str):
                            Diff = 1
                            if Block:
                                print("Found difference in matrix Parameter {}".format(attribute))
                                Block = 0
                            Pos.append([k, j])
                        elif abs(var_1[k, j] - var_2[k, j]) > Precision * abs(var_1[k, j]):
                            Diff = 1
                            if Block:
                                print("Found difference in matrix Parameter {}".format(attribute))
                                Block = 0
                            Pos.append([k, j])

        if len(Pos) > 0:
            if len(Pos) < 10:
                print("Different Positions: {}".format(Pos))
            else:
                print("Many values are different (>10)")
                if showIndices: print(Pos)

    return Diff

# END - Helper functions #
###############################################################
